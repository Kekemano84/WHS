import os
import requests
import secrets
import json
import sqlite3
import calendar
import uuid
import smtplib
import ssl
import zipfile
from datetime import datetime, timedelta
from functools import wraps
from email.message import EmailMessage
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for, send_file, flash, session, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import stripe
except Exception:
    stripe = None

try:
    import requests
except Exception:
    requests = None


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "whs-dev-secret")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DB_PATH = os.path.join(DATA_DIR, "whs.db")

def get_hmrc_rate():
    try:
        html=requests.get("https://www.gov.uk/government/publications/rates-and-allowances-travel-mileage-and-fuel-allowances/travel-mileage-and-fuel-rates-and-allowances",timeout=10).text
        if "55p" in html:
            return 0.55
        if "45p" in html:
            return 0.45
    except Exception:
        pass
    return 0.55


HMRC_MILE_RATE = get_hmrc_rate()
PERSONAL_ALLOWANCE = 12570
BASIC_RATE_LIMIT = 50270
INCOME_TAX_BASIC = 0.20
INCOME_TAX_HIGHER = 0.40
CLASS2_NI_WEEKLY = 3.45
CLASS4_NI_LOWER = 12570
CLASS4_NI_UPPER = 50270
CLASS4_NI_BASIC = 0.06
CLASS4_NI_HIGHER = 0.02

PLAN_ORDER = {"free": 0, "pro": 1, "business": 2}
PLAN_NAMES = {"free": "Free", "pro": "Free", "business": "Free"}

def display_plan_name(user_row):
    try:
        if is_admin(user_row):
            return "Admin"
    except Exception:
        pass
    return PLAN_NAMES.get(row_get(user_row, "plan", "free"), "Free")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
STRIPE_SECRET_KEY = os.environ.get("STRIPE_SECRET_KEY")
STRIPE_PRO_PRICE_ID = os.environ.get("STRIPE_PRO_PRICE_ID")
STRIPE_BUSINESS_PRICE_ID = os.environ.get("STRIPE_BUSINESS_PRICE_ID")
APP_BASE_URL = os.environ.get("APP_BASE_URL", "http://127.0.0.1:5000")
# Email / SMTP configuration.
# Supports both WHS-style variables and common Flask-Mail/Gmail variable names.
SMTP_HOST = (
    os.environ.get("SMTP_HOST")
    or os.environ.get("SMTP_SERVER")
    or os.environ.get("MAIL_SERVER")
    or os.environ.get("EMAIL_HOST")
    or "smtp.gmail.com"
)
SMTP_PORT = int(
    os.environ.get("SMTP_PORT")
    or os.environ.get("MAIL_PORT")
    or os.environ.get("EMAIL_PORT")
    or "587"
)
SMTP_USER = (
    os.environ.get("SMTP_USER")
    or os.environ.get("SMTP_EMAIL")
    or os.environ.get("MAIL_USERNAME")
    or os.environ.get("EMAIL_HOST_USER")
)
SMTP_PASSWORD = (
    os.environ.get("SMTP_PASSWORD")
    or os.environ.get("SMTP_APP_PASSWORD")
    or os.environ.get("MAIL_PASSWORD")
    or os.environ.get("EMAIL_HOST_PASSWORD")
)
SMTP_FROM = (
    os.environ.get("SMTP_FROM")
    or os.environ.get("MAIL_DEFAULT_SENDER")
    or os.environ.get("DEFAULT_FROM_EMAIL")
    or SMTP_USER
    or ""
)
SMTP_USE_TLS = str(os.environ.get("SMTP_USE_TLS") or os.environ.get("MAIL_USE_TLS") or "true").lower() in ("1", "true", "yes", "on")
SMTP_USE_SSL = str(os.environ.get("SMTP_USE_SSL") or os.environ.get("MAIL_USE_SSL") or "false").lower() in ("1", "true", "yes", "on")
GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY")


def get_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn




def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "error")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def safe_add_column(table, column, definition):
    conn = get_db()
    try:
        existing = [r["name"] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        if column not in existing:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
            conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            plan TEXT NOT NULL DEFAULT 'free',
            business_name TEXT,
            phone TEXT,
            address TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS mileage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            start_location TEXT NOT NULL,
            end_location TEXT NOT NULL,
            miles REAL NOT NULL,
            purpose TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT NOT NULL,
            amount REAL NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            invoice_number TEXT NOT NULL,
            date TEXT NOT NULL,
            customer_name TEXT NOT NULL,
            customer_email TEXT,
            description TEXT NOT NULL,
            amount REAL NOT NULL,
            status TEXT NOT NULL DEFAULT 'Unpaid',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS yard_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            trailer_id TEXT NOT NULL,
            location_type TEXT NOT NULL,
            location_detail TEXT,
            status TEXT NOT NULL DEFAULT 'Recorded',
            notes TEXT,
            source TEXT NOT NULL DEFAULT 'Manual',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS kpi_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            volume INTEGER DEFAULT 0,
            planned_hc INTEGER DEFAULT 0,
            actual_hc INTEGER DEFAULT 0,
            target_rate REAL DEFAULT 0,
            actual_rate REAL DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            errors INTEGER DEFAULT 0,
            notes TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS handovers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            manager TEXT,
            volume INTEGER DEFAULT 0,
            planned_hc INTEGER DEFAULT 0,
            actual_hc INTEGER DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            issues TEXT,
            actions TEXT,
            generated_report TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL,
            email TEXT,
            status TEXT NOT NULL DEFAULT 'Active',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS custom_locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            prefix TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS shift_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            volume INTEGER DEFAULT 0,
            available_hc INTEGER DEFAULT 0,
            target_rate REAL DEFAULT 0,
            planned_hours REAL DEFAULT 0,
            ai_plan TEXT,
            created_at TEXT NOT NULL
        )
    """)


    cur.execute("""
        CREATE TABLE IF NOT EXISTS shift_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Off',
            shift_name TEXT,
            start_time TEXT,
            end_time TEXT,
            notes TEXT,
            source TEXT DEFAULT 'Generated',
            created_at TEXT NOT NULL,
            UNIQUE(user_id, date)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS morning_briefs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            manager TEXT,
            role TEXT,
            volume INTEGER DEFAULT 0,
            available_hc INTEGER DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            safety_message TEXT,
            priorities TEXT,
            team_messages TEXT,
            break_reminder TEXT,
            equipment_reminder TEXT,
            generated_brief TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS photo_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            image_filename TEXT,
            trailer_id TEXT,
            location_detail TEXT,
            damage_notes TEXT,
            recognition_notes TEXT,
            confidence TEXT,
            created_at TEXT NOT NULL
        )
    """)


    cur.execute("""
        CREATE TABLE IF NOT EXISTS remember_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            user_agent TEXT
        )
    """)



    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_shift_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT,
            manager TEXT,
            volume INTEGER DEFAULT 0,
            planned_hc INTEGER DEFAULT 0,
            actual_hc INTEGER DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            safety TEXT,
            issues TEXT,
            actions TEXT,
            notes TEXT,
            photo_filename TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS action_tracker (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            title TEXT NOT NULL,
            owner TEXT,
            due_date TEXT,
            status TEXT NOT NULL DEFAULT 'Open',
            priority TEXT DEFAULT 'Medium',
            source TEXT,
            notes TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS absence_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            team_member_id INTEGER,
            member_name TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT,
            absence_type TEXT NOT NULL DEFAULT 'Sick',
            notes TEXT,
            created_at TEXT NOT NULL
        )
    """)


    cur.execute("""
        CREATE TABLE IF NOT EXISTS sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            address TEXT,
            target_score REAL DEFAULT 90,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS probation_reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            team_member_id INTEGER,
            member_name TEXT NOT NULL,
            review_date TEXT NOT NULL,
            outcome TEXT NOT NULL,
            manager TEXT,
            strengths TEXT,
            concerns TEXT,
            next_steps TEXT,
            generated_review TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS handover_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL DEFAULT 'Default Handover',
            section_names TEXT,
            force_vrid_uppercase INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS handover_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            handover_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            title TEXT,
            message_date TEXT,
            shift_time TEXT,
            recipient_name TEXT,
            phone_number TEXT,
            rows_json TEXT,
            generated_message TEXT,
            updated_at TEXT NOT NULL
        )
    """)
    conn.commit()

    safe_add_column("handovers", "extra_json", "TEXT")
    safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")
    safe_add_column("users", "language", "TEXT DEFAULT 'en'")
    safe_add_column("users", "favorite_tools", "TEXT DEFAULT 'morning_brief,shift_calendar,handover,yard_check'")
    safe_add_column("users", "company_theme", "TEXT DEFAULT 'whs'")
    safe_add_column("users", "show_theme_label", "INTEGER DEFAULT 1")
    safe_add_column("users", "last_login_at", "TEXT")
    safe_add_column("users", "inactive_warning_at", "TEXT")
    safe_add_column("team_members", "licence_expiry", "TEXT")
    safe_add_column("team_members", "training_expiry", "TEXT")
    safe_add_column("team_members", "training_type", "TEXT")
    safe_add_column("photo_records", "category", "TEXT DEFAULT 'Evidence'")
    safe_add_column("photo_records", "comment", "TEXT")
    safe_add_column("yard_checks", "photo_filename", "TEXT")

    conn.close()



def ensure_schema_updates():
    conn = get_db()
    cur = conn.cursor()

    additions = {
        "users": [
            ("role", "TEXT NOT NULL DEFAULT 'Admin'"),
            ("company_name", "TEXT"),
            ("stripe_customer_id", "TEXT"),
            ("stripe_subscription_id", "TEXT"),
            ("subscription_status", "TEXT DEFAULT 'manual'"),
            ("pro_expires_at", "TEXT"),
            ("pro_reason", "TEXT"),
            ("door_count", "INTEGER DEFAULT 100"),
            ("fence_count", "INTEGER DEFAULT 120"),
            ("door_start", "INTEGER DEFAULT 1"),
            ("door_end", "INTEGER DEFAULT 100"),
            ("fence_start", "INTEGER DEFAULT 1"),
            ("fence_end", "INTEGER DEFAULT 120"),
            ("language", "TEXT DEFAULT 'en'"),
            ("last_login_at", "TEXT"),
            ("inactive_warning_at", "TEXT"),
            ("business_logo_filename", "TEXT"),
            ("brand_color", "TEXT DEFAULT '#2563eb'"),
            ("default_site_id", "INTEGER"),
            ("company_theme", "TEXT DEFAULT 'whs'"),
            ("show_theme_label", "INTEGER DEFAULT 1")
        ],
        "team_members": [
            ("permissions", "TEXT DEFAULT 'View only'"),
            ("phone", "TEXT"),
            ("notes", "TEXT"),
            ("probation_start", "TEXT"),
            ("probation_end", "TEXT"),
            ("probation_status", "TEXT DEFAULT 'Not set'"),
            ("licence_expiry", "TEXT"),
            ("training_expiry", "TEXT"),
            ("training_type", "TEXT")
        ],
        "invoices": [
            ("email_sent", "INTEGER DEFAULT 0")
        ],
        "handovers": [
            ("pdf_created", "INTEGER DEFAULT 0"),
            ("extra_json", "TEXT")
        ],
        "photo_records": [
            ("ai_result", "TEXT")
        ],
        "daily_shift_logs": [
            ("site_id", "INTEGER")
        ],
        "action_tracker": [
            ("site_id", "INTEGER")
        ],
        "yard_checks": [
            ("site_id", "INTEGER")
        ],
        "handovers": [
            ("site_id", "INTEGER")
        ]
    }

    cur.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            title TEXT,
            message_date TEXT,
            shift_time TEXT,
            recipient_name TEXT,
            phone_number TEXT,
            rows_json TEXT,
            generated_message TEXT,
            updated_at TEXT NOT NULL
        )
    """)

    for table, cols in additions.items():
        existing = [row["name"] for row in cur.execute(f"PRAGMA table_info({table})").fetchall()]
        for col_name, col_def in cols:
            if col_name not in existing:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_def}")

    conn.commit()
    conn.close()




def seed_admin_user():
    conn = get_db()

    # Always make sure new columns exist before touching admin.
    existing_columns = [row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()]

    def add_col(name, sql_type):
        nonlocal existing_columns
        if name not in existing_columns:
            conn.execute(f"ALTER TABLE users ADD COLUMN {name} {sql_type}")
            existing_columns.append(name)

    add_col("role", "TEXT NOT NULL DEFAULT 'Admin'")
    add_col("company_name", "TEXT")
    add_col("stripe_customer_id", "TEXT")
    add_col("stripe_subscription_id", "TEXT")
    add_col("subscription_status", "TEXT DEFAULT 'manual'")
    add_col("mileage_rate", "REAL DEFAULT 0.55")
    add_col("door_count", "INTEGER DEFAULT 100")
    add_col("fence_count", "INTEGER DEFAULT 120")
    add_col("door_start", "INTEGER DEFAULT 1")
    add_col("door_end", "INTEGER DEFAULT 100")
    add_col("fence_start", "INTEGER DEFAULT 1")
    add_col("fence_end", "INTEGER DEFAULT 120")
    add_col("language", "TEXT DEFAULT 'en'")
    add_col("favorite_tools", "TEXT DEFAULT 'morning_brief,shift_calendar,handover,yard_check'")
    add_col("last_login_at", "TEXT")
    add_col("inactive_warning_at", "TEXT")
    add_col("password_changed_at", "TEXT")
    add_col("default_site_id", "INTEGER")
    add_col("brand_color", "TEXT DEFAULT '#f59e0b'")
    add_col("company_theme", "TEXT DEFAULT 'whs'")
    add_col("business_logo_filename", "TEXT")
    add_col("pro_expires_at", "TEXT")
    add_col("pro_reason", "TEXT")
    conn.commit()

    existing = conn.execute("SELECT id FROM users WHERE email = ?", ("admin@whs-app.com",)).fetchone()

    if not existing:
        conn.execute("""
            INSERT INTO users
            (name, email, password_hash, plan, business_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            "WHS Admin",
            "admin@whs-app.com",
            generate_password_hash(os.getenv("WHS_ADMIN_PASSWORD", "WHSAdmin2026!")),
            "pro",
            "WHS",
            datetime.now().isoformat()
        ))
        conn.commit()

    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            role = 'Admin',
            password_hash = CASE
                WHEN password_changed_at IS NULL OR password_changed_at = '' THEN ?
                ELSE password_hash
            END,
            company_name = 'WHS Admin',
            subscription_status = 'admin',
            mileage_rate = 0.55,
            door_count = 100,
            door_start = 1,
            door_end = 100,
            fence_count = 120,
            fence_start = 1,
            fence_end = 120
        WHERE email = ?
    """, (generate_password_hash(os.getenv("WHS_ADMIN_PASSWORD", "WHSAdmin2026!")), "admin@whs-app.com"))
    conn.commit()
    conn.close()

def is_admin(user=None):
    user = user or current_user()
    return bool(user and (row_get(user, "email") == "admin@whs-app.com" or row_get(user, "role") == "Admin"))


def current_user():
    if "user_id" not in session:
        return None
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    conn.close()
    return refresh_user_plan(user)


TRANSLATIONS = {
    "en": {
        "dashboard":"Dashboard","home":"Home","calendar":"Calendar","team":"Team","more":"More","warehouse_support":"Warehouse Support","notifications":"Notifications",
        "morning_brief":"Morning Brief","mileage":"Mileage","expenses":"Expenses","yard_check":"Yard Check","yard_settings":"Yard Settings","handover":"Handover",
        "daily_log":"Daily Log","actions":"Actions","absence":"Absence","evidence":"Evidence","search":"Search","plans":"Plans","settings":"Settings","logout":"Logout",
        "checks":"Checks","hmrc_mileage":"HMRC mileage","rate_55p":"55p per business mile","tax_estimates":"Tax calculations are estimates only."
    },
    "hu": {
        "dashboard":"Vezérlőpult","home":"Főoldal","calendar":"Naptár","team":"Csapat","more":"Több","warehouse_support":"Warehouse Support","notifications":"Értesítések",
        "morning_brief":"Reggeli brief","mileage":"Mérföldek","expenses":"Kiadások","yard_check":"Yard ellenőrzés","yard_settings":"Yard beállítások","handover":"Átadás",
        "daily_log":"Napi műszaknapló","actions":"Feladatok","absence":"Hiányzás","evidence":"Bizonyítékok","search":"Keresés","plans":"Csomagok","settings":"Beállítások","logout":"Kijelentkezés",
        "checks":"Ellenőrzések","hmrc_mileage":"HMRC mérföld","rate_55p":"55p üzleti mérföldenként","tax_estimates":"Az adószámítások csak becslések."
    },
    "pl": {
        "dashboard":"Panel","home":"Start","calendar":"Kalendarz","team":"Zespół","more":"Więcej","warehouse_support":"Warehouse Support","notifications":"Powiadomienia",
        "morning_brief":"Poranny briefing","mileage":"Kilometrówka","expenses":"Wydatki","yard_check":"Kontrola placu","yard_settings":"Ustawienia placu","handover":"Przekazanie zmiany",
        "daily_log":"Dziennik zmiany","actions":"Zadania","absence":"Nieobecności","evidence":"Dowody","search":"Szukaj","plans":"Plany","settings":"Ustawienia","logout":"Wyloguj",
        "checks":"Kontrole","hmrc_mileage":"Stawka HMRC","rate_55p":"55p za milę służbową","tax_estimates":"Wyliczenia podatkowe są szacunkowe."
    },
    "ro": {
        "dashboard":"Panou","home":"Acasă","calendar":"Calendar","team":"Echipă","more":"Mai mult","warehouse_support":"Warehouse Support","notifications":"Notificări",
        "morning_brief":"Briefing de dimineață","mileage":"Mileaj","expenses":"Cheltuieli","yard_check":"Verificare curte","yard_settings":"Setări curte","handover":"Predare tură",
        "daily_log":"Jurnal zilnic tură","actions":"Acțiuni","absence":"Absențe","evidence":"Dovezi","search":"Căutare","plans":"Planuri","settings":"Setări","logout":"Deconectare",
        "checks":"Verificări","hmrc_mileage":"Mileaj HMRC","rate_55p":"55p pe milă de serviciu","tax_estimates":"Calculele fiscale sunt estimări."
    },
    "es": {
        "dashboard":"Panel","home":"Inicio","calendar":"Calendario","team":"Equipo","more":"Más","warehouse_support":"Warehouse Support","notifications":"Notificaciones",
        "morning_brief":"Briefing de mañana","mileage":"Millas","expenses":"Gastos","yard_check":"Revisión de patio","yard_settings":"Ajustes de patio","handover":"Traspaso de turno",
        "daily_log":"Registro diario","actions":"Acciones","absence":"Ausencias","evidence":"Evidencias","search":"Buscar","plans":"Planes","settings":"Ajustes","logout":"Cerrar sesión",
        "checks":"Revisiones","hmrc_mileage":"Millas HMRC","rate_55p":"55p por milla de negocio","tax_estimates":"Los cálculos fiscales son estimaciones."
    },
    "de": {
        "dashboard":"Dashboard","home":"Home","calendar":"Kalender","team":"Team","more":"Mehr","warehouse_support":"Warehouse Support","notifications":"Benachrichtigungen",
        "morning_brief":"Morgenbriefing","mileage":"Fahrten","expenses":"Ausgaben","yard_check":"Hofprüfung","yard_settings":"Hof-Einstellungen","handover":"Schichtübergabe",
        "daily_log":"Tägliches Schichtlog","actions":"Aufgaben","absence":"Abwesenheit","evidence":"Nachweise","search":"Suche","plans":"Pläne","settings":"Einstellungen","logout":"Abmelden",
        "checks":"Prüfungen","hmrc_mileage":"HMRC-Fahrtkosten","rate_55p":"55p pro Geschäftsmeile","tax_estimates":"Steuerberechnungen sind Schätzungen."
    },
}


EXTRA_TRANSLATIONS = {
    "en": {"operations":"Operations","language":"Language","business_details":"Business Details","business_name":"Business Name","company_site_name":"Company / Site Name","your_name":"Your Name","your_role":"Your Role","phone":"Phone","address":"Address","default_mileage_rate":"Default Mileage Rate (£/mile)","door_count":"Door Count","fence_count":"Fence Count","save_settings":"Save Settings","favorites":"Favorites","dashboard_favorites":"Dashboard Favorites","dashboard_favorites_help":"Choose the four quick buttons shown on your mobile dashboard.","favorite":"Favorite","welcome_back":"Welcome back","dashboard_intro":"Operations dashboard for shifts, checks, mileage, expenses and handovers.","today":"Today","no_shift_time":"No shift time set","more_intro":"Warehouse Support tools and settings in one place.","holiday_settings":"Holiday Settings"},
    "hu": {"operations":"Műveletek","language":"Nyelv","business_details":"Céges adatok","business_name":"Cég neve","company_site_name":"Cég / telephely neve","your_name":"Neved","your_role":"Szerepköröd","phone":"Telefon","address":"Cím","default_mileage_rate":"Alap mérföld díj (£/mile)","door_count":"Door darabszám","fence_count":"Fence darabszám","save_settings":"Beállítások mentése","favorites":"Kedvencek","dashboard_favorites":"Vezérlőpult kedvencek","dashboard_favorites_help":"Válaszd ki a négy gyorsgombot, ami a mobil vezérlőpulton megjelenik.","favorite":"Kedvenc","welcome_back":"Üdv újra","dashboard_intro":"Műszakok, ellenőrzések, mérföldek, kiadások és átadások egy helyen.","today":"Ma","no_shift_time":"Nincs műszakidő beállítva","more_intro":"Warehouse Support eszközök és beállítások egy helyen.","holiday_settings":"Szabadság beállítások"},
    "pl": {"operations":"Operacje","language":"Język","business_details":"Dane firmy","business_name":"Nazwa firmy","company_site_name":"Firma / lokalizacja","your_name":"Twoje imię","your_role":"Twoja rola","phone":"Telefon","address":"Adres","default_mileage_rate":"Domyślna stawka za milę (£/mile)","door_count":"Liczba bram","fence_count":"Liczba płotów","save_settings":"Zapisz ustawienia","favorites":"Ulubione","dashboard_favorites":"Ulubione pulpitu","dashboard_favorites_help":"Wybierz cztery szybkie przyciski na pulpicie mobilnym.","favorite":"Ulubione","welcome_back":"Witaj ponownie","dashboard_intro":"Panel operacyjny dla zmian, kontroli, kilometrówki, wydatków i przekazań.","today":"Dzisiaj","no_shift_time":"Brak ustawionego czasu zmiany","more_intro":"Narzędzia Warehouse Support i ustawienia w jednym miejscu.","holiday_settings":"Ustawienia urlopu"},
    "ro": {"operations":"Operațiuni","language":"Limbă","business_details":"Detalii firmă","business_name":"Nume firmă","company_site_name":"Firmă / locație","your_name":"Numele tău","your_role":"Rolul tău","phone":"Telefon","address":"Adresă","default_mileage_rate":"Rată implicită mile (£/mile)","door_count":"Număr uși","fence_count":"Număr garduri","save_settings":"Salvează setările","favorites":"Favorite","dashboard_favorites":"Favorite panou","dashboard_favorites_help":"Alege cele patru butoane rapide afișate pe panoul mobil.","favorite":"Favorit","welcome_back":"Bine ai revenit","dashboard_intro":"Panou pentru ture, verificări, mileaj, cheltuieli și predări.","today":"Astăzi","no_shift_time":"Nu este setat timpul turei","more_intro":"Instrumente Warehouse Support și setări într-un singur loc.","holiday_settings":"Setări concediu"},
    "es": {"operations":"Operaciones","language":"Idioma","business_details":"Datos de empresa","business_name":"Nombre de empresa","company_site_name":"Empresa / sitio","your_name":"Tu nombre","your_role":"Tu rol","phone":"Teléfono","address":"Dirección","default_mileage_rate":"Tarifa por milla predeterminada (£/mile)","door_count":"Número de puertas","fence_count":"Número de vallas","save_settings":"Guardar ajustes","favorites":"Favoritos","dashboard_favorites":"Favoritos del panel","dashboard_favorites_help":"Elige los cuatro botones rápidos del panel móvil.","favorite":"Favorito","welcome_back":"Bienvenido de nuevo","dashboard_intro":"Panel operativo para turnos, revisiones, millas, gastos y traspasos.","today":"Hoy","no_shift_time":"Sin horario de turno","more_intro":"Herramientas Warehouse Support y ajustes en un solo lugar.","holiday_settings":"Ajustes de vacaciones"},
    "de": {"operations":"Betrieb","language":"Sprache","business_details":"Firmendaten","business_name":"Firmenname","company_site_name":"Firma / Standort","your_name":"Dein Name","your_role":"Deine Rolle","phone":"Telefon","address":"Adresse","default_mileage_rate":"Standard-Meilensatz (£/mile)","door_count":"Anzahl Türen","fence_count":"Anzahl Zäune","save_settings":"Einstellungen speichern","favorites":"Favoriten","dashboard_favorites":"Dashboard-Favoriten","dashboard_favorites_help":"Wähle die vier Schnellzugriffe für dein mobiles Dashboard.","favorite":"Favorit","welcome_back":"Willkommen zurück","dashboard_intro":"Operations-Dashboard für Schichten, Checks, Fahrten, Ausgaben und Übergaben.","today":"Heute","no_shift_time":"Keine Schichtzeit gesetzt","more_intro":"Warehouse Support Tools und Einstellungen an einem Ort.","holiday_settings":"Urlaubseinstellungen"},
}

AUTH_TRANSLATIONS = {
    "en": {
        "login_hero_text": "Mileage, expenses, shift notes, yard checks, handovers and team records in one clean manager tool.",
        "excel_export": "Excel export", "download_reports_fast": "Download reports fast",
        "pdf_handover": "PDF handover", "share_clean_shift_reports": "Share clean shift reports",
        "calendar_notes": "Calendar notes", "calendar_notes_desc": "Meetings, probation dates, reminders",
        "yard_checks": "Yard checks", "yard_checks_desc": "Doors, fences and custom locations",
        "installable_app": "Installable app", "download_whs_app": "Download WHS App",
        "install_app_desc": "Install on phone, tablet or desktop.", "download_app": "Download App",
        "create_account": "Create account", "name": "Name", "email": "Email", "password": "Password",
        "password_placeholder": "At least 6 characters", "already_have_account": "Already have an account?",
        "login": "Login", "remember_device": "Remember this device for 30 days", "no_account_yet": "No account yet?", "create_one": "Create one"
    },
    "hu": {
        "login_hero_text": "Mérföldek, kiadások, műszakjegyzetek, yard ellenőrzések, átadások és csapatadatok egy tiszta manager eszközben.",
        "excel_export": "Excel export", "download_reports_fast": "Riportok gyors letöltése",
        "pdf_handover": "PDF átadás", "share_clean_shift_reports": "Tiszta műszakriportok megosztása",
        "calendar_notes": "Naptár jegyzetek", "calendar_notes_desc": "Meetingek, próbaidő dátumok, emlékeztetők",
        "yard_checks": "Yard ellenőrzések", "yard_checks_desc": "Doorok, fence-ek és egyedi helyek",
        "installable_app": "Telepíthető app", "download_whs_app": "WHS app letöltése",
        "install_app_desc": "Telepítés telefonra, tabletre vagy desktopra.", "download_app": "App letöltése",
        "create_account": "Fiók létrehozása", "name": "Név", "email": "Email", "password": "Jelszó",
        "password_placeholder": "Legalább 6 karakter", "already_have_account": "Már van fiókod?",
        "login": "Bejelentkezés", "remember_device": "Emlékezzen erre az eszközre 30 napig", "no_account_yet": "Nincs még fiókod?", "create_one": "Hozz létre egyet"
    },
    "pl": {
        "login_hero_text": "Kilometrówka, wydatki, notatki ze zmiany, kontrole placu, przekazania i dane zespołu w jednym prostym narzędziu dla managera.",
        "excel_export": "Eksport Excel", "download_reports_fast": "Szybkie pobieranie raportów",
        "pdf_handover": "Przekazanie PDF", "share_clean_shift_reports": "Udostępniaj czytelne raporty zmianowe",
        "calendar_notes": "Notatki kalendarza", "calendar_notes_desc": "Spotkania, daty okresu próbnego, przypomnienia",
        "yard_checks": "Kontrole placu", "yard_checks_desc": "Bramy, ogrodzenia i własne lokalizacje",
        "installable_app": "Aplikacja do instalacji", "download_whs_app": "Pobierz aplikację WHS",
        "install_app_desc": "Zainstaluj na telefonie, tablecie lub komputerze.", "download_app": "Pobierz aplikację",
        "create_account": "Utwórz konto", "name": "Imię i nazwisko", "email": "Email", "password": "Hasło",
        "password_placeholder": "Minimum 6 znaków", "already_have_account": "Masz już konto?",
        "login": "Zaloguj", "remember_device": "Zapamiętaj to urządzenie przez 30 dni", "no_account_yet": "Nie masz konta?", "create_one": "Utwórz konto"
    },
    "ro": {
        "login_hero_text": "Mileaj, cheltuieli, note de tură, verificări curte, predări și evidențe de echipă într-un instrument curat pentru manageri.",
        "excel_export": "Export Excel", "download_reports_fast": "Descarcă rapid rapoarte",
        "pdf_handover": "Predare PDF", "share_clean_shift_reports": "Partajează rapoarte clare de tură",
        "calendar_notes": "Notițe calendar", "calendar_notes_desc": "Întâlniri, date de probă, memento-uri",
        "yard_checks": "Verificări curte", "yard_checks_desc": "Uși, garduri și locații personalizate",
        "installable_app": "Aplicație instalabilă", "download_whs_app": "Descarcă aplicația WHS",
        "install_app_desc": "Instalează pe telefon, tabletă sau desktop.", "download_app": "Descarcă aplicația",
        "create_account": "Creează cont", "name": "Nume", "email": "Email", "password": "Parolă",
        "password_placeholder": "Cel puțin 6 caractere", "already_have_account": "Ai deja cont?",
        "login": "Autentificare", "remember_device": "Ține minte acest dispozitiv 30 de zile", "no_account_yet": "Nu ai cont?", "create_one": "Creează unul"
    },
    "es": {
        "login_hero_text": "Kilometraje, gastos, notas de turno, revisiones de patio, entregas y registros de equipo en una herramienta limpia para managers.",
        "excel_export": "Exportar Excel", "download_reports_fast": "Descarga informes rápido",
        "pdf_handover": "Entrega PDF", "share_clean_shift_reports": "Comparte informes de turno claros",
        "calendar_notes": "Notas de calendario", "calendar_notes_desc": "Reuniones, fechas de prueba, recordatorios",
        "yard_checks": "Revisiones de patio", "yard_checks_desc": "Puertas, vallas y ubicaciones personalizadas",
        "installable_app": "Aplicación instalable", "download_whs_app": "Descargar app WHS",
        "install_app_desc": "Instala en móvil, tablet o escritorio.", "download_app": "Descargar app",
        "create_account": "Crear cuenta", "name": "Nombre", "email": "Email", "password": "Contraseña",
        "password_placeholder": "Al menos 6 caracteres", "already_have_account": "¿Ya tienes cuenta?",
        "login": "Iniciar sesión", "remember_device": "Recordar este dispositivo durante 30 días", "no_account_yet": "¿No tienes cuenta?", "create_one": "Crea una"
    },
    "de": {
        "login_hero_text": "Fahrten, Ausgaben, Schichtnotizen, Hofprüfungen, Übergaben und Teamdaten in einem klaren Manager-Tool.",
        "excel_export": "Excel-Export", "download_reports_fast": "Berichte schnell herunterladen",
        "pdf_handover": "PDF-Übergabe", "share_clean_shift_reports": "Klare Schichtberichte teilen",
        "calendar_notes": "Kalendernotizen", "calendar_notes_desc": "Meetings, Probezeitdaten, Erinnerungen",
        "yard_checks": "Hofprüfungen", "yard_checks_desc": "Tore, Zäune und eigene Standorte",
        "installable_app": "Installierbare App", "download_whs_app": "WHS App herunterladen",
        "install_app_desc": "Auf Telefon, Tablet oder Desktop installieren.", "download_app": "App herunterladen",
        "create_account": "Konto erstellen", "name": "Name", "email": "E-Mail", "password": "Passwort",
        "password_placeholder": "Mindestens 6 Zeichen", "already_have_account": "Du hast schon ein Konto?",
        "login": "Anmelden", "remember_device": "Dieses Gerät 30 Tage merken", "no_account_yet": "Noch kein Konto?", "create_one": "Eins erstellen"
    }
}
for _lang, _items in AUTH_TRANSLATIONS.items():
    EXTRA_TRANSLATIONS.setdefault(_lang, {}).update(_items)


V3_TRANSLATIONS = {'en': {'performance_dashboard': 'Shift Performance', 'weekly_report': 'Weekly Manager Report', 'ai_email_generator': 'AI Email Generator', 'ai_handover': 'AI Handover', 'probation_review': 'Probation Review', 'qr_yard_check': 'QR Yard Check', 'site_scorecard': 'Site Scorecard', 'multi_site': 'Multi Site', 'business_branding': 'Business Branding', 'weather': 'Weather', 'kanban': 'Kanban', 'generate': 'Generate', 'site': 'Site', 'sites': 'Sites', 'score': 'Score'}, 'hu': {'performance_dashboard': 'Műszak teljesítmény', 'weekly_report': 'Heti manager riport', 'ai_email_generator': 'AI email generátor', 'ai_handover': 'AI átadás', 'probation_review': 'Próbaidő értékelés', 'qr_yard_check': 'QR yard ellenőrzés', 'site_scorecard': 'Telephely scorecard', 'multi_site': 'Multi Site', 'business_branding': 'Business branding', 'weather': 'Időjárás', 'kanban': 'Kanban', 'generate': 'Generálás', 'site': 'Telephely', 'sites': 'Telephelyek', 'score': 'Pontszám'}, 'pl': {'performance_dashboard': 'Wydajność zmiany', 'weekly_report': 'Tygodniowy raport managera', 'ai_email_generator': 'Generator emaili AI', 'ai_handover': 'AI przekazanie', 'probation_review': 'Ocena okresu próbnego', 'qr_yard_check': 'QR kontrola placu', 'site_scorecard': 'Wynik lokalizacji', 'multi_site': 'Multi Site', 'business_branding': 'Branding biznesowy', 'weather': 'Pogoda', 'kanban': 'Kanban', 'generate': 'Generuj', 'site': 'Lokalizacja', 'sites': 'Lokalizacje', 'score': 'Wynik'}, 'ro': {'performance_dashboard': 'Performanță tură', 'weekly_report': 'Raport manager săptămânal', 'ai_email_generator': 'Generator email AI', 'ai_handover': 'Predare AI', 'probation_review': 'Evaluare perioadă probă', 'qr_yard_check': 'Verificare curte QR', 'site_scorecard': 'Scor locație', 'multi_site': 'Multi Site', 'business_branding': 'Branding business', 'weather': 'Vreme', 'kanban': 'Kanban', 'generate': 'Generează', 'site': 'Locație', 'sites': 'Locații', 'score': 'Scor'}, 'es': {'performance_dashboard': 'Rendimiento del turno', 'weekly_report': 'Informe semanal manager', 'ai_email_generator': 'Generador de emails AI', 'ai_handover': 'Traspaso AI', 'probation_review': 'Evaluación de prueba', 'qr_yard_check': 'Revisión QR de patio', 'site_scorecard': 'Scorecard de sitio', 'multi_site': 'Multi Site', 'business_branding': 'Branding empresarial', 'weather': 'Tiempo', 'kanban': 'Kanban', 'generate': 'Generar', 'site': 'Sitio', 'sites': 'Sitios', 'score': 'Puntuación'}, 'de': {'performance_dashboard': 'Schichtleistung', 'weekly_report': 'Wöchentlicher Managerbericht', 'ai_email_generator': 'KI E-Mail Generator', 'ai_handover': 'KI Übergabe', 'probation_review': 'Probezeitbewertung', 'qr_yard_check': 'QR Hofprüfung', 'site_scorecard': 'Standort-Scorecard', 'multi_site': 'Multi Site', 'business_branding': 'Business Branding', 'weather': 'Wetter', 'kanban': 'Kanban', 'generate': 'Generieren', 'site': 'Standort', 'sites': 'Standorte', 'score': 'Score'}}
for _lang, _items in V3_TRANSLATIONS.items():
    EXTRA_TRANSLATIONS.setdefault(_lang, {}).update(_items)

for _lang, _items in EXTRA_TRANSLATIONS.items():
    TRANSLATIONS.setdefault(_lang, {}).update(_items)

APP_I18N_FIX = {
    "en": {"appearance_themes":"Appearance / Themes","appearance_company_theme":"Appearance / Company Theme","appearance_company_theme_help":"Choose a UK logistics style, or use Custom and set your own colour. No company logos are used.","company_theme":"Company Theme","custom_accent_colour":"Custom Accent Colour","show_theme_label":"Show company theme name under WHS logo","show_theme_label_help":"Turn this off if you want only the colours to change."},
    "hu": {"appearance_themes":"Megjelenés / témák","appearance_company_theme":"Megjelenés / céges téma","appearance_company_theme_help":"Válassz UK logisztikai stílust, vagy használd az egyedi színt. Céges logókat nem használunk.","company_theme":"Céges téma","custom_accent_colour":"Egyedi kiemelő szín","show_theme_label":"Céges téma nevének megjelenítése a WHS logó alatt","show_theme_label_help":"Kapcsold ki, ha csak a színeket szeretnéd módosítani."},
    "pl": {"appearance_themes":"Wygląd / motywy","appearance_company_theme":"Wygląd / motyw firmowy","appearance_company_theme_help":"Wybierz styl logistyczny UK albo użyj własnego koloru. Logotypy firm nie są używane.","company_theme":"Motyw firmowy","custom_accent_colour":"Własny kolor akcentu","show_theme_label":"Pokaż nazwę motywu firmowego pod logo WHS","show_theme_label_help":"Wyłącz, jeśli chcesz zmieniać tylko kolory."},
    "ro": {"appearance_themes":"Aspect / teme","appearance_company_theme":"Aspect / temă companie","appearance_company_theme_help":"Alege un stil logistic UK sau folosește Custom și setează propria culoare. Nu se folosesc logo-uri de companie.","company_theme":"Temă companie","custom_accent_colour":"Culoare accent personalizată","show_theme_label":"Afișează numele temei sub logo-ul WHS","show_theme_label_help":"Oprește dacă vrei să se schimbe doar culorile."},
    "es": {"appearance_themes":"Apariencia / temas","appearance_company_theme":"Apariencia / tema de empresa","appearance_company_theme_help":"Elige un estilo logístico del Reino Unido o usa Custom y define tu propio color. No se usan logos de empresas.","company_theme":"Tema de empresa","custom_accent_colour":"Color de acento personalizado","show_theme_label":"Mostrar el nombre del tema bajo el logo WHS","show_theme_label_help":"Desactívalo si solo quieres cambiar los colores."},
    "de": {"appearance_themes":"Darstellung / Designs","appearance_company_theme":"Darstellung / Firmen-Design","appearance_company_theme_help":"Wähle einen UK-Logistikstil oder nutze Benutzerdefiniert und setze deine eigene Farbe. Firmenlogos werden nicht verwendet.","company_theme":"Firmen-Design","custom_accent_colour":"Eigene Akzentfarbe","show_theme_label":"Designnamen unter dem WHS-Logo anzeigen","show_theme_label_help":"Ausschalten, wenn nur die Farben geändert werden sollen."}
}
for _lang, _items in APP_I18N_FIX.items():
    TRANSLATIONS.setdefault(_lang, {}).update(_items)

def tr(key):
    user = current_user()
    lang = row_get(user, "language", "en") if user else session.get("language", "en")
    return TRANSLATIONS.get(lang, TRANSLATIONS["en"]).get(key, TRANSLATIONS["en"].get(key, key))

def current_language():
    user = current_user()
    return row_get(user, "language", session.get("language", "en")) if user else session.get("language", "en")

FAVORITE_TOOL_OPTIONS = [
    ("morning_brief", "📢", "Brief", "morning_brief"),
    ("shift_calendar", "📅", "Calendar", "shift_calendar"),
    ("handover", "📝", "Handover", "handover"),
    ("yard_check", "🚛", "Yard Check", "yard_check"),
    ("mileage", "🚗", "Mileage", "mileage"),
    ("expenses", "💷", "Expenses", "expenses"),
    ("team", "👥", "Team", "team"),
    ("holiday_tracker", "🌴", "Holiday Tracker", "holiday_tracker"),
    ("index", "☁️", "Weather", "index"),
    ("yard_settings", "⚙️", "Yard Settings", "yard_settings"),
    ("admin", "🛡️", "Admin", "admin_dashboard"),
]


COMPANY_THEME_OPTIONS = [
    ("whs", "WHS Default"),
    ("dhl", "DHL style colours"),
    ("amazon", "Amazon style colours"),
    ("gxo", "GXO style colours"),
    ("wincanton", "Wincanton style colours"),
    ("xpo", "XPO style colours"),
    ("evri", "Evri style colours"),
    ("royal_mail", "Royal Mail style colours"),
    ("yodel", "Yodel style colours"),
    ("dpd", "DPD style colours"),
    ("ceva", "CEVA style colours"),
    ("custom", "Custom colour"),
]

COMPANY_THEME_META = {
    # Colours are inspired by public brand-style colour palettes/screenshots. No official company logos are used.
    "whs": {"label": "WHS", "accent": "#06b6d4", "second": "#0f172a", "third": "#ffffff"},
    "dhl": {"label": "DHL", "accent": "#ffcc00", "second": "#d40511", "third": "#c9c9c9"},
    "amazon": {"label": "Amazon", "accent": "#ff9900", "second": "#232f3e", "third": "#146eb4"},
    "gxo": {"label": "GXO", "accent": "#ff3a00", "second": "#111111", "third": "#ffffff"},
    "wincanton": {"label": "Wincanton", "accent": "#005baa", "second": "#00aeef", "third": "#ffffff"},
    "xpo": {"label": "XPO", "accent": "#cc0000", "second": "#000000", "third": "#ffffff"},
    "evri": {"label": "Evri", "accent": "#009bdf", "second": "#10105a", "third": "#bfe7f7"},
    "royal_mail": {"label": "Royal Mail", "accent": "#e84142", "second": "#ffd200", "third": "#ffffff"},
    "yodel": {"label": "Yodel", "accent": "#8dc63f", "second": "#374151", "third": "#ffffff"},
    "dpd": {"label": "DPD", "accent": "#dc0032", "second": "#414042", "third": "#ffffff"},
    "ceva": {"label": "CEVA", "accent": "#1f2a44", "second": "#e30613", "third": "#a50034"},
    "custom": {"label": "Custom", "accent": "#f59e0b", "second": "#111827", "third": "#ffffff"},
}

def theme_meta_for(user):
    key = row_get(user, "company_theme", "whs") if user else "whs"
    meta = dict(COMPANY_THEME_META.get(key, COMPANY_THEME_META["whs"]))
    if key == "custom" and user and row_get(user, "brand_color", None):
        meta["accent"] = row_get(user, "brand_color")
    return meta

def get_favorite_tools(user):
    default = "morning_brief,shift_calendar,handover,yard_check"
    raw = row_get(user, "favorite_tools", default) or default
    selected = [x.strip() for x in raw.split(",") if x.strip()]
    option_map = {x[0]: x for x in FAVORITE_TOOL_OPTIONS}
    out = []
    is_admin_user = bool(user and row_get(user, "email", "") == "admin@whs-app.com")
    for key in selected:
        if key == "admin" and not is_admin_user:
            continue
        if key in option_map and key not in [x[0] for x in out]:
            out.append(option_map[key])
    for key in ["morning_brief", "shift_calendar", "handover", "yard_check"]:
        if len(out) >= 4:
            break
        if key in option_map and key not in [x[0] for x in out]:
            out.append(option_map[key])
    return out[:4]

def available_favorite_options(user):
    is_admin_user = bool(user and row_get(user, "email", "") == "admin@whs-app.com")
    return [x for x in FAVORITE_TOOL_OPTIONS if x[0] != "admin" or is_admin_user]


@app.template_filter("days_until")
def days_until(value):
    if not value:
        return None
    try:
        d = datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        return (d - datetime.today().date()).days
    except Exception:
        return None

def get_notifications(user_id):
    conn = get_db()
    today = datetime.today().date()
    soon = (today + timedelta(days=30)).isoformat()
    today_s = today.isoformat()
    items = []
    for r in conn.execute("SELECT name, probation_end FROM team_members WHERE user_id=? AND probation_end BETWEEN ? AND ? AND COALESCE(probation_status,'') NOT IN ('Passed','Terminated') ORDER BY probation_end", (user_id, today_s, soon)).fetchall():
        items.append({"level":"red", "text": f"{r['name']} probation ends on {r['probation_end']}", "link":"team"})
    for r in conn.execute("SELECT name, licence_expiry FROM team_members WHERE user_id=? AND licence_expiry BETWEEN ? AND ? ORDER BY licence_expiry", (user_id, today_s, soon)).fetchall():
        items.append({"level":"orange", "text": f"{r['name']} licence expires on {r['licence_expiry']}", "link":"team"})
    for r in conn.execute("SELECT name, training_type, training_expiry FROM team_members WHERE user_id=? AND training_expiry BETWEEN ? AND ? ORDER BY training_expiry", (user_id, today_s, soon)).fetchall():
        items.append({"level":"orange", "text": f"{r['name']} {r['training_type'] or 'training'} expires on {r['training_expiry']}", "link":"team"})
    overdue = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed' AND due_date < ?", (user_id, today_s)).fetchone()[0]
    if overdue:
        items.append({"level":"red", "text": f"{overdue} action(s) overdue", "link":"actions"})
    open_actions = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed'", (user_id,)).fetchone()[0]
    if open_actions:
        items.append({"level":"blue", "text": f"{open_actions} open action(s) require follow-up", "link":"actions"})
    conn.close()
    return items[:10]

@app.context_processor
def inject_context():
    user = current_user()
    return {
        "current_year": datetime.now().year,
        "user": user,
        "plan_names": PLAN_NAMES,
        "shift_trial": shift_calendar_trial_info(user) if user else None,
        "handover_trial": feature_trial_info(user, "handover") if user else None,
        "yard_trial": feature_trial_info(user, "yard_check") if user else None,
        "t": tr,
        "app_name": "WHS",
        "app_full_name": "Warehouse Support",
        "available_languages": [("en","English"),("hu","Magyar"),("pl","Polski"),("ro","Română"),("es","Español"),("de","Deutsch")],
        "current_language": current_language(),
        "favorite_options": FAVORITE_TOOL_OPTIONS,
        "company_theme_meta": theme_meta_for(user),
        "company_theme_meta_map": COMPANY_THEME_META,
        "notifications": get_notifications(user["id"]) if user else []
    }




def get_system_setting(key, default=None):
    conn = get_db()
    row = conn.execute("SELECT value FROM system_settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row["value"] if row and row["value"] is not None else default


def set_system_setting(key, value):
    conn = get_db()
    conn.execute("""
        INSERT INTO system_settings (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (key, value))
    conn.commit()
    conn.close()


def get_stripe_config():
    return {
        "secret_key": get_system_setting("stripe_secret_key", os.environ.get("STRIPE_SECRET_KEY", "")),
        "publishable_key": get_system_setting("stripe_publishable_key", os.environ.get("STRIPE_PUBLISHABLE_KEY", "")),
        "webhook_secret": get_system_setting("stripe_webhook_secret", os.environ.get("STRIPE_WEBHOOK_SECRET", "")),
        "pro_price_id": get_system_setting("stripe_pro_price_id", os.environ.get("STRIPE_PRO_PRICE_ID", "")),
        "app_base_url": get_system_setting("app_base_url", os.environ.get("APP_BASE_URL", "http://127.0.0.1:5000")),
        "billing_mode": get_system_setting("billing_mode", "rolling"),
        "payout_note": get_system_setting("payout_note", "Stripe payouts are controlled from your Stripe Dashboard payout settings."),
    }


def first_day_next_month_timestamp():
    now = datetime.now()
    if now.month == 12:
        next_month = datetime(now.year + 1, 1, 1)
    else:
        next_month = datetime(now.year, now.month + 1, 1)
    return int(next_month.timestamp())


def stripe_ready():
    cfg = get_stripe_config()
    return bool(stripe and cfg["secret_key"] and cfg["pro_price_id"])


def row_get(row, key, default=None):
    try:
        if row is None:
            return default
        return row[key]
    except Exception:
        return default


def get_custom_locations(user_id):
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM custom_locations WHERE user_id = ? ORDER BY name ASC", (user_id,)).fetchall()
        return rows
    except Exception:
        return []
    finally:
        conn.close()

def excel_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def style_excel_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(16, min(35, max(len(str(c.value or "")) for c in col) + 2))



def workbook_to_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def send_excel_to_self(workbook, filename, export_title):
    """Email an Excel workbook to the logged-in user's registered email.
    Fails safely when SMTP is not configured.
    """
    user = current_user()
    recipient = row_get(user, "email", "")
    if not recipient:
        return False, "No registered email address found for your account."
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SMTP_FROM]):
        return False, "Email sending is not configured yet. Add SMTP_HOST/SMTP_SERVER, SMTP_USER/SMTP_EMAIL, SMTP_PASSWORD/SMTP_APP_PASSWORD and SMTP_FROM in Render Environment Variables."

    msg = EmailMessage()
    msg["Subject"] = f"WHS Export – {export_title}"
    msg["From"] = SMTP_FROM
    msg["To"] = recipient
    msg.set_content(
        f"Hi {row_get(user, 'name', 'there')},\n\n"
        f"Your WHS {export_title} Excel export is attached.\n\n"
        "Kind regards,\nWHS Team"
    )
    html = f"""
    <div style=\"font-family:Arial,sans-serif;line-height:1.5;color:#0f172a\">
      <h2 style=\"margin:0 0 10px\">WHS Export</h2>
      <p>Hi {row_get(user, 'name', 'there')},</p>
      <p>Your <strong>{export_title}</strong> Excel export is attached.</p>
      <p style=\"color:#64748b\">This email was sent automatically from WHS – Warehouse Support.</p>
      <p>Kind regards,<br><strong>WHS Team</strong></p>
    </div>
    """
    msg.add_alternative(html, subtype="html")
    msg.add_attachment(
        workbook_to_bytes(workbook),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )

    if SMTP_USE_SSL:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ssl.create_default_context()) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            if SMTP_USE_TLS:
                server.starttls(context=ssl.create_default_context())
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    return True, f"Excel sent to {recipient}."


def simple_table_workbook(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_excel_header(ws)
    return wb


EMAIL_EXPORT_CONFIG = {
    "mileage": {
        "title": "Mileage",
        "filename": "whs-mileage.xlsx",
        "sql": "SELECT date, start_location, end_location, purpose, miles, rate, created_at FROM mileage WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "From", "To", "Purpose", "Miles", "Rate", "Recorded At"],
        "return": "mileage",
    },
    "expenses": {
        "title": "Expenses",
        "filename": "whs-expenses.xlsx",
        "sql": "SELECT date, category, description, amount, created_at FROM expenses WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "Category", "Description", "Amount", "Recorded At"],
        "return": "expenses",
    },
    "yard_check": {
        "title": "Yard Check",
        "filename": "whs-yard-check.xlsx",
        "sql": "SELECT date, location_type, location_detail, trailer_id, notes, source, created_at FROM yard_checks WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "Location Type", "Location", "Trailer ID", "Notes", "Source", "Recorded At"],
        "return": "yard_check",
    },
    "handover": {
        "title": "Handover",
        "filename": "whs-handovers.xlsx",
        "sql": "SELECT date, shift, manager, attendance, safety, operations, issues, actions, created_at FROM handovers WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "Shift", "Manager", "Attendance", "Safety", "Operations", "Issues", "Actions", "Recorded At"],
        "return": "handover",
    },
    "team": {
        "title": "Team Members",
        "filename": "whs-team-members.xlsx",
        "sql": "SELECT name, role, email, phone, status, probation_start, probation_end, probation_status, notes, created_at FROM team_members WHERE user_id = ? ORDER BY name ASC",
        "headers": ["Name", "Role", "Email", "Phone", "Status", "Probation Start", "Probation End", "Probation Status", "Notes", "Created At"],
        "return": "team",
    },
    "shift_calendar": {
        "title": "Calendar",
        "filename": "whs-calendar.xlsx",
        "sql": "SELECT date, status, shift_name, start_time, end_time, notes, holiday_hours, holiday_days, created_at FROM shift_calendar WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "Status", "Shift", "Start", "End", "Notes", "Holiday Hours", "Holiday Days", "Recorded At"],
        "return": "shift_calendar",
    },
    "holiday_tracker": {
        "title": "Holiday Tracker",
        "filename": "whs-holiday-tracker.xlsx",
        "sql": "SELECT date, status, shift_name, notes, holiday_hours, holiday_days, created_at FROM shift_calendar WHERE user_id = ? AND (status = 'Holiday' OR holiday_hours > 0 OR holiday_days > 0) ORDER BY date DESC, id DESC",
        "headers": ["Date", "Status", "Shift", "Notes", "Holiday Hours", "Holiday Days", "Recorded At"],
        "return": "holiday_tracker",
    },
    "daily_shift_log": {
        "title": "Daily Shift Log",
        "filename": "whs-daily-shift-log.xlsx",
        "sql": "SELECT date, shift, manager, volume, planned_hc, actual_hc, safety, issues, actions, created_at FROM daily_shift_logs WHERE user_id = ? ORDER BY date DESC, id DESC",
        "headers": ["Date", "Shift", "Manager", "Volume", "Planned HC", "Actual HC", "Safety", "Issues", "Actions", "Recorded At"],
        "return": "daily_shift_log",
    },
    "actions": {
        "title": "Actions",
        "filename": "whs-actions.xlsx",
        "sql": "SELECT title, owner, due_date, priority, status, notes, created_at FROM action_tracker WHERE user_id = ? ORDER BY due_date ASC, id DESC",
        "headers": ["Action", "Owner", "Due Date", "Priority", "Status", "Notes", "Recorded At"],
        "return": "actions",
    },
    "absence": {
        "title": "Absence",
        "filename": "whs-absence.xlsx",
        "sql": "SELECT member_name, absence_type, start_date, end_date, notes, created_at FROM absence_records WHERE user_id = ? ORDER BY start_date DESC, id DESC",
        "headers": ["Person", "Type", "Start", "End", "Notes", "Recorded At"],
        "return": "absence",
    },
}


@app.route("/email-excel/<kind>")
@login_required
def email_excel(kind):
    flash("Email Excel has been disabled. Please use Download Excel instead.", "info")
    return redirect(url_for("dashboard"))




def get_yard_config(user):
    is_pro_user = row_get(user, "plan") in ["pro", "business"] or is_admin(user)

    if not is_pro_user:
        return {
            "door_start": 1,
            "door_end": 100,
            "fence_start": 1,
            "fence_end": 120,
            "editable": False,
        }

    door_start = int(row_get(user, "door_start", 1) or 1)
    door_end = int(row_get(user, "door_end", row_get(user, "door_count", 100)) or 100)
    fence_start = int(row_get(user, "fence_start", 1) or 1)
    fence_end = int(row_get(user, "fence_end", row_get(user, "fence_count", 120)) or 120)

    door_start = max(1, min(door_start, 9999))
    door_end = max(door_start, min(door_end, 9999))
    fence_start = max(1, min(fence_start, 9999))
    fence_end = max(fence_start, min(fence_end, 9999))

    return {
        "door_start": door_start,
        "door_end": door_end,
        "fence_start": fence_start,
        "fence_end": fence_end,
        "editable": True,
    }


def refresh_user_plan(user):
    if not user:
        return user

    try:
        email = row_get(user, "email")
        expires_at = row_get(user, "pro_expires_at")

        if email == "admin@whs-app.com":
            return user

        if expires_at:
            expiry = datetime.fromisoformat(expires_at)
            if datetime.now() > expiry:
                conn = get_db()
                conn.execute("""
                    UPDATE users
                    SET plan = 'free',
                        subscription_status = 'expired',
                        pro_expires_at = NULL,
                        pro_reason = NULL
                    WHERE id = ?
                """, (user["id"],))
                conn.commit()
                refreshed = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
                conn.close()
                return refreshed
    except Exception:
        return user

    return user




def shift_calendar_trial_info(user):
    """Free users can use Shift Calendar for 14 days after account creation."""
    if not user:
        return {"allowed": False, "days_left": 0, "is_trial": False}

    if is_admin(user) or row_get(user, "plan") in ["pro", "business"]:
        return {"allowed": True, "days_left": None, "is_trial": False}

    try:
        created = datetime.fromisoformat(row_get(user, "created_at"))
        trial_end = created + timedelta(days=14)
        now = datetime.now()
        days_left = max((trial_end.date() - now.date()).days, 0)
        return {
            "allowed": now <= trial_end,
            "days_left": days_left,
            "is_trial": True,
            "trial_end": trial_end.date().isoformat(),
        }
    except Exception:
        return {"allowed": False, "days_left": 0, "is_trial": True}



def create_remember_token(user_id):
    token = secrets.token_urlsafe(48)
    expires = datetime.now() + timedelta(days=30)
    conn = get_db()
    conn.execute("""
        INSERT INTO remember_tokens (user_id, token, expires_at, created_at, user_agent)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, token, expires.isoformat(), datetime.now().isoformat(), request.headers.get("User-Agent", "")))
    conn.commit()
    conn.close()
    return token, expires


def consume_remember_token():
    if session.get("user_id"):
        return

    token = request.cookies.get("whs_remember")
    if not token:
        return

    conn = get_db()
    row = conn.execute("""
        SELECT rt.*, u.id AS uid
        FROM remember_tokens rt
        JOIN users u ON u.id = rt.user_id
        WHERE rt.token = ?
    """, (token,)).fetchone()
    conn.close()

    if not row:
        return

    try:
        expires = datetime.fromisoformat(row["expires_at"])
        if expires < datetime.now():
            conn = get_db()
            conn.execute("DELETE FROM remember_tokens WHERE token = ?", (token,))
            conn.commit()
            conn.close()
            return
        session["user_id"] = row["user_id"]
    except Exception:
        return


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please register or log in first.", "error")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def plan_required(required_plan):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                flash("Please register or log in first.", "error")
                return redirect(url_for("login"))
            # WHS is currently free: every registered user has full access.
            return fn(*args, **kwargs)
        return wrapper
    return decorator




def feature_trial_info(user, feature_key, days=14):
    """14 day free trial for selected features, based on account creation date."""
    if not user:
        return {"allowed": False, "days_left": 0, "expired": True}
    if is_admin(user) or PLAN_ORDER.get(row_get(user, "plan", "free"), 0) >= PLAN_ORDER["pro"]:
        return {"allowed": True, "days_left": None, "expired": False}
    created_raw = row_get(user, "created_at", None)
    try:
        created = datetime.fromisoformat(str(created_raw).replace("Z", ""))
    except Exception:
        created = datetime.now()
    expires = created + timedelta(days=days)
    now = datetime.now()
    days_left = max((expires.date() - now.date()).days, 0)
    allowed = now <= expires
    return {"allowed": allowed, "days_left": days_left, "expired": not allowed, "expires_at": expires.isoformat(timespec="seconds"), "feature": feature_key}


def free_trial_allowed(user, feature_key):
    return feature_trial_info(user, feature_key).get("allowed", False)

def money(value):
    return f"£{float(value or 0):,.2f}"


app.jinja_env.filters["money"] = money
app.jinja_env.globals["display_plan_name"] = display_plan_name


@app.template_filter("dateuk")
def dateuk(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return value


def calculate_tax(profit):
    profit = float(profit or 0)
    basic_band = max(min(profit, BASIC_RATE_LIMIT) - PERSONAL_ALLOWANCE, 0)
    higher_band = max(profit - BASIC_RATE_LIMIT, 0)
    income_tax = basic_band * INCOME_TAX_BASIC + higher_band * INCOME_TAX_HIGHER

    class4_basic_band = max(min(profit, CLASS4_NI_UPPER) - CLASS4_NI_LOWER, 0)
    class4_higher_band = max(profit - CLASS4_NI_UPPER, 0)
    class4_ni = class4_basic_band * CLASS4_NI_BASIC + class4_higher_band * CLASS4_NI_HIGHER

    class2_ni = CLASS2_NI_WEEKLY * 52 if profit > 6725 else 0
    return income_tax + class4_ni + class2_ni


def totals(user_id):
    conn = get_db()
    income = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM invoices WHERE user_id = ?", (user_id,)).fetchone()[0]
    expenses = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE user_id = ?", (user_id,)).fetchone()[0]
    miles = conn.execute("SELECT COALESCE(SUM(miles), 0) FROM mileage WHERE user_id = ?", (user_id,)).fetchone()[0]
    conn.close()

    mileage_claim = miles * HMRC_MILE_RATE
    taxable_profit = max(income - expenses - mileage_claim, 0)
    estimated_tax = calculate_tax(taxable_profit)
    net_profit = income - expenses

    return {
        "income": income,
        "expenses": expenses,
        "miles": miles,
        "mileage_claim": mileage_claim,
        "taxable_profit": taxable_profit,
        "estimated_tax": estimated_tax,
        "net_profit": net_profit,
    }


def next_invoice_number(user_id):
    conn = get_db()
    row = conn.execute("SELECT id FROM invoices WHERE user_id = ? ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
    conn.close()
    next_id = (row["id"] + 1) if row else 1
    return f"OP-{datetime.now().year}-{next_id:04d}"



def generate_morning_brief_text(date, shift, manager, role, volume, available_hc, late_trailers, safety_message, priorities, team_messages, break_reminder, equipment_reminder, custom_sections=None):
    lines = [
        "Good morning team,",
        "",
        f"Today we are running {shift or 'Day'} shift.",
        "",
        f"Manager: {manager or 'Not specified'}",
        f"Role: {role or 'Not specified'}",
        "",
        "Today's Plan:",
        f"- Expected volume: {volume}",
        f"- Available HC: {available_hc}",
        "",
    ]

    if custom_sections:
        for section in custom_sections:
            header = (section.get('header') or '').strip()
            value = (section.get('value') or '').strip()
            if not header and not value:
                continue
            lines.append(f"{header or 'Section'}:")
            lines.append(value or 'No message added.')
            lines.append("")
    else:
        default_sections = [
            ("Safety Message", safety_message or 'Keep the area safe, clean and controlled. Report hazards immediately.'),
            ("Priorities", priorities or 'Focus on trailer control, clean handovers and completing work safely.'),
            ("Team Messages", team_messages or 'Keep communication clear and support each other during the shift.'),
            ("Break Reminder", break_reminder or 'Take breaks in a controlled way and make sure the operation is covered.'),
            ("Equipment / MHE / Scanner / Key Reminder", equipment_reminder or 'Return all MHE keys, scanners and equipment at the end of the shift.'),
        ]
        for header, value in default_sections:
            lines.append(f"{header}:")
            lines.append(value)
            lines.append("")

    lines.append("Let’s keep it safe, organised and productive.")
    return "\n".join(lines) + "\n"

def generate_handover_text(date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, extra=None):
    extra = extra or {}

    def v(key, default=""):
        value = extra.get(key, default)
        if value is None:
            return ""
        if isinstance(value, (list, dict)):
            return value
        return str(value).strip()

    section_names = default_handover_section_names()
    try:
        saved_names = extra.get("section_names") or {}
        if isinstance(saved_names, dict):
            section_names.update(saved_names)
    except Exception:
        pass

    absence_rows = extra.get("absence_rows", [])
    dispatch_rows = extra.get("dispatch_rows", [])

    absence_text = "\n".join([f"- {r.get('type','')}: {r.get('count','')}" for r in absence_rows if r.get("type") or r.get("count")]) or "No absence breakdown added."
    dispatch_text = "\n".join([
        f"- {r.get('carrier','')} | VRID: {r.get('vrid','')} | Completed: {r.get('completed','')} | On Time: {r.get('on_time','')} | Issue: {r.get('issue','')}"
        for r in dispatch_rows if r.get("carrier") or r.get("vrid")
    ]) or "No dispatch activities added."

    attendance = f"{actual_hc}/{planned_hc}" if planned_hc else str(actual_hc or "")

    return f"""{shift or 'Shift'} Handover
Date: {date}
Shift: {shift or 'Not specified'}
Manager: {manager or 'Not specified'}
Attendance: {attendance}

{section_names.get('attendance', 'Attendance').upper()}
{absence_text}

{section_names.get('safety', 'Safety Metrics').upper()}
Safe Shift: {v('safe_shift', 'Safe Shift')}
SLAMs: {v('slams')}
Safety Cons: {v('safety_cons')}
Loads: {v('loads')}
Safety Rules: {v('safety_rules')}
Additional Comments: {v('safety_comments')}

{section_names.get('operations', 'Operations Pick').upper()}
Pick Audits Completed: {v('pick_audits')}
Slam Plan: {v('slam_plan')}
Picked Since: {v('picked_since_time', '06:00')} - {v('picked_since_value')}
Full Well Start: {v('full_well_start')}
Full Well End: {v('full_well_end')}
EMC / Well To Cover Start: {v('emc_start')}
EMC / Well To Cover End: {v('emc_end')}
Additional Comments: {v('ops_comments') or issues or 'No major issues reported.'}

{section_names.get('sort', 'CUK8 Sort Centre').upper()}
Deliveries Planned: {v('deliveries_planned')}
Deliveries Arrived: {v('deliveries_arrived')}
Planned Same Day Sortation: {v('same_day_sortation')}
Planned Next Day Sortation: {v('next_day_sortation')}
Additional Comments: {v('sort_comments')}

{section_names.get('dispatch', 'Dispatch').upper()}
Collections Arrived: {v('collections_arrived')}
Late Arrivals: {v('late_arrivals')}
Trailers on Doors: {v('trailers_on_doors')}
Trailers Needed to Cover Today CPTs: {v('trailers_needed_cover')}

Dispatch Activities:
{dispatch_text}

Additional Comments: {v('dispatch_comments') or actions or 'No further action required.'}

{section_names.get('suntory', 'Suntory').upper()}
Trailers on Site to Complete: {v('suntory_trailers_on_site')}
Completed on Shift: {v('suntory_completed')}
Left to Complete: {v('suntory_left')}
Additional Comments: {v('suntory_comments')}

{section_names.get('aob', 'AOB').upper()}
{v('aob')}
"""
def generate_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours):
    capacity = available_hc * target_rate * planned_hours if available_hc and target_rate and planned_hours else 0
    gap = capacity - volume
    suggested_hc = volume / (target_rate * planned_hours) if target_rate and planned_hours else 0

    if capacity <= 0:
        status = "Not enough information to calculate capacity."
    elif gap >= 0:
        status = f"Plan looks achievable. Estimated spare capacity: {gap:,.0f} units."
    else:
        status = f"Risk: estimated shortfall of {abs(gap):,.0f} units. Consider extra labour, overtime or priority control."

    return f"""AI Shift Planner

Date: {date}
Shift: {shift}

Input:
- Expected volume: {volume}
- Available HC: {available_hc}
- Target rate/person/hour: {target_rate}
- Planned working hours: {planned_hours}

Capacity:
- Estimated capacity: {capacity:,.0f} units
- Suggested HC required: {suggested_hc:.1f}

Result:
{status}

Suggested Plan:
1. Start priority trailers/lanes first.
2. Put strongest colleagues in the highest risk area.
3. Review progress after the first break.
4. If behind plan, move labour early rather than waiting until end of shift.
5. Create a daily handover with remaining risks and actions.
"""



def get_openai_client():
    if not OPENAI_API_KEY or OpenAI is None:
        return None
    return OpenAI(api_key=OPENAI_API_KEY)


def generate_ai_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours):
    fallback = generate_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours)
    client = get_openai_client()
    if not client:
        return fallback + "\n\nAI Mode: fallback planner used. Add OPENAI_API_KEY for live AI planning."

    prompt = f"""
Create a practical warehouse shift plan.
Date: {date}
Shift: {shift}
Expected volume: {volume}
Available HC: {available_hc}
Target rate per person per hour: {target_rate}
Planned hours: {planned_hours}

Give:
- capacity estimate
- risk level
- staffing recommendation
- break review points
- manager actions
- handover notes
Use simple UK warehouse English.
"""
    try:
        response = client.chat.completions.create(
            model=os.environ.get("OPENAI_TEXT_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        return fallback + f"\n\nAI Mode failed, fallback used. Error: {exc}"


def analyse_photo_ai(trailer_id, location_detail, damage_notes, recognition_notes):
    client = get_openai_client()
    if not client:
        return (
            "Photo AI fallback result:\\n"
            f"- Trailer ID entered: {trailer_id or 'Not provided'}\\n"
            f"- Location entered: {location_detail or 'Not provided'}\\n"
            f"- Damage notes: {damage_notes or 'None'}\\n"
            f"- Recognition notes: {recognition_notes or 'None'}\\n"
            "Add OPENAI_API_KEY and vision model integration for automatic image reading."
        )

    prompt = f"""
Review this manually entered trailer photo record and create a professional recognition summary.
Trailer ID: {trailer_id}
Location: {location_detail}
Damage notes: {damage_notes}
Recognition notes: {recognition_notes}

Return:
- trailer identification confidence
- possible issues
- recommended action
- handover note
"""
    try:
        response = client.chat.completions.create(
            model=os.environ.get("OPENAI_TEXT_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        return f"Photo AI failed, manual record saved. Error: {exc}"


def invoice_pdf_buffer(user, invoice):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(25 * mm, height - 30 * mm, "INVOICE")
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(25 * mm, height - 48 * mm, user["business_name"] or "WHS Business")
    pdf.setFont("Helvetica", 10)

    y = height - 56 * mm
    for line in [user["name"], user["email"], user["phone"], user["address"]]:
        if line:
            pdf.drawString(25 * mm, y, str(line))
            y -= 6 * mm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(130 * mm, height - 48 * mm, "Invoice No:")
    pdf.drawString(130 * mm, height - 56 * mm, "Date:")
    pdf.drawString(130 * mm, height - 64 * mm, "Status:")

    pdf.setFont("Helvetica", 11)
    pdf.drawString(158 * mm, height - 48 * mm, invoice["invoice_number"])
    pdf.drawString(158 * mm, height - 56 * mm, invoice["date"])
    pdf.drawString(158 * mm, height - 64 * mm, invoice["status"])

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(25 * mm, height - 90 * mm, "Bill To")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(25 * mm, height - 98 * mm, invoice["customer_name"])
    if invoice["customer_email"]:
        pdf.drawString(25 * mm, height - 106 * mm, invoice["customer_email"])

    pdf.line(25 * mm, height - 125 * mm, 185 * mm, height - 125 * mm)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(25 * mm, height - 134 * mm, "Description")
    pdf.drawRightString(185 * mm, height - 134 * mm, "Amount")
    pdf.line(25 * mm, height - 140 * mm, 185 * mm, height - 140 * mm)

    pdf.setFont("Helvetica", 11)
    pdf.drawString(25 * mm, height - 150 * mm, invoice["description"][:80])
    pdf.drawRightString(185 * mm, height - 150 * mm, f"£{invoice['amount']:,.2f}")

    pdf.line(25 * mm, height - 160 * mm, 185 * mm, height - 160 * mm)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawRightString(160 * mm, height - 176 * mm, "Total:")
    pdf.drawRightString(185 * mm, height - 176 * mm, f"£{invoice['amount']:,.2f}")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(25 * mm, 20 * mm, "Generated by WHS.")
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer


def send_invoice_email(user, invoice):
    if not invoice["customer_email"]:
        return False, "Customer email is missing."
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SMTP_FROM]):
        return False, "SMTP is not configured. Add SMTP_HOST, SMTP_USER, SMTP_PASSWORD and SMTP_FROM environment variables."

    msg = EmailMessage()
    msg["Subject"] = f"Invoice {invoice['invoice_number']} from {user['business_name'] or user['name']}"
    msg["From"] = SMTP_FROM
    msg["To"] = invoice["customer_email"]
    msg.set_content(f"""Hi {invoice['customer_name']},

Please find attached invoice {invoice['invoice_number']}.

Amount: £{invoice['amount']:,.2f}

Kind regards,
{user['name']}
""")

    pdf_data = invoice_pdf_buffer(user, invoice).read()
    msg.add_attachment(pdf_data, maintype="application", subtype="pdf", filename=f"{invoice['invoice_number']}.pdf")

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls(context=context)
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)

    return True, "Invoice email sent."




def calculate_google_maps_miles(origin, destination):
    if not GOOGLE_MAPS_API_KEY:
        return None, "Google Maps API key is missing. Add GOOGLE_MAPS_API_KEY to your environment variables."

    if requests is None:
        return None, "Requests package is missing."

    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins": origin,
        "destinations": destination,
        "units": "imperial",
        "mode": "driving",
        "key": GOOGLE_MAPS_API_KEY
    }

    try:
        response = requests.get(url, params=params, timeout=10)
        data = response.json()

        if data.get("status") != "OK":
            return None, f"Google Maps error: {data.get('status')}"

        element = data["rows"][0]["elements"][0]
        if element.get("status") != "OK":
            return None, f"Route error: {element.get('status')}"

        meters = element["distance"]["value"]
        miles = round(meters / 1609.344, 1)
        duration = element.get("duration", {}).get("text", "")
        distance_text = element.get("distance", {}).get("text", f"{miles} mi")

        return {
            "miles": miles,
            "duration": duration,
            "distance_text": distance_text,
            "origin": data.get("origin_addresses", [origin])[0],
            "destination": data.get("destination_addresses", [destination])[0],
            "claim": round(miles * HMRC_MILE_RATE, 2)
        }, None

    except Exception as exc:
        return None, f"Could not calculate mileage: {exc}"



SHIFT_STATUS_COLORS = {
    "Work": "work",
    "Off": "off",
    "Holiday": "holiday",
    "Sick": "sick",
    "Training": "training",
    "Overtime": "overtime",
    "Bank Holiday": "bankholiday",
    "Custom": "custom",
}


def parse_custom_shift_pattern(pattern_text):
    """
    Converts human-friendly pattern text to a repeating status cycle.
    Examples:
    - "3 on 4 off 4 on 3 off"
    - "2 work 2 off 3 holiday"
    - "4 day 4 off"
    """
    text = (pattern_text or "").lower().strip()
    if not text:
        return None

    text = text.replace("/", " ").replace(",", " ").replace("-", " ")
    text = text.replace("work", "on").replace("working", "on").replace("days", "on").replace("day", "on")
    text = text.replace("rest", "off").replace("offs", "off")
    text = text.replace("annual leave", "holiday")
    text = text.replace("sickness", "sick").replace("ill", "sick")
    text = re.sub(r"\s+", " ", text)

    tokens = text.split()
    cycle = []
    i = 0
    status_map = {
        "on": "Work",
        "off": "Off",
        "holiday": "Holiday",
        "sick": "Sick",
        "training": "Training",
        "overtime": "Overtime",
        "bankholiday": "Bank Holiday",
        "bank": "Bank Holiday",
        "custom": "Custom",
    }

    while i < len(tokens) - 1:
        if tokens[i].isdigit():
            count = int(tokens[i])
            word = tokens[i + 1]
            if word == "bank" and i + 2 < len(tokens) and tokens[i + 2] == "holiday":
                word = "bankholiday"
                i += 1
            status = status_map.get(word)
            if status and 0 < count <= 31:
                cycle.extend([status] * count)
                i += 2
                continue
        i += 1

    return cycle if cycle else None


def generate_shift_pattern_dates(start_date, pattern, months, shift_name, start_time, end_time, custom_pattern=""):
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = start + timedelta(days=int(months or 12) * 31)
    rows = []

    if pattern == "custom":
        cycle = parse_custom_shift_pattern(custom_pattern) or (["Work"] * 4 + ["Off"] * 4)
    elif pattern == "4on4off":
        cycle = ["Work"] * 4 + ["Off"] * 4
    elif pattern == "5on2off":
        cycle = ["Work"] * 5 + ["Off"] * 2
    elif pattern == "monfri":
        cycle = None
    elif pattern == "2days2nights4off":
        cycle = ["Work"] * 4 + ["Off"] * 4
    elif pattern == "3on4off4on3off":
        cycle = ["Work"] * 3 + ["Off"] * 4 + ["Work"] * 4 + ["Off"] * 3
    else:
        cycle = ["Work"] * 4 + ["Off"] * 4

    d = start
    i = 0
    while d <= end:
        if pattern == "monfri":
            status = "Work" if d.weekday() < 5 else "Off"
        else:
            status = cycle[i % len(cycle)]

        rows.append({
            "date": d.isoformat(),
            "status": status,
            "shift_name": shift_name,
            "start_time": start_time if status in ["Work", "Training", "Overtime"] else "",
            "end_time": end_time if status in ["Work", "Training", "Overtime"] else "",
            "notes": custom_pattern if pattern == "custom" and i == 0 else "",
            "source": "Generated",
        })
        d += timedelta(days=1)
        i += 1

    return rows

def get_week_start(date_obj=None):
    date_obj = date_obj or datetime.today().date()
    return date_obj - timedelta(days=date_obj.weekday())


def get_current_week_shift_rows(user_id):
    week_start = get_week_start()
    week_end = week_start + timedelta(days=6)

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date BETWEEN ? AND ?
        ORDER BY date ASC
    """, (user_id, week_start.isoformat(), week_end.isoformat())).fetchall()
    conn.close()

    by_date = {row["date"]: row for row in rows}
    week = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        key = d.isoformat()
        row = by_date.get(key)
        if row:
            week.append(row)
        else:
            week.append({
                "date": key,
                "status": "Not Set",
                "shift_name": "",
                "start_time": "",
                "end_time": "",
                "notes": "",
                "source": "Empty",
            })
    return week


def allowed_image(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in {"png", "jpg", "jpeg", "webp"}




def ensure_holiday_schema_updates():
    safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")
    safe_add_column("users", "annual_leave_unit", "TEXT DEFAULT 'days'")
    safe_add_column("users", "contracted_shift_hours", "REAL DEFAULT 12")
    safe_add_column("users", "break_minutes", "REAL DEFAULT 45")
    safe_add_column("users", "break_paid", "INTEGER DEFAULT 0")
    safe_add_column("users", "paid_hours_per_day", "REAL DEFAULT 12")
    safe_add_column("shift_calendar", "holiday_hours", "REAL DEFAULT 0")
    safe_add_column("shift_calendar", "holiday_days", "REAL DEFAULT 0")


def ensure_holiday_user_columns():
    safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")
    safe_add_column("users", "annual_leave_unit", "TEXT DEFAULT 'days'")
    safe_add_column("users", "contracted_shift_hours", "REAL DEFAULT 12")
    safe_add_column("users", "break_minutes", "REAL DEFAULT 45")
    safe_add_column("users", "break_paid", "INTEGER DEFAULT 0")
    safe_add_column("users", "paid_hours_per_day", "REAL DEFAULT 12")


def ensure_holiday_schema_all():
    safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")
    safe_add_column("users", "annual_leave_unit", "TEXT DEFAULT 'days'")
    safe_add_column("users", "contracted_shift_hours", "REAL DEFAULT 12")
    safe_add_column("users", "break_minutes", "REAL DEFAULT 45")
    safe_add_column("users", "break_paid", "INTEGER DEFAULT 0")
    safe_add_column("users", "paid_hours_per_day", "REAL DEFAULT 12")
    safe_add_column("shift_calendar", "holiday_hours", "REAL DEFAULT 0")
    safe_add_column("shift_calendar", "holiday_days", "REAL DEFAULT 0")

@app.before_request
def load_remembered_user():
    consume_remember_token()


@app.before_request
def apply_holiday_schema_before_request():
    try:
        ensure_holiday_schema_updates()
    except Exception:
        pass


@app.before_request
def ensure_holiday_columns_before_request():
    try:
        ensure_holiday_user_columns()
    except Exception:
        pass


@app.before_request
def ensure_holiday_schema_all_before_request():
    try:
        ensure_holiday_schema_all()
    except Exception:
        pass


def get_paid_hours_per_day(user):
    try:
        stored = row_get(user, "paid_hours_per_day")
        if stored not in (None, "") and float(stored) > 0:
            return round(float(stored), 2)
    except Exception:
        pass

    try:
        contracted = float(row_get(user, "contracted_shift_hours") or 12)
    except Exception:
        contracted = 12

    try:
        break_minutes = float(row_get(user, "break_minutes") or 0)
    except Exception:
        break_minutes = 0

    try:
        break_paid = int(row_get(user, "break_paid") or 0)
    except Exception:
        break_paid = 0

    paid = contracted if break_paid else max(contracted - (break_minutes / 60), 0)
    return round(paid, 2)



# --- FINAL HOLIDAY / PWA FIX HELPERS ---
def op_safe_add_column(table, column, definition):
    conn = get_db()
    try:
        rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
        names = [r["name"] for r in rows]
        if column not in names:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
            conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


def op_ensure_holiday_schema():
    op_safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")
    op_safe_add_column("users", "annual_leave_unit", "TEXT DEFAULT 'days'")
    op_safe_add_column("users", "contracted_shift_hours", "REAL DEFAULT 12")
    op_safe_add_column("users", "break_minutes", "REAL DEFAULT 45")
    op_safe_add_column("users", "break_paid", "INTEGER DEFAULT 0")
    op_safe_add_column("users", "paid_hours_per_day", "REAL DEFAULT 12")
    op_safe_add_column("shift_calendar", "holiday_hours", "REAL DEFAULT 0")
    op_safe_add_column("shift_calendar", "holiday_days", "REAL DEFAULT 0")


def op_get_paid_hours_per_day(user):
    try:
        stored = row_get(user, "paid_hours_per_day")
        if stored not in (None, "") and float(stored) > 0:
            return round(float(stored), 2)
    except Exception:
        pass

    try:
        contracted = float(row_get(user, "contracted_shift_hours") or 12)
    except Exception:
        contracted = 12

    try:
        break_minutes = float(row_get(user, "break_minutes") or 45)
    except Exception:
        break_minutes = 45

    try:
        break_paid = int(row_get(user, "break_paid") or 0)
    except Exception:
        break_paid = 0

    return round(contracted if break_paid else max(contracted - (break_minutes / 60), 0), 2)


def annual_leave_summary(user_id):
    op_ensure_holiday_schema()
    user = current_user()

    try:
        entitlement = float(row_get(user, "annual_leave_entitlement") or 28)
    except Exception:
        entitlement = 28

    unit = str(row_get(user, "annual_leave_unit") or "days").lower()
    if unit not in ["days", "hours"]:
        unit = "days"

    paid_hours_day = op_get_paid_hours_per_day(user)

    if unit == "hours":
        entitlement_hours = entitlement
        entitlement_days = entitlement_hours / paid_hours_day if paid_hours_day else 0
    else:
        entitlement_days = entitlement
        entitlement_hours = entitlement_days * paid_hours_day

    year = datetime.today().year
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT holiday_hours, holiday_days
            FROM shift_calendar
            WHERE user_id = ?
              AND status = 'Holiday'
              AND date BETWEEN ? AND ?
        """, (user_id, f"{year}-01-01", f"{year}-12-31")).fetchall()
    except Exception:
        rows = []
    finally:
        conn.close()

    used_hours = 0.0
    used_days = 0.0
    for row in rows:
        try:
            h = float(row["holiday_hours"] or 0)
        except Exception:
            h = 0.0
        try:
            d = float(row["holiday_days"] or 0)
        except Exception:
            d = 0.0

        if h <= 0 and d <= 0:
            d = 1.0
            h = paid_hours_day
        elif h <= 0 and d > 0:
            h = d * paid_hours_day
        elif d <= 0 and h > 0:
            d = h / paid_hours_day if paid_hours_day else 0

        used_hours += h
        used_days += d

    try:
        contracted = float(row_get(user, "contracted_shift_hours") or 12)
    except Exception:
        contracted = 12
    try:
        break_minutes = float(row_get(user, "break_minutes") or 45)
    except Exception:
        break_minutes = 45
    try:
        break_paid = int(row_get(user, "break_paid") or 0)
    except Exception:
        break_paid = 0

    return {
        "year": year,
        "unit": unit,
        "paid_hours_per_day": round(paid_hours_day, 2),
        "contracted_shift_hours": round(contracted, 2),
        "break_minutes": round(break_minutes, 2),
        "break_paid": break_paid,
        "entitlement_days": round(entitlement_days, 2),
        "entitlement_hours": round(entitlement_hours, 2),
        "used_days": round(used_days, 2),
        "used_hours": round(used_hours, 2),
        "remaining_days": round(max(entitlement_days - used_days, 0), 2),
        "remaining_hours": round(max(entitlement_hours - used_hours, 0), 2),
    }


def normalize_holiday_amounts(user, holiday_hours, holiday_days):
    try:
        holiday_hours = float(holiday_hours or 0)
    except Exception:
        holiday_hours = 0
    try:
        holiday_days = float(holiday_days or 0)
    except Exception:
        holiday_days = 0

    paid_hours = op_get_paid_hours_per_day(user)

    if holiday_hours <= 0 and holiday_days <= 0:
        holiday_days = 1
        holiday_hours = paid_hours
    elif holiday_hours > 0 and holiday_days <= 0:
        holiday_days = holiday_hours / paid_hours if paid_hours else 0
    elif holiday_days > 0 and holiday_hours <= 0:
        holiday_hours = holiday_days * paid_hours

    return round(holiday_hours, 2), round(holiday_days, 2)


def today_shift_status(user_id):
    today = datetime.today().date().isoformat()
    conn = get_db()
    try:
        row = conn.execute("""
            SELECT * FROM shift_calendar
            WHERE user_id = ? AND date = ?
            LIMIT 1
        """, (user_id, today)).fetchone()
    finally:
        conn.close()
    if row:
        return row
    return {"date": today, "status": "Not Set", "shift_name": "", "start_time": "", "end_time": "", "notes": "", "source": "Empty"}
# --- END FINAL FIX HELPERS ---


@app.before_request
def op_apply_schema_on_request():
    try:
        op_ensure_holiday_schema()
    except Exception:
        pass



def op_shift_status_for_date(user_id, date_str):
    conn = get_db()
    try:
        row = conn.execute("""
            SELECT status
            FROM shift_calendar
            WHERE user_id = ? AND date = ?
            LIMIT 1
        """, (user_id, date_str)).fetchone()
    finally:
        conn.close()
    return row["status"] if row else None


@app.route("/holiday-tracker/<int:record_id>/delete", methods=["POST"])
@login_required
def delete_holiday_record(record_id):
    user = current_user()
    conn = get_db()
    conn.execute("""
        DELETE FROM shift_calendar
        WHERE id = ? AND user_id = ? AND status = 'Holiday'
    """, (record_id, user["id"]))
    conn.commit()
    conn.close()
    flash("Holiday record deleted.", "success")
    return redirect(url_for("holiday_tracker"))


@app.route("/holiday-tracker/<int:record_id>/edit", methods=["POST"])
@login_required
def edit_holiday_record(record_id):
    op_ensure_holiday_schema()
    user = current_user()

    date = request.form.get("date")
    notes = request.form.get("notes", "Annual leave").strip() or "Annual leave"
    hours, days = normalize_holiday_amounts(user, request.form.get("holiday_hours"), request.form.get("holiday_days"))

    conn = get_db()
    conn.execute("""
        UPDATE shift_calendar
        SET date = ?,
            notes = ?,
            holiday_hours = ?,
            holiday_days = ?,
            status = 'Holiday',
            shift_name = 'Annual Leave',
            start_time = '',
            end_time = '',
            source = 'Manual'
        WHERE id = ? AND user_id = ? AND status = 'Holiday'
    """, (date, notes, hours, days, record_id, user["id"]))
    conn.commit()
    conn.close()

    flash("Holiday record updated.", "success")
    return redirect(url_for("holiday_tracker"))

@app.route("/settings/annual-leave", methods=["POST"])
@login_required
def save_annual_leave_settings():
    op_ensure_holiday_schema()
    user = current_user()

    try:
        entitlement = float(request.form.get("annual_leave_entitlement", "28") or 28)
    except Exception:
        entitlement = 28

    unit = (request.form.get("annual_leave_unit", "days") or "days").lower()
    if unit not in ["days", "hours"]:
        unit = "days"

    try:
        contracted = float(request.form.get("contracted_shift_hours", "12") or 12)
    except Exception:
        contracted = 12

    try:
        break_minutes = float(request.form.get("break_minutes", "45") or 45)
    except Exception:
        break_minutes = 45

    break_paid = 1 if request.form.get("break_paid") == "yes" else 0
    # WHS holiday logic: one holiday day can be the full shift length (e.g. 12h on 4 on / 4 off).
    # This value is used to convert hours <-> days and prevents double counting.
    try:
        paid_hours = float(request.form.get("paid_hours_per_day") or contracted)
    except Exception:
        paid_hours = contracted
    if paid_hours <= 0:
        paid_hours = contracted

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET annual_leave_entitlement = ?,
            annual_leave_unit = ?,
            contracted_shift_hours = ?,
            break_minutes = ?,
            break_paid = ?,
            paid_hours_per_day = ?
        WHERE id = ?
    """, (entitlement, unit, contracted, break_minutes, break_paid, paid_hours, user["id"]))
    conn.commit()
    conn.close()

    flash("Holiday settings saved.", "success")
    return_to = request.form.get("return_to")
    if return_to == "holiday_settings":
        return redirect(url_for("holiday_settings"))
    return redirect(url_for("holiday_tracker"))


@app.route("/holiday-tracker")
@login_required
def holiday_tracker():
    op_ensure_holiday_schema()
    user = current_user()
    summary = annual_leave_summary(user["id"])
    year = summary["year"]

    conn = get_db()
    rows = conn.execute("""
        SELECT *
        FROM shift_calendar
        WHERE user_id = ?
          AND status = 'Holiday'
          AND date BETWEEN ? AND ?
        ORDER BY date DESC
    """, (user["id"], f"{year}-01-01", f"{year}-12-31")).fetchall()
    conn.close()

    return render_template("holiday_tracker.html", page="holiday_tracker", summary=summary, rows=rows, user=user)


@app.route("/holiday-settings")
@login_required
def holiday_settings():
    op_ensure_holiday_schema()
    user = current_user()
    summary = annual_leave_summary(user["id"])
    return render_template("holiday_settings.html", page="holiday_settings", user=user, summary=summary)




@app.route("/holiday-tracker/add", methods=["POST"])
@login_required
def add_holiday_record():
    op_ensure_holiday_schema()
    user = current_user()

    start_date = request.form.get("holiday_start_date") or request.form.get("holiday_date") or datetime.today().strftime("%Y-%m-%d")
    end_date = request.form.get("holiday_end_date") or start_date
    include_off_days = request.form.get("include_off_days") == "yes"

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    except Exception:
        flash("Please select a valid start and end date.", "error")
        return redirect(url_for("holiday_tracker"))

    if end_dt < start_dt:
        start_dt, end_dt = end_dt, start_dt

    notes = request.form.get("notes", "Annual leave").strip() or "Annual leave"
    one_day_hours, one_day_days = normalize_holiday_amounts(
        user,
        request.form.get("holiday_hours"),
        request.form.get("holiday_days")
    )

    conn = get_db()
    current = start_dt
    count = 0
    skipped = 0

    while current <= end_dt:
        date_str = current.isoformat()

        existing = conn.execute("""
            SELECT status FROM shift_calendar
            WHERE user_id = ? AND date = ?
            LIMIT 1
        """, (user["id"], date_str)).fetchone()

        if existing and existing["status"] == "Off" and not include_off_days:
            skipped += 1
            current += timedelta(days=1)
            continue

        conn.execute("""
            INSERT INTO shift_calendar
            (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at, holiday_hours, holiday_days)
            VALUES (?, ?, 'Holiday', 'Annual Leave', '', '', ?, 'Manual', ?, ?, ?)
            ON CONFLICT(user_id, date) DO UPDATE SET
                status = 'Holiday',
                shift_name = 'Annual Leave',
                start_time = '',
                end_time = '',
                notes = excluded.notes,
                source = 'Manual',
                holiday_hours = excluded.holiday_hours,
                holiday_days = excluded.holiday_days
        """, (user["id"], date_str, notes, datetime.now().isoformat(), one_day_hours, one_day_days))
        count += 1
        current += timedelta(days=1)

    conn.commit()
    conn.close()

    if skipped:
        flash(f"{count} holiday day(s) added. {skipped} off day(s) skipped.", "success")
    else:
        flash(f"{count} holiday day(s) added.", "success")
    return redirect(url_for("holiday_tracker"))

@app.route("/")
@login_required
def index():
    user = current_user()
    conn = get_db()
    recent_mileage = conn.execute("SELECT * FROM mileage WHERE user_id = ? ORDER BY date DESC, id DESC LIMIT 5", (user["id"],)).fetchall()
    recent_yard = conn.execute("SELECT * FROM yard_checks WHERE user_id = ? ORDER BY date DESC, id DESC LIMIT 5", (user["id"],)).fetchall()
    conn.close()
    today = datetime.today().date()
    wanted_days = [today - timedelta(days=1), today, today + timedelta(days=1)]
    day_start = wanted_days[0].isoformat()
    day_end = wanted_days[-1].isoformat()
    conn = get_db()
    day_rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date BETWEEN ? AND ?
        ORDER BY date ASC
    """, (user["id"], day_start, day_end)).fetchall()
    conn.close()
    by_date = {row["date"]: row for row in day_rows}
    dashboard_days = []
    for label, day in [("Yesterday", wanted_days[0]), ("Today", wanted_days[1]), ("Tomorrow", wanted_days[2])]:
        key = day.isoformat()
        row = by_date.get(key)
        dashboard_days.append(row if row else {
            "date": key,
            "label": label,
            "status": "Not Set",
            "shift_name": "",
            "start_time": "",
            "end_time": "",
            "notes": "",
        })
        if row:
            dashboard_days[-1] = dict(row)
            dashboard_days[-1]["label"] = label

    return render_template(
        "dashboard.html",
        page="dashboard",
        totals=totals(user["id"]),
        recent_mileage=recent_mileage,
        recent_yard=recent_yard,
        dashboard_days=dashboard_days,
        today_shift=today_shift_status(user["id"]),
        annual_leave=annual_leave_summary(user["id"]),
        favorite_tools=get_favorite_tools(user)
    )





@app.route("/favourites", methods=["GET", "POST"])
@login_required
def favourites():
    user = current_user()
    options = available_favorite_options(user)
    allowed = [x[0] for x in options]

    if request.method == "POST":
        selected = []
        for key in request.form.getlist("favorite_tools"):
            key = key.strip()
            if key in allowed and key not in selected:
                selected.append(key)
            if len(selected) >= 4:
                break
        for key in ["morning_brief", "shift_calendar", "handover", "yard_check"]:
            if len(selected) >= 4:
                break
            if key in allowed and key not in selected:
                selected.append(key)
        conn = get_db()
        conn.execute("UPDATE users SET favorite_tools=? WHERE id=?", (",".join(selected[:4]), user["id"]))
        conn.commit()
        conn.close()
        flash("Favourites saved.", "success")
        return redirect(url_for("favourites"))

    selected_tools = get_favorite_tools(user)
    return render_template("favourites.html", page="favourites", options=options, selected_tools=selected_tools, selected_keys=[x[0] for x in selected_tools])

@app.route("/account/export-data")
@login_required
def export_account_data():
    """Download a complete account backup as JSON files inside a ZIP."""
    user = current_user()
    tables = [
        "mileage", "expenses", "yard_checks", "handovers", "team_members",
        "shift_calendar", "daily_shift_logs", "action_tracker", "absence_records",
        "evidence_library", "holiday_settings"
    ]
    conn = get_db()
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        profile = {k: user[k] for k in user.keys() if k not in ["password_hash"]}
        zf.writestr("profile.json", json.dumps(profile, indent=2, default=str))
        for table in tables:
            try:
                rows = conn.execute(f"SELECT * FROM {table} WHERE user_id=? ORDER BY id DESC", (user["id"],)).fetchall()
                data = [{k: row[k] for k in row.keys()} for row in rows]
                zf.writestr(f"{table}.json", json.dumps(data, indent=2, default=str))
            except Exception:
                zf.writestr(f"{table}.json", "[]")
    conn.close()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="whs-account-backup.zip", mimetype="application/zip")



def send_welcome_email(name, email):
    """Send a WHS multilingual welcome email after registration.
    Uses SMTP environment variables when configured; fails safely so registration still works.
    """
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SMTP_FROM]):
        print("WHS welcome email skipped: SMTP is not configured.")
        return False

    safe_name = (name or "there").strip() or "there"
    try:
        app_url = os.environ.get("APP_URL") or request.url_root.rstrip("/")
    except Exception:
        app_url = os.environ.get("APP_URL", "")

    subject = "Welcome to WHS – Warehouse Support"

    plain_text = f"""Hi {safe_name},

Welcome to WHS – Warehouse Support.

Your account has been successfully created and you can now access the platform.

WHS helps warehouse managers and supervisors with:
- Morning Briefs
- Yard Checks
- Shift Handovers
- Holiday Tracking
- Team Management
- Calendar & Planning
- Company Themes & Personalisation

To get started:
1. Complete your Account settings.
2. Select your preferred company theme.
3. Configure your Holiday Tracker.
4. Add your Team Members.
5. Create your first Morning Brief and Handover.

Open WHS: {app_url}

Thank you for being part of WHS.

Kind regards,
The WHS Team
Warehouse Support

---

HU: Üdvözlünk a WHS-ben. A fiókod elkészült, és már használhatod a napi raktári munkához szükséges eszközöket.
PL: Witamy w WHS. Twoje konto zostało utworzone i możesz korzystać z narzędzi do codziennej pracy magazynowej.
RO: Bun venit la WHS. Contul tău a fost creat și poți folosi instrumentele pentru activitatea zilnică din depozit.
ES: Bienvenido a WHS. Tu cuenta se ha creado correctamente y ya puedes usar las herramientas para operaciones diarias de almacén.
DE: Willkommen bei WHS. Dein Konto wurde erstellt und du kannst die Tools für den täglichen Lagerbetrieb nutzen.
"""

    language_sections = [
        ("English", f"Hi {safe_name}, welcome to WHS – Warehouse Support. Your account is active and ready to use."),
        ("Magyar", f"Szia {safe_name}, üdvözlünk a WHS – Warehouse Support alkalmazásban. A fiókod aktív és használatra kész."),
        ("Polski", f"Cześć {safe_name}, witamy w WHS – Warehouse Support. Twoje konto jest aktywne i gotowe do użycia."),
        ("Română", f"Salut {safe_name}, bun venit la WHS – Warehouse Support. Contul tău este activ și gata de utilizare."),
        ("Español", f"Hola {safe_name}, bienvenido a WHS – Warehouse Support. Tu cuenta está activa y lista para usar."),
        ("Deutsch", f"Hallo {safe_name}, willkommen bei WHS – Warehouse Support. Dein Konto ist aktiv und einsatzbereit."),
    ]

    tools = [
        "Morning Briefs", "Yard Checks", "Shift Handovers", "Holiday Tracking",
        "Team Management", "Calendar & Planning", "Company Themes & Personalisation"
    ]
    steps = [
        "Complete your Account settings.",
        "Select your preferred company theme.",
        "Configure your Holiday Tracker.",
        "Add your Team Members.",
        "Create your first Morning Brief and Handover."
    ]

    html_sections = "".join(
        f"""
        <div style=\"padding:14px 16px;border:1px solid #dbe7f3;border-radius:14px;margin:10px 0;background:#ffffff;\">
          <strong style=\"display:block;color:#0f172a;font-size:15px;margin-bottom:6px;\">{lang}</strong>
          <span style=\"color:#475569;font-size:14px;line-height:1.5;\">{text}</span>
        </div>
        """ for lang, text in language_sections
    )
    tools_html = "".join(f"<li style='margin:7px 0;'>✅ {tool}</li>" for tool in tools)
    steps_html = "".join(f"<li style='margin:7px 0;'>{step}</li>" for step in steps)

    html_body = f"""
<!doctype html>
<html>
  <body style=\"margin:0;padding:0;background:#eef4fb;font-family:Arial,Helvetica,sans-serif;color:#0f172a;\">
    <div style=\"max-width:720px;margin:0 auto;padding:28px 14px;\">
      <div style=\"background:linear-gradient(135deg,#071827,#009fb7);border-radius:28px;padding:28px;color:#ffffff;box-shadow:0 18px 40px rgba(15,23,42,.18);\">
        <div style=\"display:flex;align-items:center;gap:14px;margin-bottom:22px;\">
          <div style=\"width:58px;height:58px;border-radius:18px;background:linear-gradient(135deg,#18d3c2,#2563eb);display:flex;align-items:center;justify-content:center;font-weight:900;letter-spacing:.5px;\">WHS</div>
          <div>
            <div style=\"font-size:24px;font-weight:900;line-height:1.1;\">Welcome to WHS</div>
            <div style=\"opacity:.85;font-size:14px;margin-top:4px;\">Warehouse Support</div>
          </div>
        </div>
        <h1 style=\"font-size:34px;line-height:1.1;margin:0 0 12px 0;\">Your account is ready</h1>
        <p style=\"font-size:16px;line-height:1.6;margin:0;color:#dbeafe;\">Thank you for registering. WHS is built to help warehouse and logistics leaders manage daily work faster, cleaner and more consistently.</p>
        <div style=\"margin-top:24px;\">
          <a href=\"{app_url}\" style=\"background:#ffffff;color:#0f172a;text-decoration:none;font-weight:900;padding:14px 22px;border-radius:14px;display:inline-block;\">Open WHS</a>
        </div>
      </div>

      <div style=\"background:#ffffff;border-radius:24px;padding:26px;margin-top:18px;border:1px solid #dbe7f3;\">
        <h2 style=\"margin:0 0 10px 0;font-size:22px;\">Hi {safe_name},</h2>
        <p style=\"color:#475569;line-height:1.7;margin:0 0 14px 0;\">Your account has been successfully created and you currently have access to the WHS tools.</p>
        <h3 style=\"margin:20px 0 8px 0;font-size:17px;\">What you can use</h3>
        <ul style=\"padding-left:20px;color:#334155;line-height:1.6;margin-top:8px;\">{tools_html}</ul>
        <h3 style=\"margin:20px 0 8px 0;font-size:17px;\">Recommended first steps</h3>
        <ol style=\"padding-left:22px;color:#334155;line-height:1.6;margin-top:8px;\">{steps_html}</ol>
      </div>

      <div style=\"background:#f8fafc;border-radius:24px;padding:22px;margin-top:18px;border:1px solid #dbe7f3;\">
        <h2 style=\"font-size:20px;margin:0 0 10px 0;\">Welcome in all supported languages</h2>
        {html_sections}
      </div>

      <p style=\"text-align:center;color:#64748b;font-size:13px;line-height:1.6;margin:20px 0 0 0;\">
        Kind regards,<br><strong>The WHS Team</strong><br>Warehouse Support
      </p>
    </div>
  </body>
</html>
"""

    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM
        msg["To"] = email
        msg.set_content(plain_text)
        msg.add_alternative(html_body, subtype="html")
        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ssl.create_default_context()) as server:
                server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
                if SMTP_USE_TLS:
                    server.starttls(context=ssl.create_default_context())
                server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)
        print(f"WHS welcome email sent to {email}")
        return True
    except Exception as exc:
        print(f"WHS welcome email failed: {exc}")
        return False

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not name or not email or len(password) < 6:
            flash("Please enter name, email and a password with at least 6 characters.", "error")
            return redirect(url_for("register"))

        conn = get_db()
        try:
            cur = conn.execute("""
                INSERT INTO users (name, email, password_hash, plan, business_name, created_at, language, favorite_tools, annual_leave_entitlement)
                VALUES (?, ?, ?, 'free', ?, ?, ?, ?, 28)
            """, (name, email, generate_password_hash(password), "Warehouse / Site", datetime.now().isoformat(), session.get("language", "en"), "morning_brief,shift_calendar,handover,yard_check"))
            conn.commit()
            session["user_id"] = cur.lastrowid
            conn.execute("UPDATE users SET last_login_at=? WHERE id=?", (datetime.now().isoformat(timespec="seconds"), cur.lastrowid))
            conn.commit()
            # Email sending is disabled in this version.
            # The account is created immediately without SMTP configuration.
            flash("Account created successfully.", "success")
        except sqlite3.IntegrityError:
            flash("Email already registered.", "error")
            conn.close()
            return redirect(url_for("register"))
        conn.close()
        return redirect(url_for("index"))
    return render_template("auth.html", mode="register", page="auth")




@app.route("/og-test")
def og_test():
    return """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>WHS - Operations Management Platform</title>
<meta property="og:title" content="WHS - Operations Management Platform">
<meta property="og:description" content="Warehouse Support for shifts, handovers, mileage, expenses, yard checks and team records.">
<meta property="og:image" content="https://whs-6ozo.onrender.com/static/images/whs-share.png?v=16">
<meta property="og:image:secure_url" content="https://whs-6ozo.onrender.com/static/images/whs-share.png?v=16">
<meta property="og:image:type" content="image/png">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:url" content="https://whs-6ozo.onrender.com/">
<meta property="og:type" content="website">
<meta property="og:site_name" content="WHS">
</head>
<body>
<h1>WHS</h1>
<img src="/static/images/whs-share.png?v=16" style="max-width:600px;width:100%;">
</body>
</html>
"""


@app.route("/share")
def share_preview():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            conn = get_db()
            conn.execute("UPDATE users SET last_login_at=? WHERE id=?", (datetime.now().isoformat(timespec="seconds"), user["id"]))
            conn.commit(); conn.close()
            session["user_id"] = user["id"]
            response = redirect(url_for("index"))
            if request.form.get("remember_me") == "yes":
                token, expires = create_remember_token(user["id"])
                response.set_cookie(
                    "whs_remember",
                    token,
                    max_age=60 * 60 * 24 * 30,
                    httponly=True,
                    secure=True,
                    samesite="Lax"
                )
            return response
        flash("Invalid login details.", "error")
    return render_template("auth.html", mode="login", page="auth")


@app.route("/logout")
def logout():
    session.clear()
    response = redirect(url_for("login"))
    response.delete_cookie("whs_remember")
    return response



@app.route("/shift-calendar", methods=["GET", "POST"])
@login_required
def shift_calendar():
    user = current_user()
    trial = shift_calendar_trial_info(user)
    if not trial["allowed"]:
        flash("Your 14 day free Shift Calendar trial has ended. Upgrade to Pro to continue using Shift Calendar.", "error")
        return redirect(url_for("pricing"))

    statuses = ["Work", "Off", "Holiday", "Sick", "Training", "Overtime", "Bank Holiday", "Custom"]
    patterns = [
        ("4on4off", "4 on / 4 off"),
        ("5on2off", "5 days / 2 off"),
        ("monfri", "Monday-Friday"),
        ("2days2nights4off", "2 days / 2 nights / 4 off"),
        ("3on4off4on3off", "3 on / 4 off / 4 on / 3 off"),
        ("custom", "Custom pattern"),
    ]

    if request.method == "POST":
        action = request.form.get("action")

        conn = get_db()

        if action == "generate":
            start_date = request.form.get("start_date") or datetime.today().strftime("%Y-%m-%d")
            pattern = request.form.get("pattern", "4on4off")
            months = int(request.form.get("months") or 12)
            shift_name = request.form.get("shift_name", "Shift").strip()
            start_time = request.form.get("start_time", "06:00")
            end_time = request.form.get("end_time", "18:00")
            replace_existing = request.form.get("replace_existing") == "yes"

            rows = generate_shift_pattern_dates(start_date, pattern, months, shift_name, start_time, end_time, request.form.get("custom_pattern", ""))

            if replace_existing:
                conn.execute("DELETE FROM shift_calendar WHERE user_id = ? AND date >= ?", (user["id"], start_date))

            for row in rows:
                conn.execute("""
                    INSERT INTO shift_calendar
                    (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at, holiday_hours, holiday_days)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(user_id, date) DO UPDATE SET
                        status = excluded.status,
                        shift_name = excluded.shift_name,
                        start_time = excluded.start_time,
                        end_time = excluded.end_time,
                        notes = excluded.notes,
                        source = excluded.source
                """, (
                    user["id"], row["date"], row["status"], row["shift_name"], row["start_time"],
                    row["end_time"], row["notes"], row["source"], datetime.now().isoformat(), 0, 0
                ))

            conn.commit()
            conn.close()
            flash(f"Shift calendar generated for {months} month(s).", "success")
            return redirect(url_for("shift_calendar"))

        if action == "manual":
            date = request.form.get("manual_date") or datetime.today().strftime("%Y-%m-%d")
            status = request.form.get("manual_status", "Work")
            shift_name = request.form.get("manual_shift_name", "").strip()
            start_time = request.form.get("manual_start_time", "")
            end_time = request.form.get("manual_end_time", "")
            notes = request.form.get("manual_notes", "").strip()

            holiday_hours = 0
            holiday_days = 0
            if status == "Holiday":
                notes = notes or "Annual leave"
                holiday_hours, holiday_days = normalize_holiday_amounts(
                    user,
                    request.form.get("manual_holiday_hours"),
                    request.form.get("manual_holiday_days")
                )

            conn.execute("""
                INSERT INTO shift_calendar
                (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at, holiday_hours, holiday_days)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'Manual', ?, ?, ?)
                ON CONFLICT(user_id, date) DO UPDATE SET
                    status = excluded.status,
                    shift_name = excluded.shift_name,
                    start_time = excluded.start_time,
                    end_time = excluded.end_time,
                    notes = excluded.notes,
                    holiday_hours = excluded.holiday_hours,
                    holiday_days = excluded.holiday_days,
                    source = 'Manual'
            """, (user["id"], date, status, shift_name, start_time, end_time, notes, datetime.now().isoformat(), holiday_hours, holiday_days))

            conn.commit()
            conn.close()
            flash("Shift day updated.", "success")
            return redirect(url_for("shift_calendar"))

    view = request.args.get("view", "month")
    start = request.args.get("start")
    month_arg = request.args.get("month")

    if view == "list":
        try:
            start_date = datetime.strptime(start, "%Y-%m-%d").date() if start else get_week_start()
        except Exception:
            start_date = get_week_start()
        end_date = start_date + timedelta(days=27)
        prev_start = (start_date - timedelta(days=28)).isoformat()
        next_start = (start_date + timedelta(days=28)).isoformat()
        month_date = start_date.replace(day=1)
    else:
        view = "month"
        try:
            month_date = datetime.strptime(month_arg, "%Y-%m").date().replace(day=1) if month_arg else datetime.today().date().replace(day=1)
        except Exception:
            month_date = datetime.today().date().replace(day=1)
        start_date = month_date
        last_day = calendar.monthrange(month_date.year, month_date.month)[1]
        end_date = month_date.replace(day=last_day)
        prev_month_raw = (month_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        next_month_raw = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        prev_start = prev_month_raw.strftime("%Y-%m")
        next_start = next_month_raw.strftime("%Y-%m")

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date BETWEEN ? AND ?
        ORDER BY date ASC
    """, (user["id"], start_date.isoformat(), end_date.isoformat())).fetchall()
    conn.close()

    by_date = {row["date"]: row for row in rows}

    def empty_day(day):
        return {
            "date": day.isoformat(),
            "status": "Not Set",
            "shift_name": "",
            "start_time": "",
            "end_time": "",
            "notes": "",
            "source": "Empty",
        }

    calendar_days = []
    d = start_date
    while d <= end_date:
        key = d.isoformat()
        calendar_days.append(by_date.get(key) or empty_day(d))
        d += timedelta(days=1)

    month_weeks = []
    if view == "month":
        month_calendar = calendar.Calendar(firstweekday=0).monthdatescalendar(month_date.year, month_date.month)
        for week in month_calendar:
            week_items = []
            for day in week:
                item = by_date.get(day.isoformat()) or empty_day(day)
                item = dict(item)
                item["in_month"] = day.month == month_date.month
                week_items.append(item)
            month_weeks.append(week_items)

    return render_template(
        "shift_calendar.html",
        rows=calendar_days,
        month_weeks=month_weeks,
        view=view,
        current_month=month_date.strftime("%B %Y"),
        current_month_value=month_date.strftime("%Y-%m"),
        statuses=statuses,
        patterns=patterns,
        page="shift_calendar",
        prev_start=prev_start,
        next_start=next_start,
        status_colors=SHIFT_STATUS_COLORS,
        trial=shift_calendar_trial_info(user),
        today_iso=datetime.today().date().isoformat(),
    )


@app.route("/shift-calendar/export")
@login_required
@plan_required("pro")
def export_shift_calendar():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ?
        ORDER BY date ASC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Calendar"
    headers = ["Date", "Status", "Shift", "Start", "End", "Notes", "Source"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row["date"], row["status"], row["shift_name"], row["start_time"],
            row["end_time"], row["notes"], row["source"]
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="shift-calendar.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/pricing")
@login_required
def pricing():
    return render_template("pricing.html", page="pricing")


@app.route("/set-plan/<plan>")
@login_required
def set_plan(plan):
    if plan not in PLAN_ORDER:
        flash("Invalid plan.", "error")
        return redirect(url_for("pricing"))
    user = current_user()
    conn = get_db()
    conn.execute("UPDATE users SET plan = ? WHERE id = ?", (plan, user["id"]))
    conn.commit()
    conn.close()
    flash(f"Demo plan changed to {PLAN_NAMES[plan]}. For real payments use Stripe checkout setup.", "success")
    return redirect(url_for("pricing"))



@app.route("/api/calculate-mileage", methods=["POST"])
@login_required
def api_calculate_mileage():
    data = request.get_json(silent=True) or {}
    origin = (data.get("origin") or "").strip()
    destination = (data.get("destination") or "").strip()

    if not origin or not destination:
        return {"ok": False, "error": "Please enter From and To postcode/address."}, 400

    result, error = calculate_google_maps_miles(origin, destination)
    if error:
        return {"ok": False, "error": error}, 400

    return {"ok": True, "result": result}


@app.route("/mileage", methods=["GET", "POST"])
@login_required
def mileage():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        start_location = request.form.get("start_location", "").strip()
        end_location = request.form.get("end_location", "").strip()
        purpose = request.form.get("purpose", "").strip()
        miles = float(request.form.get("miles") or 0)
        rate = float(request.form.get("rate") or user["mileage_rate"] or HMRC_MILE_RATE)
        if not start_location or not end_location or miles <= 0 or rate <= 0:
            flash("Please enter start, destination, miles and valid rate.", "error")
            return redirect(url_for("mileage"))
        conn = get_db()
        conn.execute("""
            INSERT INTO mileage (user_id, date, start_location, end_location, miles, rate, purpose, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, start_location, end_location, miles, rate, purpose, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Mileage added.", "success")
        return redirect(url_for("mileage"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM mileage WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("mileage.html", rows=rows, rate=HMRC_MILE_RATE, default_rate=(user["mileage_rate"] or HMRC_MILE_RATE), page="mileage")



@app.route("/mileage/export")
@login_required
def export_mileage():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM mileage
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mileage"

    headers = ["Date", "From", "To", "Purpose", "Miles", "Rate (£/mile)", "Claim (£)", "Recorded At"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    total_miles = 0
    total_claim = 0

    for row in rows:
        rate = row["rate"] if "rate" in row.keys() and row["rate"] else HMRC_MILE_RATE
        claim = float(row["miles"] or 0) * float(rate or 0)
        total_miles += float(row["miles"] or 0)
        total_claim += claim

        ws.append([
            row["date"],
            row["start_location"],
            row["end_location"],
            row["purpose"] or "",
            float(row["miles"] or 0),
            float(rate or 0),
            round(claim, 2),
            row["created_at"]
        ])

    ws.append([])
    ws.append(["TOTAL", "", "", "", round(total_miles, 2), "", round(total_claim, 2), ""])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"mileage-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():
    user = current_user()
    categories = ["Fuel", "Parking", "Tolls", "Vehicle Maintenance", "Phone", "Equipment", "Insurance", "Office", "Other"]
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        category = request.form.get("category", "Other")
        description = request.form.get("description", "").strip()
        amount = float(request.form.get("amount") or 0)
        if not description or amount <= 0:
            flash("Please enter description and amount.", "error")
            return redirect(url_for("expenses"))
        conn = get_db()
        conn.execute("""
            INSERT INTO expenses (user_id, date, category, description, amount, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user["id"], date, category, description, amount, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Expense added.", "success")
        return redirect(url_for("expenses"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM expenses WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("expenses.html", rows=rows, categories=categories, page="expenses")



@app.route("/expenses/export")
@login_required
def export_expenses():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM expenses
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Category", "Description", "Amount (£)", "Recorded At"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    total_amount = 0

    for row in rows:
        amount = float(row["amount"] or 0)
        total_amount += amount
        ws.append([
            row["date"],
            row["category"],
            row["description"],
            round(amount, 2),
            row["created_at"]
        ])

    ws.append([])
    ws.append(["TOTAL", "", "", round(total_amount, 2), ""])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"expenses-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/invoices", methods=["GET", "POST"])
@login_required
def invoices():
    user = current_user()
    if request.method == "POST":
        invoice_number = request.form.get("invoice_number") or next_invoice_number(user["id"])
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        customer_name = request.form.get("customer_name", "").strip()
        customer_email = request.form.get("customer_email", "").strip()
        description = request.form.get("description", "").strip()
        amount = float(request.form.get("amount") or 0)
        status = request.form.get("status", "Unpaid")
        if not customer_name or not description or amount <= 0:
            flash("Please enter customer, description and amount.", "error")
            return redirect(url_for("invoices"))
        conn = get_db()
        conn.execute("""
            INSERT INTO invoices (user_id, invoice_number, date, customer_name, customer_email, description, amount, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], invoice_number, date, customer_name, customer_email, description, amount, status, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Invoice created.", "success")
        return redirect(url_for("invoices"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM invoices WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("invoices.html", rows=rows, invoice_number=next_invoice_number(user["id"]), page="invoices")


@app.route("/invoice/<int:item_id>/pdf")
@login_required
def invoice_pdf(item_id):
    user = current_user()
    conn = get_db()
    invoice = conn.execute("SELECT * FROM invoices WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()
    if not invoice:
        flash("Invoice not found.", "error")
        return redirect(url_for("invoices"))

    buffer = invoice_pdf_buffer(user, invoice)
    return send_file(buffer, as_attachment=True, download_name=f"{invoice['invoice_number']}.pdf", mimetype="application/pdf")


@app.route("/invoice/<int:item_id>/email")
@login_required
def invoice_email(item_id):
    user = current_user()
    conn = get_db()
    invoice = conn.execute("SELECT * FROM invoices WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    if not invoice:
        conn.close()
        flash("Invoice not found.", "error")
        return redirect(url_for("invoices"))

    success, message = send_invoice_email(user, invoice)
    if success:
        conn.execute("UPDATE invoices SET email_sent = 1 WHERE id = ? AND user_id = ?", (item_id, user["id"]))
        conn.commit()
        flash(message, "success")
    else:
        flash(message, "error")
    conn.close()
    return redirect(url_for("invoices"))


@app.route("/tax")
@login_required
def tax():
    user = current_user()
    return render_template("tax.html", totals=totals(user["id"]), page="tax")



@app.route("/morning-brief", methods=["GET", "POST"])
@login_required
def morning_brief():
    user = current_user()

    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        manager = request.form.get("manager", "").strip()
        role = request.form.get("role", "").strip()
        volume = int(request.form.get("volume") or 0)
        available_hc = int(request.form.get("available_hc") or 0)
        late_trailers = 0
        headers = request.form.getlist("section_header[]")
        values = request.form.getlist("section_value[]")
        custom_sections = []
        for i in range(max(len(headers), len(values))):
            header = (headers[i] if i < len(headers) else "").strip()
            value = (values[i] if i < len(values) else "").strip()
            if header or value:
                custom_sections.append({"header": header, "value": value})

        # Keep these columns populated for compatibility with older saved records.
        safety_message = custom_sections[0]["value"] if len(custom_sections) > 0 else ""
        priorities = custom_sections[1]["value"] if len(custom_sections) > 1 else ""
        team_messages = custom_sections[2]["value"] if len(custom_sections) > 2 else ""
        break_reminder = custom_sections[3]["value"] if len(custom_sections) > 3 else ""
        equipment_reminder = custom_sections[4]["value"] if len(custom_sections) > 4 else ""

        generated = generate_morning_brief_text(
            date, shift, manager, role, volume, available_hc, late_trailers,
            safety_message, priorities, team_messages, break_reminder, equipment_reminder,
            custom_sections=custom_sections
        )

        conn = get_db()
        conn.execute("""
            INSERT INTO morning_briefs
            (user_id, date, shift, manager, role, volume, available_hc, late_trailers,
             safety_message, priorities, team_messages, break_reminder, equipment_reminder, generated_brief, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user["id"], date, shift, manager, role, volume, available_hc, late_trailers,
            safety_message, priorities, team_messages, break_reminder, equipment_reminder,
            generated, datetime.now().isoformat()
        ))
        conn.commit()
        conn.close()

        flash("Morning brief generated and saved.", "success")
        return redirect(url_for("morning_brief"))

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM morning_briefs
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
        LIMIT 50
    """, (user["id"],)).fetchall()
    conn.close()

    return render_template("morning_brief.html", rows=rows, page="morning_brief")


@app.route("/morning-brief/<int:item_id>/download")
@login_required
def morning_brief_download(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM morning_briefs WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()

    if not row:
        flash("Morning brief not found.", "error")
        return redirect(url_for("morning_brief"))

    output = BytesIO()
    output.write(row["generated_brief"].encode("utf-8"))
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f"morning-brief-{row['date']}.txt", mimetype="text/plain")


@app.route("/yard-check", methods=["GET", "POST"])
@login_required
def yard_check():
    user = current_user()

    yard_cfg = get_yard_config(user)
    door_start = yard_cfg["door_start"]
    door_end = yard_cfg["door_end"]
    fence_start = yard_cfg["fence_start"]
    fence_end = yard_cfg["fence_end"]

    door_options = [f"Door {i}" for i in range(door_start, door_end + 1)]
    fence_options = [f"Fence {i}" for i in range(fence_start, fence_end + 1)]
    custom_locations = get_custom_locations(user["id"])
    custom_location_names = [r["name"] for r in custom_locations]
    locations = ["Door", "Fence", "Yard", "Loading Bay", "Workshop"] + custom_location_names + ["Other"]
    statuses = ["Recorded", "Checked", "Issue Found", "Missing", "Moved", "Loaded", "Empty"]

    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        entry_mode = request.form.get("entry_mode", "single")
        source = request.form.get("source", "Manual")
        status = request.form.get("status", "Recorded")
        notes = request.form.get("notes", "").strip()

        saved_count = 0
        conn = get_db()

        if entry_mode == "matrix":
            yard_photo_filename = None
            yard_photo = request.files.get("photo") or request.files.get("camera_photo")
            if yard_photo and yard_photo.filename and allowed_image(yard_photo.filename):
                safe = secure_filename(yard_photo.filename)
                yard_photo_filename = f"{uuid.uuid4().hex}_{safe}"
                yard_photo.save(os.path.join(UPLOAD_DIR, yard_photo_filename))

            row_locations = request.form.getlist("row_location[]")
            trailer_ids = request.form.getlist("row_trailer_id[]")
            row_notes = request.form.getlist("row_notes[]")
            markers = request.form.getlist("row_marker[]")
            row_sources = request.form.getlist("row_source[]")

            max_rows = max(len(row_locations), len(trailer_ids), len(row_notes), len(markers), len(row_sources), 0)

            for i in range(max_rows):
                location_text = (row_locations[i] if i < len(row_locations) else "").strip()
                trailer_id = (trailer_ids[i] if i < len(trailer_ids) else "").strip().upper()
                note_value = (row_notes[i] if i < len(row_notes) else "").strip()
                marker = (markers[i] if i < len(markers) else "").strip()
                row_source = (row_sources[i] if i < len(row_sources) else source).strip() or source

                # Required fields: Location and Trailer ID
                if not location_text or not trailer_id:
                    continue

                location_type = "Other"
                if location_text.lower().startswith("door"):
                    location_type = "Door"
                elif location_text.lower().startswith("fence"):
                    location_type = "Fence"
                elif location_text.lower().startswith("yard"):
                    location_type = "Yard"

                combined_notes = []
                if note_value:
                    combined_notes.append(f"Notes: {note_value}")
                if marker:
                    combined_notes.append(f"Marker: {marker}")
                combined_notes = " | ".join(combined_notes)

                conn.execute("""
                    INSERT INTO yard_checks
                    (user_id, date, trailer_id, location_type, location_detail, status, notes, source, photo_filename, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (user["id"], date, trailer_id, location_type, location_text, "Recorded", combined_notes, row_source, yard_photo_filename, datetime.now().isoformat()))
                saved_count += 1

        elif entry_mode == "batch":
            batch_text = request.form.get("batch_text", "").strip()
            lines = [line.strip() for line in batch_text.splitlines() if line.strip()]

            for line in lines:
                original = line
                clean = line.replace("|", ",").replace(" at ", ",").replace(" AT ", ",")
                parts = [p.strip() for p in clean.split(",") if p.strip()]

                trailer_id = parts[0].upper() if parts else ""
                location_type = "Yard"
                location_detail = ""
                line_notes = notes or original

                if len(parts) >= 2:
                    loc = parts[1].title()
                    if loc.startswith("Door"):
                        location_type = "Door"
                        location_detail = loc
                    elif loc.startswith("Fence"):
                        location_type = "Fence"
                        location_detail = loc
                    else:
                        location_type = "Other"
                        location_detail = loc

                if len(parts) >= 3:
                    line_notes = parts[2]

                if trailer_id:
                    conn.execute("""
                        INSERT INTO yard_checks
                        (user_id, date, trailer_id, location_type, location_detail, status, notes, source, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (user["id"], date, trailer_id, location_type, location_detail, status, line_notes, source, datetime.now().isoformat()))
                    saved_count += 1

        else:
            yard_photo_filename = None
            yard_photo = request.files.get("photo") or request.files.get("camera_photo")
            if yard_photo and yard_photo.filename:
                if allowed_image(yard_photo.filename):
                    safe = secure_filename(yard_photo.filename)
                    yard_photo_filename = f"{uuid.uuid4().hex}_{safe}"
                    yard_photo.save(os.path.join(UPLOAD_DIR, yard_photo_filename))

            trailer_id = request.form.get("trailer_id", "").strip().upper()
            location_type = request.form.get("location_type", "Yard")
            location_detail = request.form.get("location_detail", "").strip()
            custom_location = request.form.get("custom_location", "").strip()

            if location_type == "Door":
                location_detail = request.form.get("door_number", location_detail)
            elif location_type == "Fence":
                location_detail = request.form.get("fence_number", location_detail)
            elif location_type == "Other" and custom_location:
                location_detail = custom_location
            elif location_type not in ["Door", "Fence", "Yard", "Loading Bay", "Workshop", "Other"]:
                location_detail = location_detail or location_type

            if not trailer_id:
                conn.close()
                flash("Please enter trailer ID.", "error")
                return redirect(url_for("yard_check"))

            conn.execute("""
                INSERT INTO yard_checks
                (user_id, date, trailer_id, location_type, location_detail, status, notes, source, photo_filename, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (user["id"], date, trailer_id, location_type, location_detail, status, notes, source, yard_photo_filename, datetime.now().isoformat()))
            saved_count = 1

        conn.commit()
        conn.close()

        flash(f"{saved_count} trailer record(s) saved. Empty Trailer ID rows were ignored automatically.", "success")
        return redirect(url_for("yard_check"))

    search = request.args.get("search", "").strip()
    location_filter = request.args.get("location", "").strip()
    status_filter = request.args.get("status", "").strip()

    query = """
        SELECT * FROM yard_checks
        WHERE user_id = ?
    """
    params = [user["id"]]

    if search:
        like = f"%{search}%"
        query += """
            AND (
                trailer_id LIKE ?
                OR location_type LIKE ?
                OR location_detail LIKE ?
                OR status LIKE ?
                OR notes LIKE ?
                OR source LIKE ?
            )
        """
        params.extend([like, like, like, like, like, like])

    if location_filter:
        query += " AND location_type = ?"
        params.append(location_filter)

    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)

    query += " ORDER BY date DESC, id DESC LIMIT 500"

    conn = get_db()
    rows = conn.execute(query, params).fetchall()

    today_count = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND date = ?
    """, (user["id"], datetime.today().strftime("%Y-%m-%d"))).fetchone()[0]

    door_used = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND location_type = 'Door'
    """, (user["id"],)).fetchone()[0]

    fence_used = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND location_type = 'Fence'
    """, (user["id"],)).fetchone()[0]

    conn.close()

    return render_template(
        "yard_check.html",
        rows=rows,
        locations=locations,
        statuses=statuses,
        custom_locations=custom_locations,
        door_options=door_options,
        fence_options=fence_options,
        door_start=door_start,
        door_end=door_end,
        fence_start=fence_start,
        fence_end=fence_end,
        yard_config_editable=yard_cfg["editable"],
        door_used=door_used,
        fence_used=fence_used,
        today_count=today_count,
        search=search,
        location_filter=location_filter,
        status_filter=status_filter,
        page="yard_check",
        today_iso=datetime.today().strftime("%Y-%m-%d")
    )


@app.route("/yard-check/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def edit_yard_check(item_id):
    user = current_user()
    statuses = ["Recorded", "Checked", "Issue Found", "Missing", "Moved", "Loaded", "Empty"]
    custom_locations = get_custom_locations(user["id"])
    locations = ["Door", "Fence", "Yard", "Loading Bay", "Workshop"] + [r["name"] for r in custom_locations] + ["Other"]

    conn = get_db()
    row = conn.execute("SELECT * FROM yard_checks WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()

    if not row:
        conn.close()
        flash("Yard check not found.", "error")
        return redirect(url_for("yard_check"))

    if request.method == "POST":
        trailer_id = request.form.get("trailer_id", "").strip().upper()
        if not trailer_id:
            conn.close()
            flash("Trailer ID is required.", "error")
            return redirect(url_for("edit_yard_check", item_id=item_id))

        conn.execute("""
            UPDATE yard_checks
            SET date = ?, trailer_id = ?, location_type = ?, location_detail = ?, status = ?, notes = ?, source = ?
            WHERE id = ? AND user_id = ?
        """, (
            request.form.get("date") or row["date"],
            trailer_id,
            request.form.get("location_type", "Yard"),
            request.form.get("location_detail", "").strip(),
            request.form.get("status", "Recorded"),
            request.form.get("notes", "").strip(),
            request.form.get("source", "Manual"),
            item_id,
            user["id"]
        ))
        conn.commit()
        conn.close()
        flash("Yard check updated.", "success")
        return redirect(url_for("yard_check"))

    conn.close()
    return render_template("yard_edit.html", row=row, statuses=statuses, locations=locations, page="yard_check")


@app.route("/yard-check/<int:item_id>/delete", methods=["POST"])
@login_required
@plan_required("pro")
def delete_yard_check(item_id):
    user = current_user()
    conn = get_db()
    conn.execute("DELETE FROM yard_checks WHERE id = ? AND user_id = ?", (item_id, user["id"]))
    conn.commit()
    conn.close()
    flash("Yard check deleted.", "success")
    return redirect(url_for("yard_check"))



@app.route("/yard-check/edit/<int:item_id>", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def edit_yard_check_old_url(item_id):
    return edit_yard_check(item_id)


@app.route("/yard-check/delete/<int:item_id>", methods=["POST"])
@login_required
@plan_required("pro")
def delete_yard_check_old_url(item_id):
    return delete_yard_check(item_id)


@app.route("/yard-check/export")
@login_required
@plan_required("pro")
def export_yard_check():
    user = current_user()
    conn = get_db()
    rows = conn.execute("SELECT * FROM yard_checks WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Yard Check"
    headers = ["Date", "Location", "Trailer ID", "Notes", "Custom Marker", "Source", "Recorded At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        trailer_id = (row["trailer_id"] or "").strip()
        # Empty trailer rows are draft/blank rows and should not be exported.
        if not trailer_id:
            continue
        raw_notes = row["notes"] or ""
        note_text = raw_notes
        marker_text = ""
        if "Marker:" in raw_notes:
            parts = [p.strip() for p in raw_notes.split("|")]
            note_parts = []
            for part in parts:
                if part.startswith("Marker:"):
                    marker_text = part.replace("Marker:", "", 1).strip()
                elif part.startswith("Notes:"):
                    note_parts.append(part.replace("Notes:", "", 1).strip())
                else:
                    note_parts.append(part)
            note_text = " | ".join([p for p in note_parts if p])
        ws.append([
            row["date"] or "",
            row["location_detail"] or row["location_type"] or "",
            trailer_id,
            note_text,
            marker_text,
            row["source"] or "",
            row["created_at"] or "",
        ])

    widths = [16, 22, 18, 35, 28, 16, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="yard-check.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/kpi", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def kpi_dashboard():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        volume = int(request.form.get("volume") or 0)
        planned_hc = int(request.form.get("planned_hc") or 0)
        actual_hc = int(request.form.get("actual_hc") or 0)
        target_rate = float(request.form.get("target_rate") or 0)
        actual_rate = float(request.form.get("actual_rate") or 0)
        late_trailers = 0
        errors = int(request.form.get("errors") or 0)
        notes = request.form.get("notes", "").strip()
        conn = get_db()
        conn.execute("""
            INSERT INTO kpi_records (user_id, date, shift, volume, planned_hc, actual_hc, target_rate, actual_rate, late_trailers, errors, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, volume, planned_hc, actual_hc, target_rate, actual_rate, late_trailers, errors, notes, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("KPI saved.", "success")
        return redirect(url_for("kpi_dashboard"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM kpi_records WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    summary = conn.execute("""
        SELECT COALESCE(SUM(volume), 0) total_volume, COALESCE(AVG(actual_rate), 0) avg_rate,
               COALESCE(SUM(late_trailers), 0) total_late, COALESCE(SUM(errors), 0) total_errors
        FROM kpi_records WHERE user_id = ?
    """, (user["id"],)).fetchone()
    conn.close()
    return render_template("kpi.html", rows=rows, summary=summary, page="kpi")



def default_handover_section_names():
    return {
        "attendance": "Attendance",
        "safety": "Safety Metrics",
        "operations": "Operations Pick",
        "sort": "CUK8 Sort Centre",
        "dispatch": "Dispatch",
        "suntory": "Suntory",
        "aob": "AOB"
    }

def get_handover_template(user_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM handover_templates WHERE user_id=? ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
    if not row:
        names = default_handover_section_names()
        conn.execute("""
            INSERT INTO handover_templates (user_id, name, section_names, force_vrid_uppercase, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (user_id, "Default Handover", json.dumps(names), 1, datetime.now().isoformat()))
        conn.commit()
        row = conn.execute("SELECT * FROM handover_templates WHERE user_id=? ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
    conn.close()
    return row

def parse_handover_template(row):
    """Return section configuration with default sections plus custom dynamic sections.

    section_names is kept backward compatible:
    {
      "attendance": "Attendance",
      ...
      "_enabled": {"attendance": true, ...},
      "_custom": [{"id": "custom_xxx", "name": "New section"}]
    }
    """
    names = default_handover_section_names()
    names["_enabled"] = {key: True for key in default_handover_section_names().keys()}
    names["_custom"] = []
    try:
        saved = json.loads(row_get(row, "section_names", "{}") or "{}")
        if isinstance(saved, dict):
            for key in default_handover_section_names().keys():
                if saved.get(key):
                    names[key] = saved.get(key)
            if isinstance(saved.get("_enabled"), dict):
                for key in default_handover_section_names().keys():
                    names["_enabled"][key] = bool(saved["_enabled"].get(key, True))
            if isinstance(saved.get("_custom"), list):
                custom = []
                for item in saved["_custom"]:
                    if isinstance(item, dict) and item.get("name"):
                        cid = item.get("id") or ("custom_" + re.sub(r"[^a-z0-9]+", "_", item.get("name","").lower()).strip("_"))
                        custom.append({"id": cid, "name": item.get("name")})
                names["_custom"] = custom
    except Exception:
        pass
    return names

def enabled_section(section_names, key):
    try:
        return bool(section_names.get("_enabled", {}).get(key, True))
    except Exception:
        return True

def custom_sections_from_names(section_names):
    custom = section_names.get("_custom", [])
    return custom if isinstance(custom, list) else []

def compact_text(value):
    return str(value or "").strip()

@app.route("/handover/template", methods=["POST"])
@login_required
@plan_required("pro")
def save_handover_template():
    user = current_user()
    section_names = {}
    enabled = {}
    defaults = default_handover_section_names()

    for key in defaults.keys():
        section_names[key] = request.form.get(f"section_{key}", "").strip() or defaults[key]
        enabled[key] = True if request.form.get(f"enable_{key}") == "on" else False

    # Existing custom sections
    custom_sections = []
    custom_ids = request.form.getlist("custom_section_id[]")
    custom_names = request.form.getlist("custom_section_name[]")
    custom_enabled = set(request.form.getlist("custom_section_enabled[]"))
    for cid, name in zip(custom_ids, custom_names):
        cid = re.sub(r"[^a-zA-Z0-9_]+", "_", (cid or "").strip()) or f"custom_{len(custom_sections)+1}"
        name = (name or "").strip()
        if name and cid in custom_enabled:
            custom_sections.append({"id": cid, "name": name})

    # Add new section
    new_name = request.form.get("new_custom_section_name", "").strip()
    if new_name:
        base = re.sub(r"[^a-z0-9]+", "_", new_name.lower()).strip("_") or "section"
        cid = f"custom_{base}"
        existing_ids = {x["id"] for x in custom_sections}
        n = 2
        original = cid
        while cid in existing_ids:
            cid = f"{original}_{n}"
            n += 1
        custom_sections.append({"id": cid, "name": new_name})

    section_names["_enabled"] = enabled
    section_names["_custom"] = custom_sections

    force_upper = 1 if request.form.get("force_vrid_uppercase") == "on" else 0
    template_name = request.form.get("template_name", "").strip() or "Default Handover"

    conn = get_db()
    existing = conn.execute("SELECT * FROM handover_templates WHERE user_id=? ORDER BY id DESC LIMIT 1", (user["id"],)).fetchone()
    if existing:
        conn.execute("""
            UPDATE handover_templates
            SET name=?, section_names=?, force_vrid_uppercase=?, updated_at=?
            WHERE id=? AND user_id=?
        """, (template_name, json.dumps(section_names), force_upper, datetime.now().isoformat(), existing["id"], user["id"]))
    else:
        conn.execute("""
            INSERT INTO handover_templates (user_id, name, section_names, force_vrid_uppercase, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (user["id"], template_name, json.dumps(section_names), force_upper, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    flash("Handover template saved.", "success")
    return redirect(url_for("handover"))

@app.route("/handover", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def handover():
    user = current_user()
    template = get_handover_template(user["id"])
    section_names = parse_handover_template(template)
    force_vrid_uppercase = bool(row_get(template, "force_vrid_uppercase", 1))

    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        manager = request.form.get("manager", "").strip()
        volume = 0
        planned_hc = int(request.form.get("planned_hc") or 0)
        actual_hc = int(request.form.get("actual_hc") or 0)
        late_trailers = 0
        issues = request.form.get("issues", "").strip()
        actions = request.form.get("actions", "").strip()

        handover_fields = [
            "safe_shift", "slams", "safety_cons", "loads", "safety_rules", "safety_comments",
            "pick_audits", "slam_plan", "picked_since_time", "picked_since_value",
            "full_well_start", "full_well_end", "emc_start", "emc_end", "ops_comments",
            "deliveries_planned", "deliveries_arrived", "same_day_sortation", "next_day_sortation", "sort_comments",
            "collections_arrived", "late_arrivals", "trailers_on_doors", "trailers_needed_cover", "dispatch_comments",
            "suntory_trailers_on_site", "suntory_completed", "suntory_left", "suntory_comments", "aob"
        ]
        extra = {name: request.form.get(name, "").strip() for name in handover_fields}

        absence_rows = []
        absence_types = request.form.getlist("absence_type[]")
        absence_counts = request.form.getlist("absence_count[]")
        for t, c in zip(absence_types, absence_counts):
            if (t or c):
                absence_rows.append({"type": t.strip(), "count": c.strip()})

        dispatch_rows = []
        force_vrid_uppercase = True if request.form.get("force_vrid_uppercase") == "on" else False
        carriers = request.form.getlist("dispatch_carrier[]")
        vrids = request.form.getlist("dispatch_vrid[]")
        completed = request.form.getlist("dispatch_completed[]")
        on_time = request.form.getlist("dispatch_on_time[]")
        issue = request.form.getlist("dispatch_issue[]")
        max_len = max(len(carriers), len(vrids), len(completed), len(on_time), len(issue), 0)
        for i in range(max_len):
            vrid_value = vrids[i].strip() if i < len(vrids) else ""
            if force_vrid_uppercase:
                vrid_value = vrid_value.upper()
            row = {
                "carrier": carriers[i].strip() if i < len(carriers) else "",
                "vrid": vrid_value,
                "completed": completed[i].strip() if i < len(completed) else "",
                "on_time": on_time[i].strip() if i < len(on_time) else "",
                "issue": issue[i].strip() if i < len(issue) else "",
            }
            if any(row.values()):
                dispatch_rows.append(row)

        # Allow the main handover section headers to be edited directly on the form.
        posted_section_names = {}
        posted_enabled = {}
        for key in default_handover_section_names().keys():
            posted_section_names[key] = request.form.get(f"handover_section_{key}", "").strip() or section_names.get(key) or default_handover_section_names()[key]
            posted_enabled[key] = enabled_section(section_names, key)

        custom_sections = custom_sections_from_names(section_names)
        custom_values = []
        for item in custom_sections:
            cid = item.get("id")
            name = request.form.get(f"custom_section_name_{cid}", "").strip() or item.get("name")
            value = request.form.get(f"custom_section_value_{cid}", "").strip()
            custom_values.append({"id": cid, "name": name, "value": value})

        posted_section_names["_enabled"] = posted_enabled
        posted_section_names["_custom"] = [{"id": x.get("id"), "name": x.get("name")} for x in custom_sections]
        section_names = posted_section_names

        # VRID uppercase is controlled inside Dispatch Activities on each handover.
        force_vrid_uppercase = True if request.form.get("force_vrid_uppercase") == "on" else False

        extra["absence_rows"] = absence_rows
        extra["dispatch_rows"] = dispatch_rows
        extra["custom_sections"] = custom_values
        extra["section_names"] = section_names
        extra["force_vrid_uppercase"] = force_vrid_uppercase

        generated = generate_handover_text(date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, extra)
        conn = get_db()

        if request.form.get("save_handover_headers") == "on":
            existing_template = conn.execute("SELECT * FROM handover_templates WHERE user_id=? ORDER BY id DESC LIMIT 1", (user["id"],)).fetchone()
            if existing_template:
                conn.execute("""
                    UPDATE handover_templates
                    SET section_names=?, updated_at=?
                    WHERE id=? AND user_id=?
                """, (json.dumps(section_names), datetime.now().isoformat(), existing_template["id"], user["id"]))
            else:
                conn.execute("""
                    INSERT INTO handover_templates (user_id, name, section_names, force_vrid_uppercase, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, (user["id"], "Default Handover", json.dumps(section_names), 1 if force_vrid_uppercase else 0, datetime.now().isoformat()))

        cur = conn.execute("""
            INSERT INTO handovers (user_id, date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, generated_report, extra_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, generated, json.dumps(extra), datetime.now().isoformat()))
        handover_id = cur.lastrowid
        conn.execute("""
            INSERT INTO handover_audit_log (user_id, handover_id, action, details, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (user["id"], handover_id, "Created", f"Handover created for {date} / {shift}", datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Handover generated.", "success")
        return redirect(url_for("handover"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM handovers WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("handover.html", rows=rows, page="handover", template=template, section_names=section_names, custom_sections=custom_sections_from_names(section_names), enabled_section=enabled_section, force_vrid_uppercase=force_vrid_uppercase)


@app.route("/handover/<int:item_id>/download")
@login_required
@plan_required("pro")
def handover_download(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM handovers WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()
    if not row:
        flash("Handover not found.", "error")
        return redirect(url_for("handover"))
    output = BytesIO()
    output.write(row["generated_report"].encode("utf-8"))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"handover-{row['date']}.txt", mimetype="text/plain")

@app.route("/handover/<int:item_id>/email")
@login_required
@plan_required("pro")
def handover_email(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM handovers WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()
    if not row:
        flash("Handover not found.", "error")
        return redirect(url_for("handover"))
    body = row["generated_report"] or ""
    return render_template("email_export.html", title="Handover Email Preview", body=body, page="handover")


@app.route("/team", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def team():
    user = current_user()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role_select = request.form.get("role_select", "").strip()
        custom_role = request.form.get("custom_role", "").strip()
        role = custom_role if custom_role else (role_select or request.form.get("role", "").strip())
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        notes = request.form.get("notes", "").strip()
        status = request.form.get("status", "Active")
        permissions = request.form.get("permissions", "View only")
        probation_start = request.form.get("probation_start", "").strip()
        probation_end = request.form.get("probation_end", "").strip()
        probation_status = request.form.get("probation_status", "Not set")
        licence_expiry = request.form.get("licence_expiry", "").strip()
        training_type = request.form.get("training_type", "").strip()
        training_expiry = request.form.get("training_expiry", "").strip()
        if not name or not role:
            flash("Please enter name and role.", "error")
            return redirect(url_for("team"))
        conn = get_db()
        conn.execute("""INSERT INTO team_members
                     (user_id, name, role, email, phone, notes, status, permissions, probation_start, probation_end, probation_status, licence_expiry, training_type, training_expiry, created_at)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (user["id"], name, role, email, phone, notes, status, permissions, probation_start, probation_end, probation_status, licence_expiry, training_type, training_expiry, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Team member added.", "success")
        return redirect(url_for("team"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM team_members WHERE user_id = ? ORDER BY name ASC", (user["id"],)).fetchall()
    conn.close()
    return render_template("team.html", rows=rows, page="team")


@app.route("/team/<int:member_id>/update", methods=["POST"])
@login_required
@plan_required("pro")
def update_team_member(member_id):
    user = current_user()
    fields = {
        "name": request.form.get("name", "").strip(),
        "role": request.form.get("role", "").strip(),
        "email": request.form.get("email", "").strip(),
        "phone": request.form.get("phone", "").strip(),
        "status": request.form.get("status", "Active"),
        "notes": request.form.get("notes", "").strip(),
        "probation_start": request.form.get("probation_start", "").strip(),
        "probation_end": request.form.get("probation_end", "").strip(),
        "probation_status": request.form.get("probation_status", "Not set"),
        "licence_expiry": request.form.get("licence_expiry", "").strip(),
        "training_type": request.form.get("training_type", "").strip(),
        "training_expiry": request.form.get("training_expiry", "").strip(),
    }
    if not fields["name"] or not fields["role"]:
        flash("Name and role are required.", "error")
        return redirect(url_for("team"))
    conn = get_db()
    conn.execute("""
        UPDATE team_members SET name=?, role=?, email=?, phone=?, status=?, notes=?,
        probation_start=?, probation_end=?, probation_status=?, licence_expiry=?, training_type=?, training_expiry=?
        WHERE id=? AND user_id=?
    """, (fields["name"], fields["role"], fields["email"], fields["phone"], fields["status"], fields["notes"], fields["probation_start"], fields["probation_end"], fields["probation_status"], fields["licence_expiry"], fields["training_type"], fields["training_expiry"], member_id, user["id"]))
    conn.commit(); conn.close()
    flash("Team member updated.", "success")
    return redirect(url_for("team"))

@app.route("/team/<int:member_id>/delete", methods=["POST"])
@login_required
@plan_required("pro")
def delete_team_member(member_id):
    user = current_user()
    conn = get_db()
    conn.execute("DELETE FROM team_members WHERE id=? AND user_id=?", (member_id, user["id"]))
    conn.commit(); conn.close()
    flash("Team member deleted.", "success")
    return redirect(url_for("team"))

@app.route("/shift-calendar/note", methods=["POST"])
@login_required
def save_calendar_note():
    user = current_user()
    date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
    notes = request.form.get("notes", "").strip()
    return_to = request.form.get("return_to") or url_for("shift_calendar")
    conn = get_db()
    existing = conn.execute("SELECT * FROM shift_calendar WHERE user_id=? AND date=?", (user["id"], date)).fetchone()
    if existing:
        conn.execute("UPDATE shift_calendar SET notes=?, source='Manual' WHERE user_id=? AND date=?", (notes, user["id"], date))
    else:
        conn.execute("""INSERT INTO shift_calendar (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at)
                      VALUES (?, ?, 'Not Set', '', '', '', ?, 'Manual', ?)""", (user["id"], date, notes, datetime.now().isoformat()))
    conn.commit(); conn.close()
    flash("Calendar note saved.", "success")
    return redirect(return_to)


@app.route("/team/export")
@login_required
@plan_required("business")
def export_team():
    user = current_user()
    conn = get_db()
    rows = conn.execute("SELECT * FROM team_members WHERE user_id = ? ORDER BY name ASC", (user["id"],)).fetchall()
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = "Team Members"
    ws.append(["Name", "Role", "Email", "Phone", "Status", "Probation Start", "Probation End", "Probation Status", "Notes", "Created At"])
    for row in rows:
        ws.append([row["name"], row["role"], row["email"], row_get(row,"phone",""), row["status"], row_get(row,"probation_start",""), row_get(row,"probation_end",""), row_get(row,"probation_status",""), row_get(row,"notes",""), row["created_at"]])
    style_excel_header(ws)
    return excel_response(wb, "team-members.xlsx")

@app.route("/handover/export")
@login_required
@plan_required("business")
def export_handovers():
    user = current_user()
    conn = get_db()
    rows = conn.execute("SELECT * FROM handovers WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Handover"

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_fill = PatternFill("solid", fgColor="EAF3FF")
    header_fill = PatternFill("solid", fgColor="F8FAFC")
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)

    def cell(r, c, value="", bold_text=False, fill=None, align="left"):
        ws.cell(r, c, value)
        ws.cell(r, c).border = border
        ws.cell(r, c).alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bold_text:
            ws.cell(r, c).font = bold
        if fill:
            ws.cell(r, c).fill = fill
        return ws.cell(r, c)

    def merge(r1, c1, r2, c2, value="", bold_text=False, fill=None, align="left"):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell(r1, c1, value, bold_text, fill, align)
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = border
                ws.cell(rr, cc).alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
                if fill:
                    ws.cell(rr, cc).fill = fill
        return ws.cell(r1, c1)

    def enabled(names, key):
        return bool(names.get("_enabled", {}).get(key, True))

    row_num = 1
    max_col = 10

    for h in rows:
        try:
            extra = json.loads(row_get(h, "extra_json", "{}") or "{}")
        except Exception:
            extra = {}

        names = default_handover_section_names()
        names["_enabled"] = {key: True for key in default_handover_section_names()}
        names["_custom"] = []
        if isinstance(extra.get("section_names"), dict):
            saved = extra.get("section_names")
            for key in default_handover_section_names():
                if saved.get(key):
                    names[key] = saved.get(key)
            if isinstance(saved.get("_enabled"), dict):
                names["_enabled"].update(saved.get("_enabled"))
            if isinstance(saved.get("_custom"), list):
                names["_custom"] = saved.get("_custom")

        absence = extra.get("absence_rows", []) or []
        absence_text = " / ".join([f"{x.get('count','')} {x.get('type','')}" for x in absence if x.get("type") or x.get("count")])

        merge(row_num, 1, row_num, max_col, f"{h['shift'] or 'Day Shift'} Handover", True, None, "center")
        ws.cell(row_num, 1).font = title_font
        row_num += 1
        merge(row_num, 1, row_num, 2, f"Date - {h['date']}")
        merge(row_num, 3, row_num, 4, f"Shift - {h['shift']}")
        merge(row_num, 5, row_num + 1, max_col, f"Attendance - {h['actual_hc']}/{h['planned_hc']}\n{absence_text}", False, None, "center")
        row_num += 2

        if enabled(names, "safety"):
            merge(row_num, 1, row_num, 4, f"{names.get('safety')} - {extra.get('safe_shift','')}", False, None, "center")
            merge(row_num, 5, row_num, max_col, "")
            row_num += 1
            vals = [names.get("safety"), "SLAMS", extra.get("slams",""), "Safety Cons", extra.get("safety_cons",""), "LOADS", extra.get("loads",""), "Safety Rules", extra.get("safety_rules","")]
            for c, val in enumerate(vals, 1):
                cell(row_num, c, val, bold_text=(c == 1), align="center", fill=header_fill if c == 1 else None)
            for c in range(len(vals)+1, max_col+1):
                cell(row_num, c, "")
            row_num += 1

        if enabled(names, "operations"):
            merge(row_num, 1, row_num, max_col, names.get("operations"), True, section_fill)
            row_num += 1
            for label, value in [
                ("Pick Audits Completed", extra.get("pick_audits","")),
                ("Slam Plan", extra.get("slam_plan","")),
                (f"Picked since {extra.get('picked_since_time','06:00')} hrs", extra.get("picked_since_value","")),
                ("Full Well at start & end of Shift", f"{extra.get('full_well_start','')}/{extra.get('full_well_end','')}"),
                ("Well to cover EMC's at start & end of Shift", f"{extra.get('emc_start','')}/{extra.get('emc_end','')}")
            ]:
                merge(row_num, 1, row_num, max_col, f"{label} - {value}")
                row_num += 1
            merge(row_num, 1, row_num + 2, max_col, f"Additional Comments - {extra.get('ops_comments','')}")
            row_num += 3

        if enabled(names, "sort"):
            merge(row_num, 1, row_num, max_col, names.get("sort"), True, section_fill)
            row_num += 1
            for label, value in [
                ("Deliveries Planned", extra.get("deliveries_planned","")),
                ("Deliveries Arrived", extra.get("deliveries_arrived","")),
                ("Planned Same Day Sortation", extra.get("same_day_sortation","")),
                ("Planned Next Day Sortation", extra.get("next_day_sortation",""))
            ]:
                merge(row_num, 1, row_num, max_col, f"{label} - {value}")
                row_num += 1
            merge(row_num, 1, row_num + 2, max_col, f"Additional Comments - {extra.get('sort_comments','')}")
            row_num += 3

        if enabled(names, "dispatch"):
            merge(row_num, 1, row_num, max_col, names.get("dispatch"), True, section_fill)
            row_num += 1
            merge(row_num, 1, row_num, 2, "Collections Arrived")
            merge(row_num, 3, row_num, 4, extra.get("collections_arrived",""), align="center")
            merge(row_num, 5, row_num, 6, "LATE ARRIVALS", align="center")
            merge(row_num, 7, row_num, max_col, extra.get("late_arrivals",""), align="center")
            row_num += 1
            merge(row_num, 1, row_num, 2, "Trailers On Doors")
            merge(row_num, 3, row_num, 4, extra.get("trailers_on_doors",""), align="center")
            merge(row_num, 5, row_num, 7, "Trailers needed cover today CPT's")
            merge(row_num, 8, row_num, max_col, extra.get("trailers_needed_cover",""), align="center")
            row_num += 1

            dispatch_rows = extra.get("dispatch_rows", []) or []
            if dispatch_rows:
                ranges = [(1,2), (3,5), (6,7), (8,9), (10,10)]
                for (c1,c2), label in zip(ranges, ["Carrier / Dispatch", "VRID", "Completed", "On Time", "Issue"]):
                    merge(row_num, c1, row_num, c2, label, True, header_fill, "center")
                row_num += 1
                for d in dispatch_rows:
                    for (c1,c2), val in zip(ranges, [d.get("carrier",""), d.get("vrid",""), d.get("completed",""), d.get("on_time",""), d.get("issue","")]):
                        merge(row_num, c1, row_num, c2, val, align="center")
                    row_num += 1
            merge(row_num, 1, row_num + 1, max_col, f"Additional Comments - {extra.get('dispatch_comments','')}")
            row_num += 2

        if enabled(names, "suntory"):
            merge(row_num, 1, row_num, max_col, names.get("suntory"), True, section_fill)
            row_num += 1
            merge(row_num, 1, row_num, 3, "Trailers on site to complete")
            merge(row_num, 4, row_num, 4, extra.get("suntory_trailers_on_site",""), align="center")
            merge(row_num, 5, row_num, 7, "Completed on shift")
            merge(row_num, 8, row_num, max_col, extra.get("suntory_completed",""), align="center")
            row_num += 1
            merge(row_num, 1, row_num, 3, "Left to complete")
            merge(row_num, 4, row_num, max_col, extra.get("suntory_left",""), align="center")
            row_num += 1
            merge(row_num, 1, row_num + 1, max_col, f"Additional Comments - {extra.get('suntory_comments','')}")
            row_num += 2

        for cs in extra.get("custom_sections", []) or []:
            if cs.get("name"):
                merge(row_num, 1, row_num, max_col, cs.get("name"), True, section_fill)
                row_num += 1
                merge(row_num, 1, row_num + 2, max_col, cs.get("value",""))
                row_num += 3

        if enabled(names, "aob"):
            merge(row_num, 1, row_num, max_col, names.get("aob"), True, section_fill)
            row_num += 1
            merge(row_num, 1, row_num + 2, max_col, extra.get("aob",""))
            row_num += 3

        row_num += 1

    widths = [18, 14, 16, 14, 18, 14, 14, 14, 14, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in range(1, row_num + 1):
        ws.row_dimensions[r].height = 21

    # PDF-ready Excel layout
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_area = f"A1:J{max(row_num, 1)}"

    return excel_response(wb, "handover-dhl-style.xlsx")

@app.route("/admin/export")
@login_required
def export_admin_users():
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))
    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = "Users"
    ws.append(["Name", "Email", "Plan", "Role", "Company", "Phone", "Status", "Pro Until", "Created At"])
    for u in users:
        ws.append([u["name"], u["email"], u["plan"], row_get(u,"role",""), row_get(u,"company_name",""), row_get(u,"phone",""), row_get(u,"subscription_status",""), row_get(u,"pro_expires_at",""), u["created_at"]])
    style_excel_header(ws)
    return excel_response(wb, "admin-users.xlsx")

@app.route("/admin/account", methods=["POST"])
@login_required
def admin_account_update():
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    name = request.form.get("name", "").strip() or admin["name"]
    email = request.form.get("email", "").strip().lower() or admin["email"]
    current_password = request.form.get("current_password", "")
    password = request.form.get("password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if password:
        if not check_password_hash(admin["password_hash"], current_password):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("admin_dashboard"))
        if password != confirm_password:
            flash("New passwords do not match.", "error")
            return redirect(url_for("admin_dashboard"))
        has_lower = any(c.islower() for c in password)
        has_upper = any(c.isupper() for c in password)
        has_number = any(c.isdigit() for c in password)
        has_special = any(not c.isalnum() for c in password)
        if len(password) < 8 or not (has_lower and has_upper and has_number and has_special):
            flash("New password must be at least 8 characters and include lowercase, uppercase, number and special character.", "error")
            return redirect(url_for("admin_dashboard"))

    conn = get_db()
    try:
        if password:
            conn.execute("""
                UPDATE users
                SET name = ?, email = ?, role = 'Admin', password_hash = ?, password_changed_at = ?
                WHERE id = ?
            """, (name, email, generate_password_hash(password), datetime.now().isoformat(), admin["id"]))
        else:
            conn.execute("UPDATE users SET name = ?, email = ?, role = 'Admin' WHERE id = ?", (name, email, admin["id"]))
        conn.commit()
        flash("Admin login details updated.", "success")
    except sqlite3.IntegrityError:
        flash("That email address is already used by another account.", "error")
    finally:
        conn.close()
    return redirect(url_for("admin_dashboard"))


@app.route("/yard-settings/location/add", methods=["POST"])
@login_required
@plan_required("pro")
def add_yard_location():
    user = current_user()
    name = request.form.get("location_name", "").strip()
    prefix = request.form.get("location_prefix", "").strip()
    if not name:
        flash("Enter a location name.", "error")
        return redirect(url_for("yard_settings"))
    conn = get_db()
    conn.execute("INSERT INTO custom_locations (user_id, name, prefix, created_at) VALUES (?, ?, ?, ?)", (user["id"], name, prefix, datetime.now().isoformat()))
    conn.commit(); conn.close()
    flash("Custom yard location added.", "success")
    return redirect(url_for("yard_settings"))

@app.route("/yard-settings/location/<int:location_id>/delete", methods=["POST"])
@login_required
@plan_required("pro")
def delete_yard_location(location_id):
    user = current_user()
    conn = get_db()
    conn.execute("DELETE FROM custom_locations WHERE id = ? AND user_id = ?", (location_id, user["id"]))
    conn.commit(); conn.close()
    flash("Custom yard location deleted.", "success")
    return redirect(url_for("yard_settings"))

@app.route("/daily-shift-log", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def daily_shift_log():
    user = current_user()
    if request.method == "POST":
        f = request.form
        photo = request.files.get("photo")
        filename = ""
        if photo and photo.filename:
            filename = secure_filename(f"{uuid.uuid4().hex}_{photo.filename}")
            photo.save(os.path.join(UPLOAD_DIR, filename))
        conn = get_db()
        conn.execute("""INSERT INTO daily_shift_logs
            (user_id,date,shift,manager,volume,planned_hc,actual_hc,late_trailers,safety,issues,actions,notes,photo_filename,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user["id"], f.get("date") or datetime.today().strftime("%Y-%m-%d"), f.get("shift",""), f.get("manager",""), int(f.get("volume") or 0), int(f.get("planned_hc") or 0), int(f.get("actual_hc") or 0), int(f.get("late_trailers") or 0), f.get("safety",""), f.get("issues",""), f.get("actions",""), f.get("notes",""), filename, datetime.now().isoformat()))
        if f.get("actions","").strip():
            conn.execute("INSERT INTO action_tracker (user_id,date,title,owner,due_date,status,priority,source,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (user["id"], datetime.today().strftime("%Y-%m-%d"), f.get("actions").strip()[:120], f.get("manager",""), f.get("date"), "Open", "Medium", "Daily Shift Log", f.get("issues",""), datetime.now().isoformat()))
        conn.commit(); conn.close()
        flash("Daily shift log saved.", "success")
        return redirect(url_for("daily_shift_log"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM daily_shift_logs WHERE user_id=? ORDER BY date DESC,id DESC LIMIT 100", (user["id"],)).fetchall()
    conn.close()
    return render_template("daily_shift_log.html", rows=rows, page="daily_shift_log")

@app.route("/daily-shift-log/export")
@login_required
@plan_required("pro")
def export_daily_shift_log():
    user=current_user(); conn=get_db(); rows=conn.execute("SELECT * FROM daily_shift_logs WHERE user_id=? ORDER BY date DESC", (user["id"],)).fetchall(); conn.close()
    wb=Workbook(); ws=wb.active; ws.title="Daily Shift Log"; ws.append(["Date","Shift","Manager","Volume","Planned HC","Actual HC","Late Trailers","Safety","Issues","Actions","Notes"])
    for r in rows: ws.append([r["date"],r["shift"],r["manager"],r["volume"],r["planned_hc"],r["actual_hc"],r["late_trailers"],r["safety"],r["issues"],r["actions"],r["notes"]])
    style_excel_header(ws); return excel_response(wb,"daily-shift-log.xlsx")

@app.route("/daily-shift-log/<int:item_id>/email")
@login_required
@plan_required("pro")
def daily_shift_log_email(item_id):
    user=current_user(); conn=get_db(); r=conn.execute("SELECT * FROM daily_shift_logs WHERE id=? AND user_id=?", (item_id,user["id"])).fetchone(); conn.close()
    if not r: flash("Log not found.","error"); return redirect(url_for("daily_shift_log"))
    body=f"Daily Shift Log - {r['date']}\nShift: {r['shift']}\nManager: {r['manager']}\nVolume: {r['volume']}\nPlanned HC: {r['planned_hc']}\nActual HC: {r['actual_hc']}\nLate trailers: {r['late_trailers']}\n\nSafety:\n{r['safety']}\n\nIssues:\n{r['issues']}\n\nActions:\n{r['actions']}\n\nNotes:\n{r['notes']}"
    return render_template("email_export.html", title="Daily Shift Log Email", body=body, page="daily_shift_log")

@app.route("/actions", methods=["GET","POST"])
@login_required
@plan_required("pro")
def actions():
    user=current_user()
    if request.method=="POST":
        f=request.form; conn=get_db(); conn.execute("INSERT INTO action_tracker (user_id,date,title,owner,due_date,status,priority,source,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)", (user["id"], datetime.today().strftime("%Y-%m-%d"), f.get("title",""), f.get("owner",""), f.get("due_date",""), f.get("status","Open"), f.get("priority","Medium"), f.get("source","Manual"), f.get("notes",""), datetime.now().isoformat())); conn.commit(); conn.close(); flash("Action saved.","success"); return redirect(url_for("actions"))
    conn=get_db(); rows=conn.execute("SELECT * FROM action_tracker WHERE user_id=? ORDER BY CASE status WHEN 'Open' THEN 0 WHEN 'In Progress' THEN 1 ELSE 2 END, due_date ASC", (user["id"],)).fetchall(); conn.close()
    return render_template("actions.html", rows=rows, page="actions")

@app.route("/actions/<int:item_id>/update", methods=["POST"])
@login_required
@plan_required("pro")
def update_action(item_id):
    user=current_user(); f=request.form; conn=get_db(); conn.execute("UPDATE action_tracker SET title=?,owner=?,due_date=?,status=?,priority=?,notes=? WHERE id=? AND user_id=?", (f.get("title",""),f.get("owner",""),f.get("due_date",""),f.get("status","Open"),f.get("priority","Medium"),f.get("notes",""),item_id,user["id"])); conn.commit(); conn.close(); flash("Action updated.","success"); return redirect(url_for("actions"))

@app.route("/actions/<int:item_id>/delete", methods=["POST"])
@login_required
@plan_required("pro")
def delete_action(item_id):
    user=current_user(); conn=get_db(); conn.execute("DELETE FROM action_tracker WHERE id=? AND user_id=?", (item_id,user["id"])); conn.commit(); conn.close(); flash("Action deleted.","success"); return redirect(url_for("actions"))

@app.route("/actions/export")
@login_required
@plan_required("pro")
def export_actions():
    user=current_user(); conn=get_db(); rows=conn.execute("SELECT * FROM action_tracker WHERE user_id=? ORDER BY due_date ASC", (user["id"],)).fetchall(); conn.close(); wb=Workbook(); ws=wb.active; ws.title="Actions"; ws.append(["Date","Title","Owner","Due Date","Status","Priority","Source","Notes"])
    for r in rows: ws.append([r["date"],r["title"],r["owner"],r["due_date"],r["status"],r["priority"],r["source"],r["notes"]])
    style_excel_header(ws); return excel_response(wb,"actions.xlsx")

@app.route("/absence", methods=["GET","POST"])
@login_required
@plan_required("pro")
def absence():
    user=current_user(); conn=get_db()
    if request.method=="POST":
        f=request.form; conn.execute("INSERT INTO absence_records (user_id,member_name,start_date,end_date,absence_type,notes,created_at) VALUES (?,?,?,?,?,?,?)", (user["id"],f.get("member_name",""),f.get("start_date",""),f.get("end_date",""),f.get("absence_type","Sick"),f.get("notes",""),datetime.now().isoformat())); conn.commit(); flash("Absence saved.","success"); conn.close(); return redirect(url_for("absence"))
    rows=conn.execute("SELECT * FROM absence_records WHERE user_id=? ORDER BY start_date DESC", (user["id"],)).fetchall(); team=conn.execute("SELECT name FROM team_members WHERE user_id=? ORDER BY name", (user["id"],)).fetchall(); conn.close(); return render_template("absence.html", rows=rows, team=team, page="absence")

@app.route("/absence/export")
@login_required
@plan_required("pro")
def export_absence():
    user=current_user(); conn=get_db(); rows=conn.execute("SELECT * FROM absence_records WHERE user_id=? ORDER BY start_date DESC", (user["id"],)).fetchall(); conn.close(); wb=Workbook(); ws=wb.active; ws.title="Absence"; ws.append(["Name","Start","End","Type","Notes"]); [ws.append([r["member_name"],r["start_date"],r["end_date"],r["absence_type"],r["notes"]]) for r in rows]; style_excel_header(ws); return excel_response(wb,"absence.xlsx")

@app.route("/evidence", methods=["GET","POST"])
@login_required
@plan_required("pro")
def evidence():
    user=current_user()
    if request.method=="POST":
        photo=request.files.get("photo"); filename=""
        if photo and photo.filename:
            filename=secure_filename(f"{uuid.uuid4().hex}_{photo.filename}"); photo.save(os.path.join(UPLOAD_DIR, filename))
        f=request.form; conn=get_db(); conn.execute("INSERT INTO photo_records (user_id,date,image_filename,trailer_id,location_detail,damage_notes,recognition_notes,confidence,category,comment,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)", (user["id"], f.get("date") or datetime.today().strftime("%Y-%m-%d"), filename, f.get("trailer_id",""), f.get("location_detail",""), f.get("damage_notes",""), "", "", f.get("category","Evidence"), f.get("comment",""), datetime.now().isoformat())); conn.commit(); conn.close(); flash("Evidence saved.","success"); return redirect(url_for("evidence"))
    q=request.args.get("q",""); conn=get_db(); rows=conn.execute("SELECT * FROM photo_records WHERE user_id=? AND (?='' OR trailer_id LIKE ? OR location_detail LIKE ? OR damage_notes LIKE ? OR comment LIKE ?) ORDER BY date DESC,id DESC", (user["id"],q,f"%{q}%",f"%{q}%",f"%{q}%",f"%{q}%")).fetchall(); conn.close(); return render_template("evidence.html", rows=rows, q=q, page="evidence")

@app.route("/search")
@login_required
def global_search():
    user=current_user(); q=request.args.get("q","").strip(); results=[]
    if q:
        conn=get_db(); like=f"%{q}%"
        for r in conn.execute("SELECT date, trailer_id AS title, notes AS text FROM yard_checks WHERE user_id=? AND (trailer_id LIKE ? OR notes LIKE ? OR location_detail LIKE ?) LIMIT 20", (user["id"],like,like,like)).fetchall(): results.append({"module":"Yard Check","date":r["date"],"title":r["title"],"text":r["text"],"url":url_for("yard_check")})
        for r in conn.execute("SELECT date, title, notes AS text FROM action_tracker WHERE user_id=? AND (title LIKE ? OR notes LIKE ? OR owner LIKE ?) LIMIT 20", (user["id"],like,like,like)).fetchall(): results.append({"module":"Actions","date":r["date"],"title":r["title"],"text":r["text"],"url":url_for("actions")})
        for r in conn.execute("SELECT date, shift AS title, issues AS text FROM daily_shift_logs WHERE user_id=? AND (issues LIKE ? OR actions LIKE ? OR notes LIKE ?) LIMIT 20", (user["id"],like,like,like)).fetchall(): results.append({"module":"Daily Log","date":r["date"],"title":r["title"],"text":r["text"],"url":url_for("daily_shift_log")})
        for r in conn.execute("SELECT name AS title, role, notes FROM team_members WHERE user_id=? AND (name LIKE ? OR role LIKE ? OR notes LIKE ?) LIMIT 20", (user["id"],like,like,like)).fetchall(): results.append({"module":"Team","date":"","title":r["title"],"text":f"{r['role']} {r['notes'] or ''}","url":url_for("team")})
        conn.close()
    return render_template("search.html", q=q, results=results, page="search")

@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@login_required
def admin_delete_user(user_id):
    admin=current_user()
    if not is_admin(admin): flash("Admin access only.","error"); return redirect(url_for("index"))
    if user_id==admin["id"]: flash("You cannot delete your own admin account.","error"); return redirect(url_for("admin_dashboard"))
    conn=get_db()
    for table in ["mileage","expenses","invoices","yard_checks","kpi_records","handovers","team_members","shift_calendar","morning_briefs","photo_records","daily_shift_logs","action_tracker","absence_records","remember_tokens"]:
        try: conn.execute(f"DELETE FROM {table} WHERE user_id=?", (user_id,))
        except Exception: pass
    conn.execute("DELETE FROM users WHERE id=?", (user_id,)); conn.commit(); conn.close(); flash("Account deleted.","success"); return redirect(url_for("admin_dashboard"))

@app.route("/admin/inactivity-check")
@login_required
def admin_inactivity_check():
    admin=current_user()
    if not is_admin(admin): flash("Admin access only.","error"); return redirect(url_for("index"))
    now=datetime.now(); warn_before=(now-timedelta(days=180)).isoformat(); delete_before=(now-timedelta(days=365)).isoformat(); conn=get_db()
    conn.execute("UPDATE users SET inactive_warning_at=? WHERE email!='admin@whs-app.com' AND COALESCE(last_login_at,created_at) < ? AND inactive_warning_at IS NULL", (now.isoformat(timespec="seconds"), warn_before))
    stale=conn.execute("SELECT id FROM users WHERE email!='admin@whs-app.com' AND COALESCE(last_login_at,created_at) < ?", (delete_before,)).fetchall()
    for u in stale:
        for table in ["mileage","expenses","invoices","yard_checks","kpi_records","handovers","team_members","shift_calendar","morning_briefs","photo_records","daily_shift_logs","action_tracker","absence_records","remember_tokens"]:
            try: conn.execute(f"DELETE FROM {table} WHERE user_id=?", (u["id"],))
            except Exception: pass
        conn.execute("DELETE FROM users WHERE id=?", (u["id"],))
    conn.commit(); conn.close(); flash(f"Inactivity check complete. Deleted {len(stale)} inactive account(s).", "success"); return redirect(url_for("admin_dashboard"))

@app.route("/operations")
@login_required
@plan_required("business")
def operations():
    user = current_user()
    conn = get_db()
    kpi = conn.execute("""
        SELECT COALESCE(SUM(volume), 0) total_volume, COALESCE(AVG(actual_rate), 0) avg_rate,
               COALESCE(SUM(late_trailers), 0) total_late, COALESCE(SUM(errors), 0) total_errors
        FROM kpi_records WHERE user_id = ?
    """, (user["id"],)).fetchone()
    yard_count = conn.execute("SELECT COUNT(*) FROM yard_checks WHERE user_id = ?", (user["id"],)).fetchone()[0]
    team_count = conn.execute("SELECT COUNT(*) FROM team_members WHERE user_id = ?", (user["id"],)).fetchone()[0]
    handover_count = conn.execute("SELECT COUNT(*) FROM handovers WHERE user_id = ?", (user["id"],)).fetchone()[0]
    photo_count = conn.execute("SELECT COUNT(*) FROM photo_records WHERE user_id = ?", (user["id"],)).fetchone()[0]
    conn.close()
    return render_template("operations.html", kpi=kpi, yard_count=yard_count, team_count=team_count, handover_count=handover_count, photo_count=photo_count, page="operations")


@app.route("/shift-planner", methods=["GET", "POST"])
@login_required
@plan_required("business")
def shift_planner():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        volume = int(request.form.get("volume") or 0)
        available_hc = int(request.form.get("available_hc") or 0)
        target_rate = float(request.form.get("target_rate") or 0)
        planned_hours = float(request.form.get("planned_hours") or 0)
        ai_plan = generate_ai_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours)
        conn = get_db()
        conn.execute("""
            INSERT INTO shift_plans (user_id, date, shift, volume, available_hc, target_rate, planned_hours, ai_plan, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, volume, available_hc, target_rate, planned_hours, ai_plan, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("AI shift plan generated.", "success")
        return redirect(url_for("shift_planner"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM shift_plans WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("shift_planner.html", rows=rows, page="shift_planner")


@app.route("/photo-recognition", methods=["GET", "POST"])
@login_required
@plan_required("business")
def photo_recognition():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        trailer_id = request.form.get("trailer_id", "").strip().upper()
        location_detail = request.form.get("location_detail", "").strip()
        damage_notes = request.form.get("damage_notes", "").strip()
        recognition_notes = request.form.get("recognition_notes", "").strip()
        confidence = request.form.get("confidence", "Manual check")
        ai_result = analyse_photo_ai(trailer_id, location_detail, damage_notes, recognition_notes)
        image_filename = None

        uploaded = request.files.get("image")
        if uploaded and uploaded.filename:
            if not allowed_image(uploaded.filename):
                flash("Only PNG, JPG, JPEG or WEBP images are allowed.", "error")
                return redirect(url_for("photo_recognition"))
            safe = secure_filename(uploaded.filename)
            image_filename = f"{uuid.uuid4().hex}_{safe}"
            uploaded.save(os.path.join(UPLOAD_DIR, image_filename))

        conn = get_db()
        conn.execute("""
            INSERT INTO photo_records (user_id, date, image_filename, trailer_id, location_detail, damage_notes, recognition_notes, confidence, ai_result, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, image_filename, trailer_id, location_detail, damage_notes, recognition_notes, confidence, ai_result, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Photo recognition record saved.", "success")
        return redirect(url_for("photo_recognition"))

    conn = get_db()
    rows = conn.execute("SELECT * FROM photo_records WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("photo_recognition.html", rows=rows, page="photo_recognition")



@app.route("/billing/checkout/<plan>")
@login_required
def billing_checkout(plan):
    if plan not in ["pro", "business"]:
        flash("Invalid billing plan.", "error")
        return redirect(url_for("pricing"))

    user = current_user()

    if not stripe or not STRIPE_SECRET_KEY:
        flash("Stripe is not configured yet. For now, plan selection works in demo/manual mode.", "error")
        return redirect(url_for("pricing"))

    price_id = STRIPE_PRO_PRICE_ID if plan == "pro" else STRIPE_BUSINESS_PRICE_ID
    if not price_id:
        flash("Stripe price ID is missing for this plan.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = STRIPE_SECRET_KEY
    checkout = stripe.checkout.Session.create(
        mode="subscription",
        payment_method_types=["card"],
        customer_email=user["email"],
        line_items=[{"price": price_id, "quantity": 1}],
        success_url=f"{APP_BASE_URL}/billing/success/{plan}",
        cancel_url=f"{APP_BASE_URL}/pricing",
    )
    return redirect(checkout.url, code=303)


@app.route("/billing/success/<plan>")
@login_required
def billing_success(plan):
    if plan not in PLAN_ORDER:
        flash("Invalid plan.", "error")
        return redirect(url_for("pricing"))
    user = current_user()
    conn = get_db()
    conn.execute("UPDATE users SET plan = ?, subscription_status = 'active' WHERE id = ?", (plan, user["id"]))
    conn.commit()
    conn.close()
    flash(f"Subscription activated: {PLAN_NAMES[plan]}.", "success")
    return redirect(url_for("pricing"))



@app.route("/admin")
@login_required
def admin_dashboard():
    user = current_user()
    if not is_admin(user):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()

    stats = {
        "users": conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "free": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'free'").fetchone()[0],
        "pro": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'pro'").fetchone()[0],
        "business": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'business'").fetchone()[0],
        "yard": conn.execute("SELECT COUNT(*) FROM yard_checks").fetchone()[0],
        "kpi": conn.execute("SELECT COUNT(*) FROM kpi_records").fetchone()[0],
        "handovers": conn.execute("SELECT COUNT(*) FROM handovers").fetchone()[0],
        "team": conn.execute("SELECT COUNT(*) FROM team_members").fetchone()[0],
        "daily_logs": conn.execute("SELECT COUNT(*) FROM daily_shift_logs").fetchone()[0],
        "actions": conn.execute("SELECT COUNT(*) FROM action_tracker").fetchone()[0],
        "absence": conn.execute("SELECT COUNT(*) FROM absence_records").fetchone()[0],
    }
    conn.close()

    return render_template("admin.html", users=users, stats=stats, page="admin")



@app.route("/admin/demo-users/create", methods=["POST"])
@login_required
def admin_create_demo_users():
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    demo_password = "WHS2026!"
    demo_users = [
        ("Demo User 1", "demo1@whs.ai"),
        ("Demo User 2", "demo2@whs.ai"),
        ("Demo User 3", "demo3@whs.ai"),
        ("Demo User 4", "demo4@whs.ai"),
        ("Demo User 5", "demo5@whs.ai"),
    ]

    conn = get_db()
    created = 0
    updated = 0
    now = datetime.now().isoformat(timespec="seconds")

    for name, email in demo_users:
        existing = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        password_hash = generate_password_hash(demo_password)
        if existing:
            conn.execute("""
                UPDATE users
                SET name = ?,
                    password_hash = ?,
                    plan = 'business',
                    role = 'Manager',
                    subscription_status = 'demo_business',
                    pro_expires_at = NULL,
                    pro_reason = 'Demo Business account'
                WHERE email = ?
            """, (name, password_hash, email))
            updated += 1
        else:
            conn.execute("""
                INSERT INTO users (name, email, password_hash, plan, business_name, created_at, role, subscription_status, pro_expires_at, pro_reason)
                VALUES (?, ?, ?, 'business', 'WHS Demo', ?, 'Manager', 'demo_business', NULL, 'Demo Business account')
            """, (name, email, password_hash, now))
            created += 1

    conn.commit()
    conn.close()

    flash(f"Demo Business users ready. Created: {created}, updated/reset: {updated}. Password: {demo_password}", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/demo-users/delete", methods=["POST"])
@login_required
def admin_delete_demo_users():
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    emails = [f"demo{i}@whs.ai" for i in range(1, 6)]
    conn = get_db()
    deleted = 0
    for email in emails:
        user = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        if user:
            uid = user["id"]
            # Delete common user data tables when they exist.
            for table in [
                "mileage", "expenses", "invoices", "yard_checks", "handovers", "team_members",
                "shift_calendar", "daily_logs", "actions", "sites", "custom_locations",
                "handover_templates", "calendar_notes", "shift_plans"
            ]:
                try:
                    conn.execute(f"DELETE FROM {table} WHERE user_id = ?", (uid,))
                except Exception:
                    pass
            conn.execute("DELETE FROM users WHERE id = ?", (uid,))
            deleted += 1
    conn.commit()
    conn.close()
    flash(f"Deleted {deleted} demo users.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/plan/<plan>")
@login_required
def admin_set_user_plan(user_id, plan):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    if plan not in ["free", "pro", "business"]:
        flash("Invalid plan.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()

    if plan in ["pro", "business"]:
        conn.execute("""
            UPDATE users
            SET plan = ?,
                subscription_status = 'manual_admin',
                pro_expires_at = NULL,
                pro_reason = ?
            WHERE id = ?
        """, (plan, f"Manual {PLAN_NAMES[plan]} set by admin", user_id))
    else:
        conn.execute("""
            UPDATE users
            SET plan = 'free',
                subscription_status = 'manual_admin',
                pro_expires_at = NULL,
                pro_reason = NULL
            WHERE id = ?
        """, (user_id,))

    conn.commit()
    conn.close()

    flash(f"User plan changed to {PLAN_NAMES[plan]}.", "success")
    return redirect(url_for("admin_dashboard"))



@app.route("/admin/user/<int:user_id>/trial/pro30")
@login_required
def admin_set_user_trial_pro30(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    expires_at = (datetime.now() + timedelta(days=30)).isoformat(timespec="seconds")

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            subscription_status = 'trial_admin',
            pro_expires_at = ?,
            pro_reason = '30 day Pro trial set by admin'
        WHERE id = ?
    """, (expires_at, user_id))
    conn.commit()
    conn.close()

    flash("30 day Pro trial activated for user.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/gift/pro")
@login_required
def admin_set_user_gift_pro(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            subscription_status = 'gift_admin',
            pro_expires_at = NULL,
            pro_reason = 'Free Pro gift set by admin'
        WHERE id = ?
    """, (user_id,))
    conn.commit()
    conn.close()

    flash("Free Pro gift activated for user.", "success")
    return redirect(url_for("admin_dashboard"))



@app.route("/admin/user/<int:user_id>/trial/business30")
@login_required
def admin_set_user_trial_business30(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    expires_at = (datetime.now() + timedelta(days=30)).isoformat(timespec="seconds")
    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'business',
            subscription_status = 'trial_admin',
            pro_expires_at = ?,
            pro_reason = '30 day Business trial set by admin'
        WHERE id = ?
    """, (expires_at, user_id))
    conn.commit()
    conn.close()
    flash("30 day Business trial activated.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/gift/business")
@login_required
def admin_set_user_gift_business(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'business',
            subscription_status = 'gift_admin',
            pro_expires_at = NULL,
            pro_reason = 'Free Business gift set by admin'
        WHERE id = ?
    """, (user_id,))
    conn.commit()
    conn.close()
    flash("Free Business gift activated for user.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/role/<role>")
@login_required
def admin_set_user_role(user_id, role):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    allowed = ["Admin", "Manager", "FLM", "User"]
    if role not in allowed:
        flash("Invalid role.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    conn.execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
    conn.commit()
    conn.close()

    flash(f"User role changed to {role}.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/billing/checkout/pro")
@login_required
def billing_checkout_pro():
    user = current_user()
    cfg = get_stripe_config()

    if not stripe or not cfg["secret_key"] or not cfg["pro_price_id"]:
        flash("Stripe is not configured yet. Admin must add Stripe keys and Pro Price ID.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = cfg["secret_key"]

    subscription_data = {
        "metadata": {
            "user_id": str(user["id"]),
            "plan": "pro",
        }
    }

    # rolling = normal monthly subscription from today.
    # calendar_prorated = first invoice is proportional until the first day of next month,
    # then billing renews on the 1st monthly.
    if cfg.get("billing_mode") == "calendar_prorated":
        subscription_data["billing_cycle_anchor"] = first_day_next_month_timestamp()
        subscription_data["proration_behavior"] = "create_prorations"

    session = stripe.checkout.Session.create(
        mode="subscription",
        customer_email=user["email"],
        line_items=[{"price": cfg["pro_price_id"], "quantity": 1}],
        subscription_data=subscription_data,
        success_url=f"{cfg['app_base_url']}/billing/success?session_id={{CHECKOUT_SESSION_ID}}",
        cancel_url=f"{cfg['app_base_url']}/pricing",
        metadata={
            "user_id": str(user["id"]),
            "plan": "pro",
        },
    )

    return redirect(session.url, code=303)


@app.route("/billing/success")
@login_required
def billing_success_legacy_2():
    flash("Payment completed. If your plan has not updated yet, it will update after Stripe webhook confirmation.", "success")
    return redirect(url_for("pricing"))


@app.route("/billing/portal")
@login_required
def billing_portal():
    user = current_user()
    cfg = get_stripe_config()

    if not stripe or not cfg["secret_key"]:
        flash("Stripe is not configured yet.", "error")
        return redirect(url_for("pricing"))

    if not row_get(user, "stripe_customer_id"):
        flash("No Stripe customer found for this account yet.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = cfg["secret_key"]

    portal = stripe.billing_portal.Session.create(
        customer=user["stripe_customer_id"],
        return_url=f"{cfg['app_base_url']}/pricing",
    )

    return redirect(portal.url, code=303)


@app.route("/stripe/webhook", methods=["POST"])
def stripe_webhook():
    cfg = get_stripe_config()

    if not stripe or not cfg["webhook_secret"]:
        return "Stripe webhook not configured", 400

    payload = request.get_data()
    sig_header = request.headers.get("Stripe-Signature")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, cfg["webhook_secret"])
    except Exception:
        return "Invalid webhook", 400

    event_type = event["type"]
    obj = event["data"]["object"]

    conn = get_db()

    if event_type == "checkout.session.completed":
        user_id = obj.get("metadata", {}).get("user_id")
        customer_id = obj.get("customer")
        subscription_id = obj.get("subscription")

        if user_id:
            conn.execute("""
                UPDATE users
                SET plan = 'pro',
                    stripe_customer_id = ?,
                    stripe_subscription_id = ?,
                    subscription_status = 'active'
                WHERE id = ?
            """, (customer_id, subscription_id, user_id))
            conn.commit()

    elif event_type in ["invoice.payment_succeeded", "customer.subscription.updated"]:
        subscription_id = obj.get("subscription") or obj.get("id")
        status = obj.get("status", "active")

        if subscription_id:
            plan = "pro" if status in ["active", "trialing"] else "free"
            conn.execute("""
                UPDATE users
                SET plan = ?, subscription_status = ?
                WHERE stripe_subscription_id = ?
            """, (plan, status, subscription_id))
            conn.commit()

    elif event_type in ["customer.subscription.deleted", "customer.subscription.paused"]:
        subscription_id = obj.get("id")
        if subscription_id:
            conn.execute("""
                UPDATE users
                SET plan = 'free', subscription_status = ?
                WHERE stripe_subscription_id = ?
            """, (obj.get("status", "canceled"), subscription_id))
            conn.commit()

    conn.close()
    return "ok", 200



@app.route("/admin/billing", methods=["GET", "POST"])
@login_required
def admin_billing():
    user = current_user()
    if not is_admin(user):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        fields = [
            "stripe_publishable_key",
            "stripe_secret_key",
            "stripe_webhook_secret",
            "stripe_pro_price_id",
            "app_base_url",
            "billing_mode",
            "payout_note",
        ]
        for field in fields:
            set_system_setting(field, request.form.get(field, "").strip())

        flash("Billing settings saved.", "success")
        return redirect(url_for("admin_billing"))

    cfg = get_stripe_config()
    return render_template("admin_billing.html", cfg=cfg, page="admin")


@app.route("/settings/yard", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def yard_settings():
    user = current_user()

    if request.method == "POST":
        door_start = int(request.form.get("door_start") or 1)
        door_end = int(request.form.get("door_end") or 100)
        fence_start = int(request.form.get("fence_start") or 1)
        fence_end = int(request.form.get("fence_end") or 120)

        door_start = max(1, min(door_start, 9999))
        door_end = max(door_start, min(door_end, 9999))
        fence_start = max(1, min(fence_start, 9999))
        fence_end = max(fence_start, min(fence_end, 9999))

        conn = get_db()
        conn.execute("""
            UPDATE users
            SET door_start = ?,
                door_end = ?,
                door_count = ?,
                fence_start = ?,
                fence_end = ?,
                fence_count = ?
            WHERE id = ?
        """, (
            door_start,
            door_end,
            door_end,
            fence_start,
            fence_end,
            fence_end,
            user["id"]
        ))
        conn.commit()
        conn.close()

        flash("Yard configuration saved.", "success")
        return redirect(url_for("yard_settings"))

    cfg = get_yard_config(user)
    custom_locations = get_custom_locations(user["id"])
    return render_template("yard_settings.html", cfg=cfg, custom_locations=custom_locations, page="yard_settings")




@app.route("/set-language", methods=["POST"])
def set_language():
    lang = request.form.get("language", "en")
    if lang not in TRANSLATIONS:
        lang = "en"
    session["language"] = lang
    user = current_user()
    if user:
        conn = get_db()
        conn.execute("UPDATE users SET language=? WHERE id=?", (lang, user["id"]))
        conn.commit(); conn.close()
    return redirect(request.form.get("next") or request.referrer or url_for("login"))


def user_sites(user_id):
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM sites WHERE user_id=? ORDER BY name ASC", (user_id,)).fetchall()
        if not rows:
            conn.execute("INSERT INTO sites (user_id, name, address, target_score, created_at) VALUES (?, ?, ?, ?, ?)",
                         (user_id, "Main Site", "", 90, datetime.now().isoformat()))
            conn.commit()
            rows = conn.execute("SELECT * FROM sites WHERE user_id=? ORDER BY name ASC", (user_id,)).fetchall()
        return rows
    finally:
        conn.close()

def selected_site_id(user):
    raw = request.args.get("site_id") or row_get(user, "default_site_id")
    try:
        return int(raw) if raw else None
    except Exception:
        return None

def generate_weekly_report_text(user_id, date_from=None, date_to=None):
    conn = get_db()
    if not date_to:
        date_to = datetime.today().date()
    if not date_from:
        date_from = date_to - timedelta(days=6)
    df, dt = date_from.isoformat(), date_to.isoformat()
    logs = conn.execute("SELECT * FROM daily_shift_logs WHERE user_id=? AND date BETWEEN ? AND ? ORDER BY date ASC", (user_id, df, dt)).fetchall()
    actions_open = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed'", (user_id,)).fetchone()[0]
    overdue = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed' AND due_date < ?", (user_id, datetime.today().date().isoformat())).fetchone()[0]
    conn.close()
    total_volume = sum([int(row_get(x, "volume", 0) or 0) for x in logs])
    planned_hc = sum([int(row_get(x, "planned_hc", 0) or 0) for x in logs])
    actual_hc = sum([int(row_get(x, "actual_hc", 0) or 0) for x in logs])
    late = sum([int(row_get(x, "late_trailers", 0) or 0) for x in logs])
    lines = [
        f"Weekly Manager Report ({df} to {dt})",
        "",
        f"Total volume: {total_volume}",
        f"HC vs Plan: {actual_hc}/{planned_hc}",
        f"Late trailers: {late}",
        f"Open actions: {actions_open}",
        f"Overdue actions: {overdue}",
        "",
        "Key shift notes:"
    ]
    for x in logs[-7:]:
        note = (row_get(x, "issues", "") or row_get(x, "notes", "") or "").strip()
        if note:
            lines.append(f"- {x['date']} {row_get(x,'shift','')}: {note[:180]}")
    return "\n".join(lines)

@app.route("/performance")
@login_required
@plan_required("pro")
def performance_dashboard():
    user = current_user()
    conn = get_db()
    summary = conn.execute("""
        SELECT COALESCE(SUM(volume),0) volume,
               COALESCE(SUM(planned_hc),0) planned_hc,
               COALESCE(SUM(actual_hc),0) actual_hc,
               COALESCE(SUM(late_trailers),0) late_trailers,
               COUNT(*) logs
        FROM daily_shift_logs WHERE user_id=?
    """, (user["id"],)).fetchone()
    open_actions = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed'", (user["id"],)).fetchone()[0]
    overdue = conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed' AND due_date < ?", (user["id"], datetime.today().date().isoformat())).fetchone()[0]
    conn.close()
    hc_percent = round((float(summary["actual_hc"] or 0) / float(summary["planned_hc"] or 1)) * 100, 1)
    volume_per_hc = round(float(summary["volume"] or 0) / float(summary["actual_hc"] or 1), 1)
    return render_template("performance.html", page="performance", summary=summary, hc_percent=hc_percent, volume_per_hc=volume_per_hc, open_actions=open_actions, overdue=overdue)

@app.route("/weekly-report")
@login_required
@plan_required("pro")
def weekly_report():
    user = current_user()
    body = generate_weekly_report_text(user["id"])
    return render_template("email_export.html", title="Weekly Manager Report", body=body, page="weekly_report")

@app.route("/weekly-report/pdf")
@login_required
@plan_required("pro")
def weekly_report_pdf():
    user = current_user()
    body = generate_weekly_report_text(user["id"])
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Weekly Manager Report", styles["Title"]), Spacer(1, 8)]
    for line in body.split("\n"):
        story.append(Paragraph(line or "&nbsp;", styles["Normal"]))
        story.append(Spacer(1, 4))
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="weekly-manager-report.pdf", mimetype="application/pdf")

@app.route("/ai-email", methods=["GET","POST"])
@login_required
@plan_required("pro")
def ai_email_generator():
    user=current_user()
    body=""
    if request.method=="POST":
        topic=request.form.get("topic","").strip()
        tone=request.form.get("tone","professional").strip()
        detail=request.form.get("details","").strip()
        body = f"Subject: {topic or 'Operations update'}\n\nHi team,\n\nI wanted to share a {tone} update regarding {topic or 'today’s operation'}.\n\n{detail}\n\nPlease make sure this is followed up and any further issues are raised as soon as possible.\n\nKind regards,\n{user['name']}"
    return render_template("ai_email.html", page="ai_email", body=body)

@app.route("/ai-handover", methods=["GET","POST"])
@login_required
@plan_required("pro")
def ai_handover():
    user=current_user()
    body=""
    if request.method=="POST":
        safety=request.form.get("safety","").strip()
        operations=request.form.get("operations","").strip()
        dispatch=request.form.get("dispatch","").strip()
        aob=request.form.get("aob","").strip()
        body=f"Shift Handover\n\nSafety:\n{safety or 'No safety issues reported.'}\n\nOperations:\n{operations or 'Operation completed as planned.'}\n\nDispatch:\n{dispatch or 'Dispatch update not provided.'}\n\nAOB:\n{aob or 'No further updates.'}\n\nKind regards,\n{user['name']}"
    return render_template("ai_handover.html", page="ai_handover", body=body)

@app.route("/actions/kanban")
@login_required
@plan_required("pro")
def actions_kanban():
    user=current_user()
    conn=get_db()
    rows=conn.execute("SELECT * FROM action_tracker WHERE user_id=? ORDER BY due_date ASC", (user["id"],)).fetchall()
    conn.close()
    columns={"Open":[], "In Progress":[], "Closed":[]}
    for r in rows:
        columns.setdefault(row_get(r,"status","Open") or "Open", []).append(r)
    return render_template("actions_kanban.html", page="actions", columns=columns)

@app.route("/sites", methods=["GET","POST"])
@login_required
@plan_required("business")
def sites():
    user=current_user()
    if request.method=="POST":
        conn=get_db()
        conn.execute("INSERT INTO sites (user_id,name,address,target_score,created_at) VALUES (?,?,?,?,?)",
                     (user["id"], request.form.get("name","").strip() or "New Site", request.form.get("address","").strip(), float(request.form.get("target_score") or 90), datetime.now().isoformat()))
        conn.commit(); conn.close()
        flash("Site added.", "success")
        return redirect(url_for("sites"))
    return render_template("sites.html", page="sites", rows=user_sites(user["id"]))

@app.route("/site-scorecard")
@login_required
@plan_required("business")
def site_scorecard():
    user=current_user()
    sites_rows=user_sites(user["id"])
    score_rows=[]
    conn=get_db()
    for s in sites_rows:
        logs=conn.execute("SELECT COALESCE(SUM(planned_hc),0) planned, COALESCE(SUM(actual_hc),0) actual, COALESCE(SUM(late_trailers),0) late, COALESCE(SUM(volume),0) volume FROM daily_shift_logs WHERE user_id=? AND COALESCE(site_id,0)=?", (user["id"], s["id"])).fetchone()
        actions=conn.execute("SELECT COUNT(*) FROM action_tracker WHERE user_id=? AND status!='Closed' AND COALESCE(site_id,0)=?", (user["id"], s["id"])).fetchone()[0]
        planned=float(logs["planned"] or 0); actual=float(logs["actual"] or 0); late=float(logs["late"] or 0)
        hc_score = 100 if planned == 0 else min(100, (actual/planned)*100)
        score = max(0, min(100, round(hc_score - (late*2) - actions, 1)))
        score_rows.append({"site":s, "score":score, "volume":logs["volume"], "late":late, "actions":actions})
    conn.close()
    score_rows=sorted(score_rows, key=lambda x:x["score"], reverse=True)
    return render_template("site_scorecard.html", page="site_scorecard", rows=score_rows)

@app.route("/branding", methods=["GET","POST"])
@login_required
@plan_required("business")
def branding():
    user=current_user()
    if request.method=="POST":
        logo=request.files.get("logo")
        filename=row_get(user,"business_logo_filename","")
        if logo and logo.filename:
            filename=secure_filename(f"{user['id']}_{uuid.uuid4().hex}_{logo.filename}")
            logo.save(os.path.join(UPLOAD_DIR, filename))
        conn=get_db()
        conn.execute("UPDATE users SET business_name=?, company_name=?, brand_color=?, business_logo_filename=? WHERE id=?",
                     (request.form.get("business_name","").strip(), request.form.get("company_name","").strip(), request.form.get("brand_color","#2563eb").strip(), filename, user["id"]))
        conn.commit(); conn.close()
        flash("Branding saved.", "success")
        return redirect(url_for("branding"))
    return render_template("branding.html", page="branding", row=user)

@app.route("/probation-review", methods=["GET","POST"])
@login_required
@plan_required("business")
def probation_review():
    user=current_user()
    conn=get_db()
    if request.method=="POST":
        member_id=request.form.get("team_member_id")
        member=conn.execute("SELECT * FROM team_members WHERE id=? AND user_id=?", (member_id,user["id"])).fetchone()
        member_name=row_get(member,"name", request.form.get("member_name",""))
        outcome=request.form.get("outcome","Pass")
        strengths=request.form.get("strengths","")
        concerns=request.form.get("concerns","")
        next_steps=request.form.get("next_steps","")
        generated=f"Probation Review\n\nEmployee: {member_name}\nDate: {request.form.get('review_date')}\nOutcome: {outcome}\nManager: {user['name']}\n\nStrengths:\n{strengths}\n\nConcerns:\n{concerns}\n\nNext steps:\n{next_steps}\n\nManager signature: ____________________\nEmployee signature: ____________________"
        conn.execute("INSERT INTO probation_reviews (user_id,team_member_id,member_name,review_date,outcome,manager,strengths,concerns,next_steps,generated_review,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                     (user["id"], member_id, member_name, request.form.get("review_date") or datetime.today().date().isoformat(), outcome, user["name"], strengths, concerns, next_steps, generated, datetime.now().isoformat()))
        if member:
            conn.execute("UPDATE team_members SET probation_status=? WHERE id=? AND user_id=?", (outcome, member_id, user["id"]))
        conn.commit(); conn.close()
        flash("Probation review generated.", "success")
        return redirect(url_for("probation_review"))
    members=conn.execute("SELECT * FROM team_members WHERE user_id=? ORDER BY name", (user["id"],)).fetchall()
    rows=conn.execute("SELECT * FROM probation_reviews WHERE user_id=? ORDER BY review_date DESC,id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("probation_review.html", page="probation_review", members=members, rows=rows)

@app.route("/probation-review/<int:item_id>/pdf")
@login_required
@plan_required("business")
def probation_review_pdf(item_id):
    user=current_user()
    conn=get_db()
    row=conn.execute("SELECT * FROM probation_reviews WHERE id=? AND user_id=?", (item_id,user["id"])).fetchone()
    conn.close()
    if not row:
        flash("Review not found.","error")
        return redirect(url_for("probation_review"))
    buffer=BytesIO()
    doc=SimpleDocTemplate(buffer,pagesize=A4)
    styles=getSampleStyleSheet()
    story=[Paragraph("Probation Review",styles["Title"]),Spacer(1,8)]
    for line in row["generated_review"].split("\n"):
        story.append(Paragraph(line or "&nbsp;",styles["Normal"]))
        story.append(Spacer(1,4))
    doc.build(story); buffer.seek(0)
    return send_file(buffer,as_attachment=True,download_name=f"probation-review-{row['member_name']}.pdf",mimetype="application/pdf")

@app.route("/qr-yard-check")
@login_required
@plan_required("business")
def qr_yard_check():
    user=current_user()
    base=APP_BASE_URL.rstrip("/")
    link=f"{base}{url_for('yard_check')}"
    return render_template("qr_yard_check.html", page="qr_yard_check", link=link)

@app.route("/admin/revenue")
@login_required
def admin_revenue():
    admin=current_user()
    if not is_admin(admin):
        flash("Admin access only.","error")
        return redirect(url_for("index"))
    conn=get_db()
    total=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    pro=conn.execute("SELECT COUNT(*) FROM users WHERE plan='pro'").fetchone()[0]
    business=conn.execute("SELECT COUNT(*) FROM users WHERE plan='business'").fetchone()[0]
    active=conn.execute("SELECT COUNT(*) FROM users WHERE last_login_at IS NOT NULL").fetchone()[0]
    conn.close()
    monthly=round(pro*4.99 + business*6.99, 2)
    annual=round(monthly*12, 2)
    return render_template("admin_revenue.html", page="admin_revenue", total=total, pro=pro, business=business, active=active, monthly=monthly, annual=annual)

@app.route("/more")
@login_required
def more_menu():
    user = current_user()
    return render_template("more.html", page="more", user=user)

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    user = current_user()
    if request.method == "POST":
        conn = get_db()
        conn.execute("""
            UPDATE users SET business_name = ?, company_name = ?, name = ?, role = ?, phone = ?, address = ?, mileage_rate = ?, door_count = ?, fence_count = ?, language = ?, favorite_tools = ?, company_theme = ?, brand_color = ?, show_theme_label = ? WHERE id = ?
        """, (
            request.form.get("business_name", "").strip(),
            request.form.get("company_name", "").strip(),
            request.form.get("name", "").strip(),
            request.form.get("role", "Admin").strip(),
            request.form.get("phone", "").strip(),
            request.form.get("address", "").strip(),
            float(request.form.get("mileage_rate") or 0.55),
            int(request.form.get("door_count") or 100),
            int(request.form.get("fence_count") or 120),
            request.form.get("language", "en"),
            ",".join([request.form.get(f"favorite_{i}", "") for i in range(1,5) if request.form.get(f"favorite_{i}", "")]),
            request.form.get("company_theme", "whs"),
            request.form.get("brand_color", "#f59e0b"),
            1 if request.form.get("show_theme_label") == "on" else 0,
            user["id"]
        ))
        conn.commit()
        conn.close()
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))
    return render_template("settings.html", row=user, page="settings", selected_favorites=[x[0] for x in get_favorite_tools(user)], company_theme_options=COMPANY_THEME_OPTIONS)




@app.route("/whatsapp-message", methods=["GET", "POST"])
@login_required
def whatsapp_message():
    user = current_user()
    conn = get_db()
    if request.method == "POST":
        title = request.form.get("title", "").strip() or "Start Up"
        message_date = request.form.get("message_date", "").strip()
        shift_time = request.form.get("shift_time", "").strip()
        recipient_name = request.form.get("recipient_name", "").strip()
        phone_number = request.form.get("phone_number", "").strip()
        rows_json = request.form.get("rows_json", "{}").strip() or "{}"
        generated_message = request.form.get("generated_message", "").strip()
        updated_at = datetime.now().isoformat()
        existing = conn.execute("SELECT id FROM whatsapp_drafts WHERE user_id=?", (user["id"],)).fetchone()
        if existing:
            conn.execute("""
                UPDATE whatsapp_drafts
                SET title=?, message_date=?, shift_time=?, recipient_name=?, phone_number=?, rows_json=?, generated_message=?, updated_at=?
                WHERE user_id=?
            """, (title, message_date, shift_time, recipient_name, phone_number, rows_json, generated_message, updated_at, user["id"]))
        else:
            conn.execute("""
                INSERT INTO whatsapp_drafts
                (user_id, title, message_date, shift_time, recipient_name, phone_number, rows_json, generated_message, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (user["id"], title, message_date, shift_time, recipient_name, phone_number, rows_json, generated_message, updated_at))
        conn.commit()
        conn.close()
        flash("WhatsApp message draft saved.", "success")
        return redirect(url_for("whatsapp_message"))

    draft = conn.execute("SELECT * FROM whatsapp_drafts WHERE user_id=?", (user["id"],)).fetchone()
    conn.close()
    if draft and row_get(draft, "rows_json", ""):
        rows_json = row_get(draft, "rows_json", "{}")
    else:
        rows_json = json.dumps({
            "main": [
                {"label":"Absence", "value":"1", "prefix":"x"},
                {"label":"Holiday", "value":"4", "prefix":"x"},
                {"label":"FLM", "value":"2", "prefix":"x"},
                {"label":"Clerk", "value":"2", "prefix":"x"},
                {"label":"Pick", "value":"6", "prefix":"x"},
                {"label":"Pack", "value":"1", "prefix":"x"},
                {"label":"Run Off", "value":"1", "prefix":"x"},
                {"label":"Despatch", "value":"5", "prefix":"x"},
                {"label":"SC", "value":"4", "prefix":"x"},
                {"label":"Suntory", "value":"2", "prefix":"x"}
            ],
            "notes": [
                {"label":"Slam Plan", "value":"1500"},
                {"label":"Pick/Pack", "value":"58 to pick/pack Inc EMC"},
                {"label":"Total Well", "value":"- 416"},
                {"label":"Despatch plan", "value":"18 Inc additionals"},
                {"label":"SC", "value":"1 x ND"},
                {"label":"Suntory", "value":"7.5 to invert, 3 to collect inverted"}
            ]
        })
    return render_template(
        "whatsapp_message.html",
        page="whatsapp_message",
        user=user,
        draft=draft,
        rows_json=rows_json,
        generated_message=row_get(draft, "generated_message", "") if draft else "",
        today=datetime.today().date().isoformat(),
    )

@app.route("/notifications", methods=["GET", "POST"])
@login_required
def notifications_page():
    user = current_user()
    if request.method == "POST":
        flash("Notification settings saved.", "success")
        return redirect(url_for("notifications_page"))
    return render_template("notifications.html", page="notifications", user=user, notifications=get_notifications(user["id"]))

@app.route("/manifest.json")
def manifest_json():
    manifest = {
        "name": "WHS",
        "short_name": "WHS",
        "description": "Free logistics support app for warehouse teams.",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#0f172a",
        "theme_color": "#16a34a",
        "orientation": "portrait-primary",
        "icons": [
            {"src": "/static/icons/whs-logo.svg", "sizes": "any", "type": "image/svg+xml", "purpose": "any maskable"}
        ]
    }
    return app.response_class(json.dumps(manifest), mimetype="application/manifest+json")


@app.route("/service-worker.js")
def service_worker():
    js = """
const CACHE_NAME = 'whs-ai-v25-whatsapp';
const urlsToCache = [
  '/',
  '/static/css/style.css'
];

self.addEventListener('install', event => {
  event.waitUntil(
    caches.open(CACHE_NAME).then(cache => cache.addAll(urlsToCache)).catch(() => null)
  );
});

self.addEventListener('fetch', event => {
  event.respondWith(
    fetch(event.request).catch(() => caches.match(event.request))
  );
});
"""
    return app.response_class(js, mimetype="application/javascript")



init_db()
ensure_schema_updates()
seed_admin_user()
@app.route("/api/weather")
@login_required
def api_weather():
    """Live weather for dashboard. Browser GPS first, then saved site/address, then Warrington fallback."""
    user = current_user()

    def as_float(value):
        try:
            return float(value)
        except Exception:
            return None

    lat = as_float(request.args.get("lat"))
    lon = as_float(request.args.get("lon"))
    source = "browser location"

    if lat is None or lon is None:
        query = (
            row_get(user, "address", "")
            or row_get(user, "site_name", "")
            or row_get(user, "company_name", "")
            or ""
        ).strip()
        if not query or "whs" in query.lower() or "admin" in query.lower():
            query = "Warrington, United Kingdom"

        try:
            geo = requests.get(
                "https://geocoding-api.open-meteo.com/v1/search",
                params={"name": query, "count": 1, "language": "en", "format": "json"},
                timeout=6,
            )
            result = ((geo.json() if geo.ok else {}).get("results") or [{}])[0]
            lat = result.get("latitude") or 53.3900
            lon = result.get("longitude") or -2.5969
            source = result.get("name") or "Warrington"
        except Exception:
            lat, lon, source = 53.3900, -2.5969, "Warrington"

    try:
        resp = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude": lat,
                "longitude": lon,
                "current": "temperature_2m,relative_humidity_2m,precipitation,weather_code,wind_speed_10m,pressure_msl,visibility",
                "hourly": "temperature_2m,precipitation_probability,precipitation,weather_code,wind_speed_10m,visibility",
                "daily": "weather_code,temperature_2m_max,temperature_2m_min,precipitation_probability_max,sunrise,sunset",
                "timezone": "auto",
                "forecast_days": 2,
            },
            timeout=8,
        )
        resp.raise_for_status()
        data = resp.json()
        current = data.get("current") or {}
        hourly = data.get("hourly") or {}
        daily = data.get("daily") or {}
        if not current:
            raise RuntimeError("No current weather returned")

        def get(values, i, default=None):
            try:
                value = values[i]
                return default if value is None else value
            except Exception:
                return default

        times = hourly.get("time") or []
        now_index = 0
        if current.get("time") in times:
            now_index = times.index(current.get("time"))

        hourly_items = []
        for i in range(now_index, min(now_index + 8, len(times))):
            hourly_items.append({
                "time": get(times, i),
                "temperature": get(hourly.get("temperature_2m") or [], i, 0),
                "rain_probability": get(hourly.get("precipitation_probability") or [], i, 0),
                "precipitation": get(hourly.get("precipitation") or [], i, 0),
                "weather_code": get(hourly.get("weather_code") or [], i, 0),
                "wind_speed": get(hourly.get("wind_speed_10m") or [], i, 0),
                "visibility": get(hourly.get("visibility") or [], i, 10000),
            })

        daily_items = []
        for i, day in enumerate(daily.get("time") or []):
            daily_items.append({
                "date": day,
                "weather_code": get(daily.get("weather_code") or [], i, 0),
                "max": get(daily.get("temperature_2m_max") or [], i, 0),
                "min": get(daily.get("temperature_2m_min") or [], i, 0),
                "rain_probability": get(daily.get("precipitation_probability_max") or [], i, 0),
                "sunrise": get(daily.get("sunrise") or [], i, ""),
                "sunset": get(daily.get("sunset") or [], i, ""),
            })

        rain_now = current.get("precipitation") or 0
        wind = current.get("wind_speed_10m") or 0
        visibility = current.get("visibility") or 10000
        next_rain_prob = max([(x.get("rain_probability") or 0) for x in hourly_items[:4]] or [0])
        next_rain_mm = max([(x.get("precipitation") or 0) for x in hourly_items[:4]] or [0])

        risk_points = 0
        notes = []
        if rain_now >= 1 or next_rain_mm >= 1 or next_rain_prob >= 70:
            risk_points += 2
            notes.append("Rain risk for yard work")
        elif next_rain_prob >= 40:
            risk_points += 1
            notes.append("Possible rain later")
        if wind >= 35:
            risk_points += 2
            notes.append("Strong wind")
        elif wind >= 22:
            risk_points += 1
            notes.append("Moderate wind")
        if visibility < 3000:
            risk_points += 2
            notes.append("Low visibility")
        elif visibility < 6000:
            risk_points += 1
            notes.append("Reduced visibility")
        if not notes:
            notes = ["Good conditions for yard activity"]

        risk_level = "High" if risk_points >= 4 else ("Medium" if risk_points >= 2 else "Low")

        return jsonify({
            "ok": True,
            "source": source,
            "temperature": current.get("temperature_2m"),
            "humidity": current.get("relative_humidity_2m"),
            "precipitation": current.get("precipitation"),
            "weather_code": current.get("weather_code"),
            "wind_speed": current.get("wind_speed_10m"),
            "pressure": current.get("pressure_msl"),
            "visibility": current.get("visibility"),
            "hourly": hourly_items,
            "daily": daily_items,
            "risk": {"level": risk_level, "points": risk_points, "notes": notes},
        })

    except Exception as exc:
        return jsonify({
            "ok": True,
            "source": "Warrington fallback",
            "temperature": 12,
            "humidity": 80,
            "precipitation": 0,
            "weather_code": 3,
            "wind_speed": 12,
            "pressure": 1012,
            "visibility": 10000,
            "hourly": [],
            "daily": [{
                "date": "",
                "max": 16,
                "min": 10,
                "sunrise": "2026-06-11T04:42",
                "sunset": "2026-06-11T21:37",
                "rain_probability": 0,
                "weather_code": 3,
            }],
            "risk": {"level": "Low", "points": 0, "notes": ["Fallback weather shown."]},
            "error": str(exc),
        })




@app.route("/dashboard")
@login_required
def dashboard():
    return index()



@app.route("/more")
@login_required
def more():
    return render_template("more.html", user=current_user(), page="more")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)




